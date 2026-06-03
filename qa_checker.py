"""
qa_checker.py — Deterministic / structural QA checks for TMX / MQXLIFF bilingual files.

35 categories total — 32 enabled by default. The three opt-IN checks
(declared in ``OPT_IN_CHECK_IDS``) are ``confusable_pairs``,
``custom_forbidden_regex`` and ``custom_required_regex``; everything else
ships ON. Spell-check is opt-OUT via the Configuration UI.

The 35 checks are grouped into 5 categories displayed in this order
across the Configuration expander, the in-tab cards and the HTML
report's "Sort by groups" toggle (single source of truth =
``CATEGORY_METADATA`` insertion order):

  1. Content (20) — empty_target, untranslated_segment,
     inconsistent_translation, repeated_words, length_ratio,
     polarity_mismatch, urls, emails, first_letter_case,
     alluppercase_mismatch, camelcase_mismatch, bracket_balance,
     repeated_punctuation, final_punctuation_mismatch, double_spaces,
     whitespace_edges, control_characters, mixed_scripts,
     custom_forbidden_regex*, custom_required_regex*.
  2. Numeric Elements (7) — number_mismatch, range_ratio_mismatch,
     unit_mismatch, symbol_mismatch, alphanum_id_mismatch,
     number_format_mismatch, date_format_mismatch.
  3. Tags (4) — tag_count, tag_malformed, placeholders, tag_order.
  4. Terminology (2) — forbidden_terms, glossary_violation.
  5. Spelling (2) — spellcheck, confusable_pairs*.

  (*) opt-IN — ships OFF; see ``OPT_IN_CHECK_IDS``.

The spell-check subsystem was removed in Task #28 (because it was always-on
and noisy) and reintroduced in Task #33 as a togglable category (initially
opt-in, then promoted to opt-OUT in a follow-up so it ships enabled). It uses
spylls for pure-Python Hunspell, downloads dictionaries on demand from the
LibreOffice repo (38 languages supported in total) and ships 5 of them
(en/es/fr/de/it) bundled inside the .exe so it works fully offline by default.

Public API
----------
- run_qa_checks(xml_content, filename, config) -> dict
- prepare_qa_download(xml_content, filename, *,
                      target_overrides=None) -> bytes
- export_qa_report(results, fmt, *, target_overrides=None) -> bytes
- parse_glossary(file_bytes, filename, case_sensitive=False) -> list[dict]
- parse_forbidden_terms(text) -> list[str]
- DEFAULT_CHECK_CONFIG, CATEGORY_METADATA, DEFAULT_PROFILE
"""

from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter
from functools import lru_cache
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from lxml import etree

XLIFF_NS = {
    "xliff": "urn:oasis:names:tc:xliff:document:1.2",
    "mq": "MQXliff",
}
INLINE_TAG_NAMES = {"ph", "bpt", "ept", "it", "bx", "ex", "x", "g", "mrk", "sub"}
# Character-run formatting that memoQ renders as *styled text*, not as a
# visible tag chip (see ``_is_formatting_run``).
_FORMATTING_CTYPES = {"bold", "italic", "underline"}

# -----------------------------------------------------------------------------
# Categories & severity
# -----------------------------------------------------------------------------

SEVERITY_HIGH = "high"
SEVERITY_LOW = "low"

# Task #63 — internal sub-priority WITHIN the High severity bucket. Lower
# rank = more critical. Kept for backwards compatibility only; Task #68
# replaced the HTML report's "Sort by severity" toggle with "Sort by groups"
# (driven by GROUP_RANK + CATEGORY_RANK), so this map no longer feeds any
# UI surface but is still exported for downstream consumers.
# Tiers (per pharma/clinical reviewer priorities — Task #67 reorder:
# Content & coherence now leads the body so the reviewer triages the
# linguistic-integrity bulk first, then clinical-safety, then
# terminology):
#   1 = Critical (file cannot be delivered: empty/untranslated, broken tags,
#       inconsistent translation across segments).
#   2 = Content & coherence (everything else in the Content group).
#   3 = Clinical / patient-safety data (numbers, units, dates, IDs, URLs,
#       emails, operator/currency symbols).
#   4 = Terminology (client glossary / forbidden terms).
# Categories that are Low severity are absent from this map; their JS rank
# defaults to 99 — irrelevant because severity itself already pushes them
# below every High issue.
HIGH_PRIORITY_RANK: Dict[str, int] = {
    # Tier 1 — Critical (structural / would break delivery)
    "empty_target": 1,
    "untranslated_segment": 1,
    "tag_count": 1,
    "tag_malformed": 1,
    "tag_order": 1,
    "placeholders": 1,  # Task #66 — placeholders are structural, promoted from clinical tier.
    "inconsistent_translation": 1,
    # Tier 2 — Content & coherence (Task #67 — promoted ahead of clinical/terminology
    # because the bulk of QA findings live here; reviewer wants them grouped first)
    "mixed_scripts": 2,
    "control_characters": 2,
    "bracket_balance": 2,
    "repeated_punctuation": 2,
    "length_ratio": 2,
    "repeated_words": 2,
    "first_letter_case": 2,
    "final_punctuation_mismatch": 2,
    "double_spaces": 2,
    "whitespace_edges": 2,
    "alluppercase_mismatch": 2,  # Task #66
    "camelcase_mismatch": 2,     # Task #66
    "custom_forbidden_regex": 2, # Task #61
    "custom_required_regex": 2,  # Task #61
    "polarity_mismatch": 2,      # Task #71 — negation polarity flip, content coherence
    # Tier 3 — Clinical safety / data integrity
    "number_mismatch": 3,
    "range_ratio_mismatch": 3,   # Task #71 — clinical numeric structure
    "unit_mismatch": 3,
    "date_format_mismatch": 3,
    "alphanum_id_mismatch": 3,
    "urls": 3,
    "emails": 3,
    "symbol_mismatch": 3,
    # Tier 4 — Terminology
    "forbidden_terms": 4,
    "glossary_violation": 4,
}

# Task #68 — Ordered listing of every check. The dict insertion order IS
# the display order across the whole app (Streamlit QA Configuration
# expander, in-tab category cards, HTML report sort-by-tags toggle, CSV
# rows). Groups are listed contiguously in the user-approved sequence
# Content → Numeric Elements → Tags → Terminology → Spelling, and inside
# each group the categories follow the order the reviewer agreed on. The
# `severity` field is kept for back-compat (CSV column, internal stats)
# but is no longer rendered in any UI surface (Task #68).
CATEGORY_METADATA: Dict[str, Dict[str, Any]] = {
    # ── Content ────────────────────────────────────────────────────────
    "empty_target":               {"label": "Empty target",                          "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "untranslated_segment":       {"label": "Untranslated segment",                  "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "inconsistent_translation":   {"label": "Inconsistent translation",              "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "repeated_words":             {"label": "Adjacent repeated word(s)",             "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "length_ratio":               {"label": "Length ratio out of range",             "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "polarity_mismatch":          {"label": "Polarity mismatch (negation flip)",     "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "urls":                       {"label": "URLs missing or altered",               "icon": "🔗", "group": "Content",          "severity": SEVERITY_HIGH},
    "emails":                     {"label": "Emails missing or altered",             "icon": "🔗", "group": "Content",          "severity": SEVERITY_HIGH},
    "first_letter_case":          {"label": "First-letter capitalization differs",   "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "alluppercase_mismatch":      {"label": "ALLUPPERCASE casing differs",           "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "camelcase_mismatch":         {"label": "CamelCase token missing in target",     "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "bracket_balance":            {"label": "Bracket / quote balance",               "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "repeated_punctuation":       {"label": "Repeated punctuation",                  "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "final_punctuation_mismatch": {"label": "Final punctuation differs",             "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "double_spaces":              {"label": "Double spaces",                         "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "whitespace_edges":           {"label": "Leading/trailing whitespace",           "icon": "✏️", "group": "Content",          "severity": SEVERITY_HIGH},
    "control_characters":         {"label": "Invisible / control characters",        "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    "mixed_scripts":              {"label": "Mixed scripts in same word",            "icon": "🧩", "group": "Content",          "severity": SEVERITY_HIGH},
    # User-driven custom regex (no-op until the user pastes patterns).
    "custom_forbidden_regex":     {"label": "Custom forbidden regex (target)",           "icon": "🧩", "group": "Content",      "severity": SEVERITY_HIGH},
    "custom_required_regex":      {"label": "Custom required regex (source → target)",   "icon": "🧩", "group": "Content",      "severity": SEVERITY_HIGH},
    # ── Numeric Elements ──────────────────────────────────────────────
    "number_mismatch":            {"label": "Numbers differ source / target",        "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    "range_ratio_mismatch":       {"label": "Range / ratio structure differs",       "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    "unit_mismatch":              {"label": "Unit of measurement differs",           "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    "symbol_mismatch":            {"label": "Symbol mismatch (operator / currency)", "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    "alphanum_id_mismatch":       {"label": "Alphanumeric IDs differ",               "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    "number_format_mismatch":     {"label": "Number format wrong for locale",        "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_LOW},
    "date_format_mismatch":       {"label": "Date order differs (DD/MM swap)",       "icon": "🔢", "group": "Numeric Elements", "severity": SEVERITY_HIGH},
    # ── Tags ───────────────────────────────────────────────────────────
    "tag_count":                  {"label": "Tag count mismatch",                    "icon": "🏷️", "group": "Tags",             "severity": SEVERITY_HIGH},
    "tag_malformed":              {"label": "Broken / malformed tags",               "icon": "🏷️", "group": "Tags",             "severity": SEVERITY_HIGH},
    "placeholders":               {"label": "Placeholders / variables",              "icon": "🏷️", "group": "Tags",             "severity": SEVERITY_HIGH},
    "tag_order":                  {"label": "Tag order mismatch",                    "icon": "🏷️", "group": "Tags",             "severity": SEVERITY_HIGH},
    # ── Terminology ────────────────────────────────────────────────────
    "forbidden_terms":            {"label": "Forbidden terms in target",             "icon": "📘", "group": "Terminology",      "severity": SEVERITY_HIGH},
    "glossary_violation":         {"label": "Glossary term not used",                "icon": "📘", "group": "Terminology",      "severity": SEVERITY_HIGH},
    # ── Spelling ───────────────────────────────────────────────────────
    "spellcheck":                 {"label": "Spelling (target)",                     "icon": "🔤", "group": "Spelling",         "severity": SEVERITY_LOW},
    "confusable_pairs":           {"label": "Confusable word (real-word error)",     "icon": "🔤", "group": "Spelling",         "severity": SEVERITY_LOW},
}

# Group → (background, text) highlight colors. Used by both the in-app
# QA panel (`_qa_highlight` in app.py) and the HTML report (`_highlight`
# inside `export_qa_report`) so visual language is consistent. Five
# groups, one color each — every check inherits its group's color.
# Task #66: Content shifted from pinkish red `#ffcccc/#c0392b` to a peach
# orange `#ffd9b3/#d35400` so it reads as orange (warm) and is clearly
# distinct from Spelling's pink magenta `#f4cce0/#a01060` (the previous
# Content red and the Spelling pink looked nearly identical at low
# saturation). The severity 🔴 High color `#c0392b` is independent and
# still used elsewhere for the severity badge / cell border.
GROUP_HIGHLIGHT_COLORS: Dict[str, Tuple[str, str]] = {
    "Content":            ("#ffd9b3", "#d35400"),  # peach / burnt orange
    "Numeric Elements":   ("#cce0ff", "#1f4e79"),  # blue
    "Tags":               ("#e8d4b8", "#6b4423"),  # tan / brown
    "Terminology":        ("#e8d5f2", "#6c3483"),  # purple
    "Spelling":           ("#f4cce0", "#a01060"),  # magenta / pink
}

# Task #68 — primary sort key for the HTML "Sort by groups" toggle. Order
# follows the user-approved sequence Content → Numeric Elements → Tags →
# Terminology → Spelling so the report mirrors the legend swatches and
# the Streamlit Configuration expander.
GROUP_RANK: Dict[str, int] = {
    "Content":          0,
    "Numeric Elements": 1,
    "Tags":             2,
    "Terminology":      3,
    "Spelling":         4,
}

# Task #68 — within-group sort key. Derived from CATEGORY_METADATA's
# insertion order so the in-group ordering inside the HTML "Sort by groups"
# toggle matches exactly what the reviewer sees in the Configuration
# expander. Unknown categories fall back to rank 999 (after everything).
CATEGORY_RANK: Dict[str, int] = {
    cid: i for i, cid in enumerate(CATEGORY_METADATA.keys())
}


def get_highlight_color(category_id: Optional[str]) -> Tuple[str, str]:
    """Return ``(background, text)`` highlight colors for a category.

    Falls back to the Content red palette for unknown categories so legacy
    call sites that don't pass a ``category_id`` keep their previous look.
    """
    if not category_id:
        return GROUP_HIGHLIGHT_COLORS["Content"]
    meta = CATEGORY_METADATA.get(category_id) or {}
    group = meta.get("group", "Content")
    return GROUP_HIGHLIGHT_COLORS.get(group, GROUP_HIGHLIGHT_COLORS["Content"])


# Per-category boundary lookarounds for the list/tuple branch of the QA
# highlighter. Returns (left_lookbehind, right_lookahead) regex snippets.
# A naïve global sub paints incidental substring matches — e.g. "2"
# inside "28005" for number_mismatch, "AB1" inside "AB12" for
# alphanum_id_mismatch, "http://a.com" inside "http://a.com/path" for
# urls. Each category needs a different "what counts as continuation"
# definition:
#   - spellcheck       → letter-only (current behavior, preserved so
#                        accents and digit-suffixed words behave as
#                        spell-check intends).
#   - urls / emails    → block \w plus URL chars (./:@?#~+%-) on the LEFT
#                        so a short URL doesn't match starting inside a
#                        longer one (e.g. `a.com/path` won't paint as
#                        `.com/path`). RIGHT side blocks only \w so a
#                        trailing sentence period (`…@example.es.`) or a
#                        closing bracket doesn't suppress the paint. The
#                        same-cell substring risk on the right is moot:
#                        the painted token comes from `_extract_set` on
#                        this very cell, so a short token cannot appear
#                        only as a substring of a longer one — the regex
#                        would have captured the longer one whole. Task #67
#                        regression: emails followed by `.` weren't painted.
#   - everything else  → block \w (letters + digits + underscore) so
#                        numbers and IDs don't match inside larger ones.
# This is the rule used by both `_qa_highlight` (app.py) and
# `_highlight` (inside `export_qa_report`).
_BOUNDARY_LETTER_ONLY = (r"(?<![^\W\d_])", r"(?![^\W\d_])")
_BOUNDARY_WORD = (r"(?<!\w)", r"(?!\w)")
_BOUNDARY_URL_LIKE = (r"(?<![\w./:@?#&=;~+%-])", r"(?!\w)")


def get_highlight_boundary(category_id: Optional[str]) -> Tuple[str, str]:
    """Return regex lookarounds (left, right) to wrap list-branch tokens.

    Centralized so app.py and qa_checker.py stay in sync.
    """
    if category_id == "spellcheck":
        return _BOUNDARY_LETTER_ONLY
    if category_id in ("urls", "emails"):
        return _BOUNDARY_URL_LIKE
    # Task #65 / #79 — double_spaces tokens are a run of 2+ space
    # characters (ASCII space *or* a Unicode space variant — NBSP, narrow,
    # thin, ideographic, etc.; see _DOUBLE_SPACE_CHARS). The default word
    # boundary `(?<!\w)…(?!\w)` rejects every match because the surrounding
    # letters ARE word chars, so the painted span never appeared in the
    # cell. The pattern itself (2+ consecutive spaces) is unambiguous, so
    # no boundary is needed.
    if category_id == "double_spaces":
        return ("", "")
    # Task #69 — tag/placeholder marker spans like ``[1]`` / ``[/1]``
    # are inserted inline by `_element_text_with_markers`. The default
    # word boundary `(?<!\w)…(?!\w)` reads the `[` / `]` as non-word
    # chars (fine on the inner side) but a marker glued to a letter
    # (e.g. ``word[1]word``) still highlights because both `[` and `]`
    # are non-word — actually it works. The real failure is the OUTER
    # side: with `word[1]word`, lookbehind sees `d` (word char) → fails.
    # Drop the boundary for tag/placeholder categories so the brackets
    # themselves act as the anchor.
    if category_id in ("tag_count", "tag_order", "placeholders"):
        return ("", "")
    return _BOUNDARY_WORD

ALL_CHECK_IDS = list(CATEGORY_METADATA.keys())

# Default profile = every check except the three opt-IN ones:
# * `confusable_pairs` — needs user-supplied word groups; high noise.
# * `custom_forbidden_regex` / `custom_required_regex` (Task #68) — user-
#   supplied Python regex patterns. The textareas in the QA Configuration
#   panel only appear once one of these toggles is ON, so the panel stays
#   uncluttered for users who never reach for custom regex.
# The UI layer in app.py mirrors these defaults on first render. Users can
# still disable any built-in category via the QA Configuration UI,
# including spell-check if it's too noisy for their corpus.
OPT_IN_CHECK_IDS = frozenset({
    "confusable_pairs",
    "custom_forbidden_regex",
    "custom_required_regex",
})
DEFAULT_PROFILE = set(ALL_CHECK_IDS) - OPT_IN_CHECK_IDS

DEFAULT_CHECK_CONFIG: Dict[str, Any] = {
    "enabled_checks": set(DEFAULT_PROFILE),
    "forbidden_terms": [],
    "glossary": [],
    "glossary_case_sensitive": False,
    # Task #46 — when True, glossary matching expands single-word terms via
    # Hunspell stems (re-uses spellcheck.get_dictionary cache). "paciente"
    # in the glossary then matches "pacientes" / "Pacientes" in the text.
    # Multi-word terms always fall back to literal whole-word match.
    "glossary_inflected_forms": False,
    "spellcheck_ignore": [],  # list[str] of words to skip during spell-check
    # Confusable pairs (Task #42). User-supplied groups only — no built-in
    # language packs. Loaded via the textarea or .xlsx upload in the QA tab.
    "confusable_pairs_custom": [],            # list[tuple[str, ...]]
    # length_ratio thresholds (Task #36). Tuneable without touching code.
    # Mirrors the module-level _LENGTH_RATIO_* constants defined further
    # down (kept in sync; tests assert these exact defaults: 20 / 0.3 / 3.0).
    "length_ratio_min_chars": 20,
    "length_ratio_low": 0.3,
    "length_ratio_high": 3.0,
    # Task #51 — Inconsistent translation threshold. 1.0 = exact source
    # match only; values <1.0 enable fuzzy matching via SequenceMatcher
    # on the source string (mirrors Duplicates tab). Recommended
    # range: 0.80 – 1.00.
    "inconsistent_translation_threshold": 1.00,
    # Task #55 — minimum length (in characters, after .strip()) of the
    # grouping axis for `inconsistent_translation`. Stops the check from
    # firing on short UI strings ("OK", "Cancel"…) where one-to-many
    # translations are legitimate.
    "inconsistent_translation_min_chars": 8,
    # Task #61 — User-supplied custom regex patterns. Each entry is a
    # dict produced by ``parse_custom_regex_patterns()``:
    #   {"line": int, "raw": str, "pattern": re.Pattern}
    # The case-sensitive flag is honored at parse time (it controls
    # ``re.IGNORECASE``); we keep a copy in the config so downstream
    # consumers (HTML report, tests) can introspect what the user asked.
    "custom_forbidden_regex_patterns": [],
    "custom_required_regex_patterns": [],
    "custom_regex_case_sensitive": False,
    # Per-pattern wallclock budget (seconds) enforced via the PyPI `regex`
    # timeout guard. Patterns that exceed this on a given segment are
    # skipped for that segment and counted in a summary notice.
    "custom_regex_timeout_seconds": 0.25,
}


# -----------------------------------------------------------------------------
# XML parsing helpers
# -----------------------------------------------------------------------------

def _detect_format(xml_content: bytes, filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".tmx"):
        return "tmx"
    if name.endswith(".mqxliff") or name.endswith(".xliff") or name.endswith(".xlf"):
        return "mqxliff"
    head = xml_content[:2048].lower()
    if b"<tmx" in head:
        return "tmx"
    return "mqxliff"


def _normalize_xml_input(xml_content: bytes) -> bytes:
    if xml_content[:3] == b"\xef\xbb\xbf":
        xml_content = xml_content[3:]
    elif xml_content[:2] in (b"\xff\xfe", b"\xfe\xff"):
        try:
            encoding = "utf-16-le" if xml_content[:2] == b"\xff\xfe" else "utf-16-be"
            text = xml_content[2:].decode(encoding)
            xml_content = text.encode("utf-8")
            xml_content = re.sub(rb'encoding=["\'][^"\']*["\']', rb'encoding="UTF-8"', xml_content)
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass
    xml_content = re.sub(rb"[\x00-\x08\x0b\x0c\x0e-\x1f]", b"", xml_content)
    return xml_content


# Inline-tag categories for the text and marker renderers below. XLIFF 1.2:
#   self-closing → ph, x, it, bx, ex   (skeletal placeholders)
#   paired native → g, mrk, sub        (wrap their own children)
#   paired siblings → bpt / ept        (sit next to each other, share id)
_INLINE_TAGS_SELF_CLOSING = {"ph", "x", "it", "bx", "ex"}
_INLINE_TAGS_PAIRED_WRAP = {"g", "mrk", "sub"}
# All inline tag names that should act as a word-boundary when extracting
# the visible text. Anything in this set inserts a single space between two
# adjacent non-whitespace characters when ``_element_text`` walks the tree.
_INLINE_TAG_NAMES_FOR_BOUNDARY = (
    _INLINE_TAGS_SELF_CLOSING | _INLINE_TAGS_PAIRED_WRAP | {"bpt", "ept"}
)


def _append_text_with_boundary(parts: List[str], text: str,
                               crossed_inline_tag: bool) -> None:
    """Append *text* to *parts*. When ``crossed_inline_tag`` is true and we
    would otherwise glue two non-whitespace characters together, insert a
    single space first. This prevents word-glue bugs like
    ``<source>word<ph/>other</source>`` collapsing to ``"wordother"``
    (which used to feed spellcheck / glossary / alphanum_id_mismatch /
    repeated_words and produce silent false positives on every segment
    with inline formatting).

    Crucially we only insert when BOTH adjacent characters are non-space —
    if either side already had a space, the original separator is kept and
    we do not introduce a phantom double-space (which would in turn fire
    the ``double_spaces`` check)."""
    if not text:
        return
    if crossed_inline_tag and parts:
        prev = parts[-1]
        prev_ch = prev[-1] if prev else ""
        next_ch = text[0]
        if prev_ch and next_ch and not prev_ch.isspace() and not next_ch.isspace():
            parts.append(" ")
    parts.append(text)


def _element_text(element) -> str:
    """Return the visible text of *element*, inserting a single space at
    every inline-tag boundary that would otherwise glue two word chars
    together. See :func:`_append_text_with_boundary` for the rationale.

    Boundary tracking uses a single ``pending`` flag that is raised whenever
    an inline-tag boundary is crossed and lowered when the next visible text
    is emitted. This keeps the boundary space correct even when comment /
    processing-instruction nodes (which carry no visible text of their own)
    sit between an inline tag and the following text."""
    if element is None:
        return ""
    parts: List[str] = []
    pending = {"boundary": False}

    def emit(text):
        if not text:
            return
        _append_text_with_boundary(parts, text, pending["boundary"])
        pending["boundary"] = False

    def visit(el, enter_inline: bool = False):
        if enter_inline:
            pending["boundary"] = True
        emit(el.text)
        for child in el:
            if not isinstance(child.tag, str):
                # Comment / processing-instruction node: no visible text and
                # no inline boundary of its own, but its tail must still be
                # emitted (inheriting any pending boundary from a preceding
                # inline tag).
                emit(child.tail)
                continue
            tag = etree.QName(child.tag).localname
            is_inline = tag in _INLINE_TAG_NAMES_FOR_BOUNDARY
            if tag in ("bpt", "ept") and _is_formatting_run(child, tag):
                # memoQ character-formatting run (bold/italic/underline): its
                # body is a bare ``{}`` / empty placeholder, NOT translatable
                # text. Skip the body so it can't masquerade as content — e.g.
                # a segment that is only a bold toggle wrapped around a real
                # chip would otherwise yield source_text "{} {}" and trip
                # empty_target / untranslated_segment. Keep the boundary +
                # tail so surrounding words still get a separator.
                pending["boundary"] = True
                emit(child.tail)
                continue
            # Recurse so paired wraps (g/mrk/sub) and any nested text are
            # captured.
            visit(child, enter_inline=is_inline)
            if is_inline:
                pending["boundary"] = True
            emit(child.tail)

    visit(element)
    return "".join(parts)


def _compute_marker_id(tag_name: str, tid: str,
                       name_counters: Dict[str, int]) -> str:
    """Return the marker-id portion (without brackets) for a single inline
    tag. When ``tid`` is empty, fall back to ``"<name>#<N>"`` where N is a
    per-tag-name positional counter (1-based) maintained in
    ``name_counters``.

    The per-name positional fallback (instead of a per-segment numeric
    counter) prevents false positives like ``missing [A] in target,
    extra [1] in target`` when one side has ``id="A"`` on a ``<g>`` and
    the other side has an unattributed ``<g>``: with the old logic the
    reviewer saw a phantom ``[1]`` marker that didn't correspond to
    anything in the source. With the new convention the unattributed
    ``<g>`` shows as ``[g#1]``, making it explicit that the tag is a
    nameless ``g`` (still surfaced as missing/extra when counts differ,
    but no longer disguised as a different numeric id).
    """
    if tid:
        return tid
    name_counters[tag_name] = name_counters.get(tag_name, 0) + 1
    return f"{tag_name}#{name_counters[tag_name]}"


def _element_text_with_markers(element, include_formatting_markers: bool = True) -> str:
    """Like :func:`_element_text` but inserts ``[id]`` / ``[id]…[/id]``
    placeholders where inline tags live.

    Used only by the QA-tab / HTML-report renderer so reviewers can see
    *where* tags were and *which* are missing or swapped. Every text-based
    QA check keeps consuming ``source_text`` / ``target_text`` (tag-free),
    so spell-check / glossary / number-mismatch can't false-positive on
    a ``[1]`` literal in the cleaned text.

    Marker id: the tag's ``id`` attribute (memoQ exports usually populate
    it). Missing id falls back to ``<name>#<N>`` via
    :func:`_compute_marker_id`, kept in lock-step with :func:`_tag_markers`
    so the QA cards/HTML report and the underlying check never disagree
    about which tags are present.

    ``include_formatting_markers=False`` omits the ``[id]`` marker for memoQ
    character-formatting runs (``bpt``/``ept`` flagged by
    :func:`_is_formatting_run`) while keeping their tail text. memoQ renders
    those runs as invisible *style*, not as visible chips, so a marker for
    them must NOT shield edge whitespace in :func:`_check_whitespace_edges`
    (real chips such as a tab still emit their marker and keep shielding).
    The full-marker default is what the reviewer-facing display uses.
    """
    if element is None:
        return ""
    parts: List[str] = []
    name_counters: Dict[str, int] = {}

    def _marker(child) -> str:
        tag_name = etree.QName(child.tag).localname
        tid = child.get("id") or child.get("i") or ""
        return _compute_marker_id(tag_name, tid, name_counters)

    def visit(el):
        if el.text:
            parts.append(el.text)
        for child in el:
            if not isinstance(child.tag, str):
                # Comment / processing-instruction node: emit no marker,
                # but keep its tail text.
                if child.tail:
                    parts.append(child.tail)
                continue
            tag = etree.QName(child.tag).localname
            if tag in _INLINE_TAGS_SELF_CLOSING:
                parts.append(f"[{_marker(child)}]")
            elif tag in _INLINE_TAGS_PAIRED_WRAP:
                mid = _marker(child)
                parts.append(f"[{mid}]")
                visit(child)
                parts.append(f"[/{mid}]")
            elif tag == "bpt":
                if include_formatting_markers or not _is_formatting_run(child, tag):
                    parts.append(f"[{_marker(child)}]")
            elif tag == "ept":
                if include_formatting_markers or not _is_formatting_run(child, tag):
                    parts.append(f"[/{_marker(child)}]")
            else:
                # Non-inline child element (rare in trans-unit/seg bodies):
                # recurse for its text content but emit no marker.
                visit(child)
            if child.tail:
                parts.append(child.tail)

    visit(element)
    return "".join(parts)


def _tag_markers(tags: List[Tuple[str, str, bool]]) -> List[str]:
    """Convert a ``[(name, id, is_formatting), …]`` list (as produced by
    ``_inline_tags``; the third element is tolerated but unused here)
    into the same marker strings emitted by
    :func:`_element_text_with_markers`. Missing ids fall back to
    ``<name>#<N>`` via :func:`_compute_marker_id` so the two renderers
    stay in lock-step (see that function's docstring for why).

    Tag-type-aware: ``ept`` → ``[/id]`` (closing half of an unbalanced
    pair), and ``g`` / ``mrk`` / ``sub`` emit BOTH ``[id]`` AND
    ``[/id]`` because the display renderer wraps their content. Other
    inline tags (``ph``/``x``/``it``/``bx``/``bpt``/``ex``) emit a
    single ``[id]``.
    """
    out: List[str] = []
    name_counters: Dict[str, int] = {}
    for name, tid, *_ in tags:
        mid = _compute_marker_id(name, tid, name_counters)
        if name == "ept":
            out.append(f"[/{mid}]")
        elif name in _INLINE_TAGS_PAIRED_WRAP:
            out.append(f"[{mid}]")
            out.append(f"[/{mid}]")
        else:
            out.append(f"[{mid}]")
    return out


def _is_formatting_run(child, tag: str) -> bool:
    """True for a memoQ *character-formatting* run (bold/italic/underline)
    that memoQ renders as styled text — NOT as a visible tag chip — and that
    it does not count for tag-mismatch QA.

    memoQ stores these as paired ``bpt``/``ept`` whose body is an *empty*
    bare placeholder (literally ``{}`` or nothing) carrying no native inline
    code, optionally with a formatting ``ctype`` on the opening half. The real
    content placeholders (the chips memoQ shows — ``<rpr…>`` runs, ``<mq:ch/>``
    tabs, images, breaks) always carry native code, i.e. an unescaped ``<…>``
    in their body, so they keep ``False`` here and stay in the comparison.
    ``ph`` / ``x`` / ``g`` and the other self-contained inline tags are never
    treated as formatting runs.

    The body test is deliberately **narrow**: only an empty or literal ``{}``
    body counts. A named brace placeholder such as ``{b}`` / ``{/b}`` (which
    other CAT tools and some memoQ exports use for real content tags) carries
    a token inside the braces and must keep ``False`` so genuine broken
    pairing on those tags still raises ``tag_order``.

    Why this matters: translators routinely split or merge bold/italic runs
    to fit target word order, so the *same* visible formatting yields a
    different number of ``bpt``/``ept`` halves on each side. Counting those
    produced phantom "tag count differs … unexpected [N]" errors on segments
    memoQ itself reports clean (0 QA errors)."""
    if tag not in ("bpt", "ept"):
        return False
    if (child.get("ctype") or "").lower() in _FORMATTING_CTYPES:
        return True
    # Empty or literal ``{}`` body → memoQ character-formatting placeholder.
    # Anything else (native ``<…>`` code, or a named token like ``{b}``) is a
    # real content tag and stays in the comparison.
    body = (child.text or "").strip()
    return body in ("", "{}")


def _inline_tags(element) -> List[Tuple[str, str, bool]]:
    tags: List[Tuple[str, str, bool]] = []
    if element is None:
        return tags
    for child in element.iter():
        if child is element:
            continue
        if not isinstance(child.tag, str):
            # Skip comment / processing-instruction nodes.
            continue
        tag = etree.QName(child.tag).localname
        if tag in INLINE_TAG_NAMES:
            tag_id = child.get("id") or child.get("i") or ""
            tags.append((tag, tag_id, _is_formatting_run(child, tag)))
    return tags


def _get_tmx_tuv_by_lang(tu, lang_code: str):
    if not lang_code:
        return None
    code = lang_code.lower()
    for tuv in tu.xpath("tuv"):
        tuv_lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", "") or tuv.get("lang", "")
        if tuv_lang.lower().startswith(code):
            segs = tuv.xpath("seg")
            if segs:
                return segs[0]
    return None


def _autodetect_tmx_languages(tree) -> Tuple[str, str]:
    # TMX spec: ``srclang`` lives on <header>, not on the root <tmx>. Reading
    # it from the root silently returned "" for virtually every TMX file and
    # made the function fall back to "first TUV seen = source", which flipped
    # source/target whenever the TUV order disagreed with the declared
    # language. We now read the header attribute, accepting either a direct
    # child or a namespaced one for robustness.
    header = tree.find("header")
    if header is None:
        header = tree.find(".//{*}header")
    src_attr = (header.get("srclang", "") if header is not None else "") or ""
    if not src_attr:
        # Last-resort fallback for malformed TMX where srclang accidentally
        # ended up on the root (some legacy exporters do this).
        src_attr = tree.get("srclang", "") or ""
    langs: List[str] = []
    for tuv in tree.xpath("//tu/tuv")[:50]:
        lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", "") or tuv.get("lang", "")
        if lang and lang not in langs:
            langs.append(lang)
        if len(langs) >= 2:
            break
    if src_attr and any(l.lower().startswith(src_attr.lower()) for l in langs):
        source = src_attr
        target = next((l for l in langs if not l.lower().startswith(src_attr.lower())), "")
    else:
        source = langs[0] if langs else "en"
        target = langs[1] if len(langs) > 1 else "es"
    return source, target


def _autodetect_mqxliff_languages(tree) -> Tuple[str, str]:
    files = tree.xpath("//xliff:file", namespaces=XLIFF_NS)
    if files:
        f0 = files[0]
        return f0.get("source-language", "en"), f0.get("target-language", "es")
    return "en", "es"


# Pseudo-tag scrubber. memoQ exports often emit inline placeholders as
# escaped text (e.g. ``<mq:rxt-req displaytext="{0}" val="{0}" />``) that
# arrive in source/target as literal characters, NOT as XML children. They
# are not real XLIFF inline tags (those are handled via the XML structure)
# so they pollute every text-based QA check — spell-check sees `mq`, `val`,
# `displaytext`; unit/number checks see digits inside attributes; etc.
#
# The pattern is intentionally narrow: it requires ``<`` followed by a
# letter (so ``x < 5 and y > 3`` and similar prose are never touched),
# allows tag-name chars (letters/digits/colon/dash/dot/underscore), an
# optional attribute area that disallows ``<``/``>`` inside (so a stray
# unclosed ``<`` does not consume the rest of the segment), and a closing
# ``/?>``. Matches across both opening, self-closing and closing tags.
_PSEUDO_TAG_RE = re.compile(
    r"</?[A-Za-z][A-Za-z0-9:_.\-]*(?:\s[^<>]*)?/?>"
)


def _strip_pseudo_tags(text: str) -> str:
    """Remove inline pseudo-tag fragments from a segment's visible text.

    Returns the text with each matched pseudo-tag replaced by a single
    space (so adjacent words don't accidentally fuse) and collapsed runs
    of whitespace normalized back to single spaces. ``None`` / empty
    input is returned unchanged.

    Early-return when no pseudo-tag was found: collapsing 2+ spaces
    unconditionally on every segment used to silently swallow legitimate
    double spaces (e.g. ``"Antonio es  guapo"``), neutralizing the
    ``double_spaces`` check end-to-end. We only normalize whitespace when
    we actually inserted spaces ourselves via the substitution.
    """
    if not text:
        return text
    if not _PSEUDO_TAG_RE.search(text):
        return text
    cleaned = _PSEUDO_TAG_RE.sub(" ", text)
    # Collapse the spaces we just inserted; preserve trailing/leading
    # whitespace markers that other checks (whitespace_edges) rely on by
    # only collapsing runs of >1 internal spaces.
    return re.sub(r"  +", " ", cleaned)


def _extract_segments(xml_content: bytes, filename: str) -> Tuple[str, str, str, List[Dict[str, Any]]]:
    """Return (format, source_lang, target_lang, segments)."""
    fmt = _detect_format(xml_content, filename)
    xml_norm = _normalize_xml_input(xml_content)
    parser = etree.XMLParser(remove_blank_text=False, strip_cdata=False, recover=False)
    try:
        tree = etree.fromstring(xml_norm, parser=parser)
    except etree.XMLSyntaxError as exc:
        raise ValueError(f"Could not parse {fmt.upper()} file: {exc}")

    segments: List[Dict[str, Any]] = []

    if fmt == "tmx":
        source_lang, target_lang = _autodetect_tmx_languages(tree)
        for idx, tu in enumerate(tree.xpath("//tu"), start=1):
            src_seg = _get_tmx_tuv_by_lang(tu, source_lang)
            tgt_seg = _get_tmx_tuv_by_lang(tu, target_lang)
            src_raw = _element_text(src_seg)
            tgt_raw = _element_text(tgt_seg)
            segments.append({
                "id": idx,
                # Cleaned visible text — what every QA check should see by
                # default. Pseudo-tag fragments stripped so they don't
                # contaminate spellcheck / unit / number / etc.
                "source_text": _strip_pseudo_tags(src_raw),
                "target_text": _strip_pseudo_tags(tgt_raw),
                # Raw original — kept for the small set of checks that
                # legitimately need to inspect tag-like sequences in
                # text (currently: tag_malformed).
                "source_text_raw": src_raw,
                "target_text_raw": tgt_raw,
                # Display variants with inline tags rendered as ``[id]`` /
                # ``[id]…[/id]`` markers. Surfaced in the QA tab and the
                # HTML report so reviewers can see where tags lived and
                # which ones are missing/swapped. Never fed back into any
                # text-based check.
                "source_text_display": _strip_pseudo_tags(_element_text_with_markers(src_seg)),
                "target_text_display": _strip_pseudo_tags(_element_text_with_markers(tgt_seg)),
                # Edge-only variant: real chips keep their markers (shield
                # internal whitespace) but invisible formatting runs do not,
                # so a space behind a leading/trailing format toggle still
                # fires whitespace_edges. See _check_whitespace_edges.
                "target_text_edges": _strip_pseudo_tags(
                    _element_text_with_markers(tgt_seg, include_formatting_markers=False)),
                "source_tags": _inline_tags(src_seg),
                "target_tags": _inline_tags(tgt_seg),
                "source_element": src_seg,
                "target_element": tgt_seg,
            })
    else:
        source_lang, target_lang = _autodetect_mqxliff_languages(tree)
        for idx, tu in enumerate(tree.xpath("//xliff:trans-unit", namespaces=XLIFF_NS), start=1):
            sources = tu.xpath("xliff:source", namespaces=XLIFF_NS)
            targets = tu.xpath("xliff:target", namespaces=XLIFF_NS)
            src = sources[0] if sources else None
            tgt = targets[0] if targets else None
            src_raw = _element_text(src)
            tgt_raw = _element_text(tgt)
            segments.append({
                "id": idx,
                "source_text": _strip_pseudo_tags(src_raw),
                "target_text": _strip_pseudo_tags(tgt_raw),
                "source_text_raw": src_raw,
                "target_text_raw": tgt_raw,
                "source_text_display": _strip_pseudo_tags(_element_text_with_markers(src)),
                "target_text_display": _strip_pseudo_tags(_element_text_with_markers(tgt)),
                # See the sibling block above: edge-only variant excludes
                # invisible formatting-run markers so whitespace_edges still
                # fires on a space hidden behind a leading/trailing format
                # toggle, while real chips keep shielding internal whitespace.
                "target_text_edges": _strip_pseudo_tags(
                    _element_text_with_markers(tgt, include_formatting_markers=False)),
                "source_tags": _inline_tags(src),
                "target_tags": _inline_tags(tgt),
                "source_element": src,
                "target_element": tgt,
            })

    return fmt, source_lang, target_lang, segments


# -----------------------------------------------------------------------------
# Glossary & forbidden terms parsing
# -----------------------------------------------------------------------------

def parse_forbidden_terms(text: str) -> List[str]:
    if not text:
        return []
    terms = []
    seen = set()
    for raw in text.replace("\r", "\n").split("\n"):
        term = raw.strip()
        if not term:
            continue
        key = term.lower()
        if key in seen:
            continue
        seen.add(key)
        terms.append(term)
    return terms


def _read_xlsx_rows(file_bytes: bytes) -> List[List[str]]:
    """Read first sheet of an .xlsx workbook into rows of strings.

    Uses openpyxl (already a project dependency). Legacy ``.xls`` (BIFF)
    is **not** supported — openpyxl can only read the OOXML ``.xlsx``
    format, and modern Excel exports ``.xlsx`` by default.
    Empty leading/trailing cells are preserved as ``""`` so the caller can
    decide whether the row has enough columns; row gets dropped only when
    every cell is empty. Returns ``[]`` on any parse error.
    """
    if not file_bytes:
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.active
    if ws is None:
        return []
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    return rows


def parse_glossary(file_bytes: bytes, filename: str, *, case_sensitive: bool = False) -> List[Dict[str, str]]:
    """Parse glossary file (.txt tab-separated, .csv comma-separated, or .xlsx).

    Excel path reads the first sheet, takes columns A & B, and applies the
    same optional-header detection used for TXT/CSV.
    """
    if not file_bytes:
        return []
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx")
    is_csv = name.endswith(".csv")
    rows: List[List[str]] = []
    if is_xlsx:
        rows = _read_xlsx_rows(file_bytes)
    else:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1", errors="replace")
        if is_csv:
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row:
                    rows.append([c.strip() for c in row])
        else:
            for line in text.splitlines():
                if not line.strip():
                    continue
                parts = line.split("\t") if "\t" in line else re.split(r"\s{2,}", line)
                rows.append([p.strip() for p in parts])

    entries: List[Dict[str, str]] = []
    if not rows:
        return entries

    header_keywords = {"source", "target", "src", "tgt", "term", "translation", "origen", "destino"}
    first = [c.lower() for c in rows[0][:2]]
    if any(c in header_keywords for c in first) and len(rows) > 1:
        rows = rows[1:]

    seen = set()
    for row in rows:
        if len(row) < 2:
            continue
        src, tgt = row[0].strip(), row[1].strip()
        if not src or not tgt:
            continue
        # Optional column C = reviewer note. When present it will be
        # surfaced in the QA card and HTML report under the standard
        # glossary_violation message ("Glossary mismatch: 'X' → 'Y'").
        # Empty string when the column is missing.
        note = row[2].strip() if len(row) >= 3 else ""
        key = (src, tgt) if case_sensitive else (src.lower(), tgt.lower())
        if key in seen:
            continue
        seen.add(key)
        entries.append({"source": src, "target": tgt, "note": note})
    return entries


# -----------------------------------------------------------------------------
# Per-segment checks (only deterministic / structural ones)
# -----------------------------------------------------------------------------

_URL_RE = re.compile(r"\bhttps?://[^\s<>\"']+", re.IGNORECASE)
_EMAIL_RE = re.compile(r"(?<!\w)[\w.%+\-]+@[\w.\-]+\.[^\W\d_]{2,}(?!\w)", re.UNICODE)
_PLACEHOLDER_RE = re.compile(
    r"%[\d]*\$?[sdifoxX]"          # printf-style: %s, %d, %1$s, %x
    r"|\{\{[^{}]+\}\}"             # mustache: {{name}}
    r"|\{[A-Za-z0-9_\.\-]+\}"      # python/icu: {name}
    r"|\$\{[^}]+\}"                # template: ${var}
    r"|<[A-Za-z]+/>",              # self-closing pseudo-tag: <br/>
)
# Space-like characters that count toward a "double space": ASCII space
# (U+0020) plus the Unicode space separators (category Zs) a source file or a
# memoQ export can smuggle in — non-breaking (U+00A0), narrow no-break
# (U+202F), thin (U+2009), hair, figure/punctuation/en/em spaces, ideographic
# (U+3000), etc. Two or more in a row are always redundant whitespace the
# translator usually can't see, so we flag any consecutive run regardless of
# which space variants it mixes. Tab, newline and zero-width characters are
# deliberately excluded: memoQ stores them as inline tags or renders them with
# no width at all, so flagging them would produce warnings the reviewer can
# neither see nor fix.
_DOUBLE_SPACE_CHARS = (
    "\u0020\u00a0\u1680\u2000\u2001\u2002\u2003\u2004\u2005\u2006"
    "\u2007\u2008\u2009\u200a\u202f\u205f\u3000"
)
_DOUBLE_SPACE_CHARS_SET = frozenset(_DOUBLE_SPACE_CHARS)
_DOUBLE_SPACE_RE = re.compile("[" + _DOUBLE_SPACE_CHARS + "]{2,}")


def dotify_double_space(matched: str) -> str:
    """Render a matched run of consecutive spaces as visible middle dots.

    The ``double_spaces`` token is a run of 2+ space characters (ASCII *or*
    one of the Unicode space separators in ``_DOUBLE_SPACE_CHARS``). HTML
    collapses runs of whitespace visually, so the painted span would be
    invisible — swap each space char for a middle dot ``·`` so the reviewer
    sees exactly how many (and that non-ASCII variants such as a non-breaking
    space are hiding there). Shared by the in-app QA panel (app.py) and the
    exported HTML report so both renderers stay in lock-step. Any non-space
    character (which should never appear in a double_spaces token) is passed
    through untouched.
    """
    return "".join(
        "·" if ch in _DOUBLE_SPACE_CHARS_SET else ch for ch in matched
    )


_BRACKET_PAIRS = [("(", ")"), ("[", "]"), ("{", "}")]

# Repeated punctuation: forbid runs of 2 dots (but allow exact 3-dot ellipsis "..."),
# 4+ dots, and any run of >=2 of !, ?, comma, semicolon, colon. Lookbehind
# guards against matching the tail of a longer run (e.g. last 2 dots of "...").
_REPEATED_PUNCT_RE = re.compile(
    r"(?:(?<!\.)\.{2}(?!\.)|(?<!\.)\.{4,}|(?<!!)!{2,}|(?<!\?)\?{2,}|(?<!,),{2,}|(?<!;);{2,}|(?<!:):{2,})"
)

# Invisible / control characters that survive XML parsing and are almost always
# the result of a bad copy-paste (zero-width spaces, bidi controls, BOM, DEL).
# Excludes ordinary whitespace (\t, \n, \r, regular space, NBSP U+00A0).
_INVISIBLE_CHAR_RE = re.compile(
    "["
    "\u007f"            # DEL
    "\u00ad"            # SOFT HYPHEN
    "\u200b\u200c\u200d\u200e\u200f"   # ZWSP, ZWNJ, ZWJ, LRM, RLM
    "\u202a-\u202e"     # bidi embedding / override
    "\u2060\u2061\u2062\u2063\u2064"   # word joiner & invisible math operators
    "\ufeff"            # BOM / ZWNBSP
    "]"
)
_INVISIBLE_CHAR_NAMES = {
    "\u007f": "DEL", "\u00ad": "SOFT HYPHEN",
    "\u200b": "ZERO WIDTH SPACE", "\u200c": "ZERO WIDTH NON-JOINER",
    "\u200d": "ZERO WIDTH JOINER", "\u200e": "LEFT-TO-RIGHT MARK",
    "\u200f": "RIGHT-TO-LEFT MARK",
    "\u202a": "LRE", "\u202b": "RLE", "\u202c": "PDF",
    "\u202d": "LRO", "\u202e": "RLO",
    "\u2060": "WORD JOINER",
    "\ufeff": "BOM / ZWNBSP",
}

# Mixed-scripts: Latin, Cyrillic, Greek character ranges (letters only).
_LATIN_RE = re.compile(r"[A-Za-z\u00C0-\u024F\u1E00-\u1EFF]")
_CYRILLIC_RE = re.compile(r"[\u0400-\u04FF\u0500-\u052F]")
_GREEK_RE = re.compile(r"[\u0370-\u03FF\u1F00-\u1FFF]")
# A "word" for mixed-scripts purposes: a maximal run of letters (any script).
_WORD_RE = re.compile(r"[^\W\d_]+", re.UNICODE)

# Numeric tokens: keep digit-runs with internal . , : / - so "1,234.56", "12:30",
# "2025-04-30", "1/2" stay together, are listed once each, and compared as a multiset.
_NUMBER_RE = re.compile(r"\d+(?:[.,:/\-]\d+)*")

# Sentence-final punctuation we compare. Includes CJK fullwidth variants and ellipsis.
_FINAL_PUNCT_CHARS = set(".!?…。．！？")

# Untranslated-segment threshold: skip very short segments (codes, "OK", "Email"),
# require at least one letter, ignore pure tag/whitespace differences.
_UNTRANSLATED_MIN_LEN = 8

# --- Task #36 patterns / thresholds -----------------------------------------
# Mixed alphanumeric IDs: codes that contain BOTH letters and digits and are
# typically untranslatable (subject IDs, lot numbers, sample codes, etc.).
# Detection is intentionally split in two passes (linear-time, ReDoS-safe):
#   1. ``_ALPHANUM_TOKEN_RE`` extracts every maximal run of letters / digits /
#      hyphens with no anchoring backtracking — purely linear.
#   2. ``_extract_alphanum_ids`` walks the matches and keeps only those that
#      are real ID-shaped tokens (both letters AND digits, leading/trailing
#      hyphens trimmed, single-character tokens skipped). Hyphen-only and
#      letter-only / digit-only runs are dropped here.
# This avoids the catastrophic backtracking that the previous nested
# lookahead pattern exhibited on long hyphen-heavy inputs (e.g. "A-"*25000).
# Internal separators allowed inside an alphanumeric ID: hyphen, slash, dot,
# underscore (covers SM-RA-0147, 1234/AB, A.B.C-12, lot_45_x, etc.).
_ALPHANUM_TOKEN_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9./_-]*[A-Za-z0-9]")
# Used to suppress double-reporting against number_mismatch: pure digit-runs
# (with internal . , : / -) are already covered there.
_PURE_NUMBER_RE = re.compile(r"^\d+(?:[.,:/\-]\d+)*$")


def _extract_alphanum_ids(text: str) -> List[str]:
    """Return mixed alphanumeric IDs found in *text*.

    Linear-time: a single regex pass + per-token Python validation. Tokens
    must contain at least one letter AND one digit to qualify, and must be
    at least 3 characters long after trimming separator chars. Pure words
    and pure numbers are filtered out so this check stays orthogonal to
    ``number_mismatch`` (digit-only) and the various text-only checks.
    Comparison is case-sensitive — IDs like subject codes or DNI suffixes
    are treated as untranslatable opaque strings, so case must match.
    """
    if not text:
        return []
    out: List[str] = []
    for tok in _ALPHANUM_TOKEN_RE.findall(text):
        # Strip stray edge separators (the regex anchors are alphanumerics,
        # so this is mostly defensive; cheap and keeps results clean).
        tok = tok.strip("./_-")
        if len(tok) < 3:
            continue
        # Skip pure-digit tokens — they belong to number_mismatch.
        if _PURE_NUMBER_RE.match(tok):
            continue
        has_letter = False
        has_digit = False
        for c in tok:
            if c.isalpha():
                has_letter = True
            elif c.isdigit():
                has_digit = True
            if has_letter and has_digit:
                break
        if has_letter and has_digit:
            out.append(tok)
    return out

# Adjacent repeated word(s). Case-insensitive; minimum length 2 to skip
# legitimate "I I" / "a a" sequences that may exist in noisy TM data.
_REPEATED_WORD_RE = re.compile(
    r"\b([^\W\d_]{2,})\s+\1\b", re.IGNORECASE | re.UNICODE
)

# Categories whose ``span_*`` text semantically refers to the LAST
# occurrence in the segment, not the first. The single-string highlight
# branch in both ``_qa_highlight`` (app.py) and the HTML-report
# ``_highlight`` (qa_checker.export_qa_report) consults this set so the
# correct character gets painted.
LAST_OCCURRENCE_CATEGORIES = frozenset({"final_punctuation_mismatch"})

# Length-ratio thresholds. Most language pairs land between 0.5x and 2.0x.
# Only applied above a minimum source length to avoid false positives on
# short labels / button captions.
# Default thresholds for length_ratio. Exposed via DEFAULT_CHECK_CONFIG
# (keys: length_ratio_min_chars / length_ratio_low / length_ratio_high) so
# users can tune them without code changes.
_LENGTH_RATIO_MIN_CHARS = 20
_LENGTH_RATIO_LOW = 0.3
_LENGTH_RATIO_HIGH = 3.0

# --- Task #37: units, locale-aware numbers, dates -----------------------------
# Languages whose decimal separator is the comma (and thousand separator the
# dot). Anything not in here is treated as dot-decimal. ISO-639-1 base codes
# only — region suffixes (`pt-BR`, `pt-PT`, `en-US`, ...) are stripped via
# split('-') so `pt-BR` and `pt-PT` are both treated as comma-decimal (correct
# for both European and Brazilian Portuguese).
_COMMA_DECIMAL_LANGS = {
    "es", "fr", "de", "it", "pt", "nl", "ru", "sv", "da", "no", "nb", "nn",
    "pl", "cs", "sk", "hu", "ro", "fi", "lt", "lv", "et", "bg", "hr", "sl",
    "sr", "uk", "el", "tr", "ca", "eu", "gl", "is", "sq", "mk", "be", "az",
}


def _lang_uses_comma_decimal(lang: str) -> bool:
    """True when the locale uses ``,`` as decimal and ``.`` as thousand."""
    if not lang:
        return False
    return lang.split("-")[0].split("_")[0].lower() in _COMMA_DECIMAL_LANGS


def _normalize_number_token(token: str, lang: str) -> str:
    """Return a canonical numeric-value string so ``number_mismatch`` can
    compare source and target by value, not by surface form.

    Key insight: deciding "is this dot a decimal or a thousand separator?"
    is unambiguous from the digit-suffix length in almost every real case,
    and only needs to fall back on locale for the genuinely ambiguous
    3-digit suffix (``1.000`` / ``1,000``):

    - BOTH separators present → the LAST one is the decimal (universal).
    - Single separator with a 1- or 2-digit suffix → decimal, regardless
      of locale. (This is the fix for the rejected-review bug: an EN-style
      ``5.5`` left in an ES segment must compare equal to ``5.5`` in EN
      source so ``number_format_mismatch`` (Warning) handles the locale
      issue alone — no double-report as ``number_mismatch`` (Error).)
    - Single separator with a 4+-digit suffix → decimal too (no thousand
      grouping is ever 4+ digits long).
    - Single separator with a 3-digit suffix → ambiguous; locale decides.
      Comma-decimal locale treats it as the thousand separator; dot-decimal
      locale treats it as decimal.

    Decimals always normalize to ``.`` in the output; thousand separators
    are stripped. Only separator semantics change — leading and trailing
    zeros are preserved verbatim.
    """
    if "." in token and "," in token:
        last_dot = token.rfind(".")
        last_comma = token.rfind(",")
        if last_comma > last_dot:
            return token.replace(".", "").replace(",", ".")
        return token.replace(",", "")
    if "." in token:
        suffix = token.rsplit(".", 1)[1]
        if len(suffix) == 3 and suffix.isdigit():
            # Ambiguous: could be decimal (1.000 = 1) OR thousand (1.000 =
            # 1000). Locale rules apply — comma-decimal locale → ``.`` is
            # the thousand separator (strip), dot-decimal locale → ``.`` is
            # the decimal (keep). Cross-locale rewrite false positives that
            # arise here (e.g. EN ``1,000`` vs ES ``1.000`` both meaning
            # 1000) are caught by the raw-token guard inside
            # ``_check_number_mismatch``, so genuine semantic mismatches
            # like EN decimal ``1.234`` vs target integer ``1234`` are
            # preserved.
            return token.replace(".", "") if _lang_uses_comma_decimal(lang) else token
        # 1-, 2- or 4+-digit suffix → unambiguously a decimal point.
        return token
    if "," in token:
        suffix = token.rsplit(",", 1)[1]
        if len(suffix) == 3 and suffix.isdigit():
            # Mirror of the dot branch above.
            return token.replace(",", ".") if _lang_uses_comma_decimal(lang) else token.replace(",", "")
        # 1-, 2- or 4+-digit suffix → unambiguously a decimal comma.
        return token.replace(",", ".")
    return token


def _multiset_pick_raw(raw: List[str], norm: List[str], extra: "Counter[str]") -> List[str]:
    """Given paired ``raw``/``norm`` token lists and a ``Counter`` of leftover
    normalized values from multiset subtraction, return the corresponding
    raw tokens in original order — one raw token per leftover count. Used
    by ``number_mismatch`` and ``unit_mismatch`` so painted spans honour
    duplicate-count differences (e.g. source ``mg, mg`` vs target ``mg, g``
    surfaces exactly one extra ``mg`` on the source side)."""
    if not extra:
        return []
    remaining = Counter(extra)
    picked: List[str] = []
    for r, n in zip(raw, norm):
        if remaining.get(n, 0) > 0:
            picked.append(r)
            remaining[n] -= 1
    return picked


def _violates_number_locale(token: str, lang: str) -> bool:
    """True when *token* uses a decimal separator that's wrong for *lang*.

    Conservative: only flags 1- or 2-digit fractions (``5.5`` / ``5.55`` in
    a comma-decimal locale, ``5,5`` / ``5,55`` in a dot-decimal locale).
    Tokens that look like thousand-grouped integers (3-digit suffix) are
    NEVER flagged — they're locale-correct under the opposite convention
    too. Mixed dot+comma is skipped because it's unambiguous either way.
    """
    if "." in token and "," in token:
        return False
    is_comma_locale = _lang_uses_comma_decimal(lang)
    if "." in token:
        last = token.rsplit(".", 1)[1]
        if len(last) in (1, 2) and last.isdigit():
            return is_comma_locale
        return False
    if "," in token:
        last = token.rsplit(",", 1)[1]
        if len(last) in (1, 2) and last.isdigit():
            return not is_comma_locale
        return False
    return False


# Curated SI / clinical unit lexicon. Sorted longest-first inside the regex
# alternation so ``mg/kg`` matches before ``mg``, ``mL/min`` before ``mL``,
# etc. Case-insensitive matching at the regex level (``mL`` ≡ ``ml``); the
# normalizer also folds ``μ`` (U+03BC) and ``µ`` (U+00B5) so ``μg`` ≡ ``µg``.
# Length units (mm/cm/m) are deliberately omitted — too noisy in clinical
# segments where ``5 m`` could be many things; ``mmHg`` IS in the lexicon.
_UNIT_LEXICON: List[str] = [
    # Composite rates / concentrations — must come first to win the
    # longest-match contest against their bare-unit prefixes.
    "mL/kg/h", "mL/min", "mL/h",
    "mg/kg", "mg/m2", "mcg/kg", "mcg/m2",
    "ng/mL", "pg/mL", "mg/mL", "µg/mL", "μg/mL", "ug/mL",
    "mmol/L", "nmol/L", "mEq/L", "IU/L", "U/L",
    "/min", "/h",
    # Pressure / activity / dose
    "mmHg", "mSv", "kcal", "kJ",
    # Mass
    "mcg", "µg", "μg", "ug", "mg", "ng", "pg", "kg", "g",
    # Molar
    "mmol", "nmol", "pmol", "mol",
    # Activity
    "mEq", "IU", "UI",
    # Volume
    "mL", "cL", "dL", "µL", "μL", "uL", "L",
    # Misc rate / freq
    "bpm", "Hz", "Bq", "Gy", "Sv",
    # Time
    "min", "ms", "wk", "mo", "yr", "h", "s", "d",
    # Temperature
    "°C", "°F",
    # Percent
    "%",
]

# Build the unit alternation. Sorted longest-first to ensure greedy matching.
_UNIT_ALT = "|".join(re.escape(u) for u in sorted(set(_UNIT_LEXICON), key=len, reverse=True))
# (number, unit) pair extractor. Number allows internal `.` or `,` (one),
# unit followed by anything that's NOT a letter/digit/µ/μ so we don't match
# the prefix of a longer word like "mlitros" or "gramos". Case-insensitive.
_UNIT_PAIR_RE = re.compile(
    rf"(\d+(?:[.,]\d+)?)\s*(?P<unit>{_UNIT_ALT})(?![\w\u00b5\u03bc])",
    re.IGNORECASE | re.UNICODE,
)


def _normalize_unit(unit: str) -> str:
    """Lowercase + fold both micro-sign code points so ``mL``≡``ml`` and
    ``µg``≡``μg``. Comparison stays case-insensitive within the SI ASCII
    block — that's the deliberate medical convention (capitalisation of
    `L` or `mL` is style, not semantics)."""
    return unit.replace("\u03bc", "\u00b5").lower()


# Magnitude families for the unit-mismatch hint. Values are scale factors
# in SI base units (kg / L / mol / s) so a ratio of target_scale / source_scale
# tells us by how much the displayed amount has been multiplied/divided.
_UNIT_FAMILY_SCALES: Dict[str, Dict[str, float]] = {
    "mass": {
        "pg": 1e-15, "ng": 1e-12, "µg": 1e-9, "ug": 1e-9, "mcg": 1e-9,
        "mg": 1e-6, "g": 1e-3, "kg": 1.0,
    },
    "volume": {
        "µl": 1e-9, "ul": 1e-9, "ml": 1e-6, "cl": 1e-5, "dl": 1e-4, "l": 1e-3,
    },
    "molar": {
        "pmol": 1e-12, "nmol": 1e-9, "mmol": 1e-3, "mol": 1.0,
    },
    "time": {
        "ms": 1e-3, "s": 1.0, "min": 60.0, "h": 3600.0, "d": 86400.0,
        "wk": 604800.0, "mo": 2629800.0, "yr": 31557600.0,
    },
}


def _scale_ratio(src_unit: str, tgt_unit: str) -> Optional[float]:
    """Return tgt/src scale ratio when both units share a family, else None."""
    su, tu = _normalize_unit(src_unit), _normalize_unit(tgt_unit)
    for fam in _UNIT_FAMILY_SCALES.values():
        if su in fam and tu in fam:
            try:
                return fam[tu] / fam[su]
            except ZeroDivisionError:
                return None
    return None


def _format_magnitude_hint(src_unit: str, tgt_unit: str) -> Optional[str]:
    r = _scale_ratio(src_unit, tgt_unit)
    if r is None or r == 1:
        return None
    if r > 1:
        f = r if r != int(r) else int(r)
        return f"{src_unit} → {tgt_unit} = ×{f:g}" if isinstance(f, float) else f"{src_unit} → {tgt_unit} = ×{f}"
    inv = 1.0 / r
    f = inv if inv != int(inv) else int(inv)
    return f"{src_unit} → {tgt_unit} = ÷{f:g}" if isinstance(f, float) else f"{src_unit} → {tgt_unit} = ÷{f}"


# Date pattern: DD-MM-YYYY / MM-DD-YYYY / YYYY-MM-DD with `/`, `.`, `-` as
# separator. We only flag DAY/MONTH order swaps (positions 1&2) so the
# 4-digit-year forms with year first are out of scope by design.
_DATE_RE = re.compile(r"\b(\d{1,2})([/.\-])(\d{1,2})\2(\d{2,4})\b")


def _make_issue(category_id: str, seg: Dict[str, Any], message: str,
                span_source=None, span_target=None,
                note: str = "",
                tag_sig_missing: Optional[Tuple[str, ...]] = None,
                tag_sig_extra: Optional[Tuple[str, ...]] = None) -> Dict[str, Any]:
    meta = CATEGORY_METADATA[category_id]
    return {
        "category_id": category_id,
        "severity": meta["severity"],
        "segment_id": seg["id"],
        "message": message,
        # Surface the marker-augmented text so reviewers see inline tags
        # as ``[1]`` / ``[1]…[/1]`` in every QA card and the HTML report.
        # Falls back to the cleaned text when the segment came in
        # without the display field (e.g. legacy callers / tests).
        "source": seg.get("source_text_display", seg.get("source_text", "")),
        "target": seg.get("target_text_display", seg.get("target_text", "")),
        "span_source": span_source,
        "span_target": span_target,
        # Optional reviewer note (currently used by glossary_violation
        # when the user-supplied glossary has a column C). Always a
        # string; empty when the issue type doesn't carry notes.
        "note": note or "",
        # Carry the category id on every issue so downstream renderers
        # (in-app `_qa_highlight`, HTML report `_highlight`) can look up
        # the right group color without needing the parent category dict.
        "category_id_for_color": category_id,
        # Tag-error grouping signature — populated by `tag_count`,
        # `tag_order`, `placeholders`. Two issues share a signature iff
        # their (missing, extra) sets/sequences match, and the HTML
        # report collapses same-signature rows into a single summary
        # row with a lazy "Show all" expander.
        "tag_sig_missing": tuple(tag_sig_missing) if tag_sig_missing else None,
        "tag_sig_extra": tuple(tag_sig_extra) if tag_sig_extra else None,
    }


def _check_empty_target(seg, cfg, ctx):
    src = seg["source_text"].strip()
    tgt = seg["target_text"].strip()
    if src and not tgt:
        return [_make_issue("empty_target", seg, "Source has text but target is empty.")]
    return []


@lru_cache(maxsize=20000)
def _term_pattern(term: str, case_sensitive: bool = False) -> re.Pattern:
    """Build a Unicode-aware whole-word matcher for a forbidden / glossary
    term. Word boundaries are computed via lookarounds against ``\\w`` so
    accented characters (``í``, ``ñ``, ``ü``) count as word chars and
    ``ill`` does not match inside ``pill`` / ``pillar``. Multi-word terms
    keep their internal whitespace literal.

    Memoized: the compiled pattern depends only on ``term`` and
    ``case_sensitive``, so a glossary of N terms is compiled once per run
    instead of once per (term × segment). Python's own ``re`` cache holds
    only 512 patterns, so a 3000-term glossary thrashed it and recompiled
    on nearly every call — this cache removes that entirely. Output is
    byte-for-byte identical; only the compilation work is saved."""
    flags = re.UNICODE if case_sensitive else (re.IGNORECASE | re.UNICODE)
    return re.compile(rf"(?<!\w){re.escape(term)}(?!\w)", flags)


def _check_forbidden_terms(seg, cfg, ctx):
    terms = cfg.get("forbidden_terms") or []
    if not terms:
        return []
    tgt_text = seg["target_text"] or ""
    issues = []
    for term in terms:
        if not term:
            continue
        # Whole-word, Unicode-aware: avoids "ill" inside "pillar",
        # "are" inside "share", etc. Always case-insensitive — forbidden
        # terms are user-supplied and intent is "this word, anywhere".
        if _term_pattern(term, case_sensitive=False).search(tgt_text):
            issues.append(_make_issue("forbidden_terms", seg,
                                       f"Forbidden term found: '{term}'.",
                                       span_target=term))
    return issues


# --- Task #46 helpers — Hunspell-stem expansion for glossary matching -------
# Single-word tokenizer (mirrors spellcheck._TOKEN_RE so what we feed Hunspell
# matches what spell-check feeds it). Apostrophe / hyphen allowed inside.
_GLOSSARY_TOKEN_RE = re.compile(
    r"[^\W\d_]+(?:[\u2019'\-][^\W\d_]+)*", re.UNICODE
)
# A "stem-eligible" term must look like a real word for Hunspell to have
# any hope of stemming it: letters only, optional internal apostrophe /
# hyphen between letter runs. Biomedical alphanumeric terms (``IL-6``,
# ``HLA-B27``, ``COVID-19``, ``ABC-123``) and any term carrying digits,
# slashes, dots, etc. fall back to literal whole-word matching even when
# inflected mode is ON — Hunspell would just return nothing for them and
# we'd silently miss the glossary violation.
_STEM_ELIGIBLE_RE = re.compile(
    r"^[^\W\d_]+(?:[\u2019'\-][^\W\d_]+)*$", re.UNICODE
)


def _get_lang_dict(lang: str, ctx: Dict[str, Any]) -> Optional[Any]:
    """Lazy-load + cache a spylls Dictionary for ``lang`` inside ``ctx``.

    Re-uses the on-disk / in-memory cache of :mod:`spellcheck` so a run with
    inflected-forms ON pays for the dictionary at most once per language even
    when spell-check itself is disabled. Returns None on any failure (the
    glossary check then silently falls back to literal whole-word match for
    that segment, never crashes).
    """
    if not lang:
        return None
    cache = ctx.setdefault("_lang_dicts", {})
    if lang in cache:
        return cache[lang]
    try:
        import spellcheck as _sc
        norm = _sc.normalize_lang_code(lang)
        cache[lang] = _sc.get_dictionary(norm) if norm else None
    except Exception:
        cache[lang] = None
    return cache[lang]


def _word_stems(word: str, dictionary: Optional[Any]) -> set:
    """Return the lowercased Hunspell stems for ``word``.

    Falls back to ``{word.lower()}`` when the dictionary is missing or the
    lookup yields nothing — so an unknown / mistyped word still matches
    itself literally instead of dropping out of the glossary check.
    """
    base = (word or "").lower()
    if not base:
        return set()
    if dictionary is None:
        return {base}
    try:
        forms = list(dictionary.lookuper.good_forms(word))
    except Exception:
        return {base}
    stems = {getattr(f, "stem", "").lower() for f in forms if getattr(f, "stem", None)}
    return stems or {base}


def _text_contains_glossary_term(text: str, term: str,
                                   dictionary: Optional[Any],
                                   token_stem_cache: Optional[Dict[str, set]] = None
                                   ) -> bool:
    """True iff ``text`` contains ``term`` OR any inflected form of it.

    Falls back to literal whole-word matching via :func:`_term_pattern` when:
      * ``term`` contains whitespace (multi-word — Hunspell only stems
        single words), or
      * ``term`` is not stem-eligible (contains digits, slashes, dots, or
        any non-letter glyph beyond `'`/`-` between letter runs). This
        covers biomedical alphanumeric terms like ``IL-6``, ``COVID-19``,
        ``HLA-B27``, ``ABC-123`` — Hunspell would return nothing for them
        and we'd silently miss real glossary violations.

    ``token_stem_cache`` is an optional dict (per-segment) that memoizes
    ``token -> stems`` so repeated Hunspell lookups across many glossary
    entries on the same segment are paid only once per distinct token.
    """
    if not text or not term:
        return False
    if (" " in term.strip() or "\t" in term
            or not _STEM_ELIGIBLE_RE.match(term)):
        return _term_pattern(term, case_sensitive=False).search(text) is not None
    target_stems = _word_stems(term, dictionary)
    term_lower = term.lower()
    for tok in _GLOSSARY_TOKEN_RE.findall(text):
        tl = tok.lower()
        if tl == term_lower:
            return True
        if token_stem_cache is not None:
            stems = token_stem_cache.get(tl)
            if stems is None:
                stems = _word_stems(tok, dictionary)
                token_stem_cache[tl] = stems
        else:
            stems = _word_stems(tok, dictionary)
        if stems & target_stems:
            return True
    return False


def _check_glossary_violation(seg, cfg, ctx):
    glossary = cfg.get("glossary") or []
    if not glossary:
        return []
    case_sensitive = cfg.get("glossary_case_sensitive", False)
    inflected = bool(cfg.get("glossary_inflected_forms", False))
    src_text = seg["source_text"] or ""
    tgt_text = seg["target_text"] or ""
    src_dict = _get_lang_dict((ctx or {}).get("source_lang", ""), ctx) if inflected else None
    tgt_dict = _get_lang_dict((ctx or {}).get("target_lang", ""), ctx) if inflected else None
    # Per-segment, per-side memoization of token -> stems. Without this,
    # a glossary of N entries × M tokens would call Hunspell N×M times per
    # segment; with it, each distinct token is stemmed once per segment.
    src_stem_cache: Dict[str, set] = {} if inflected else None
    tgt_stem_cache: Dict[str, set] = {} if inflected else None
    issues = []
    for entry in glossary:
        gs = entry["source"]
        gt = entry["target"]
        if inflected:
            # Stem expansion implicitly lowercases comparison; the
            # case-sensitive toggle is honored only in the literal path.
            in_src = _text_contains_glossary_term(src_text, gs, src_dict, src_stem_cache)
        else:
            # Whole-word matching prevents "cell" matching inside "cellular",
            # "cura" inside "curado", etc. Honors the glossary's case mode.
            in_src = _term_pattern(gs, case_sensitive=case_sensitive).search(src_text) is not None
        # Short-circuit: a violation can only be reported when the source
        # term is present. If it isn't, the target side is irrelevant — skip
        # the (often expensive) target match entirely. Identical output to
        # the previous "compute both, then test in_src and not in_tgt".
        if not in_src:
            continue
        if inflected:
            in_tgt = _text_contains_glossary_term(tgt_text, gt, tgt_dict, tgt_stem_cache)
        else:
            in_tgt = _term_pattern(gt, case_sensitive=case_sensitive).search(tgt_text) is not None
        if not in_tgt:
            issues.append(_make_issue("glossary_violation", seg,
                                       f"Glossary mismatch: '{gs}' → '{gt}'",
                                       span_source=gs, span_target=gt,
                                       note=entry.get("note", "")))
    return issues


def _check_double_spaces(seg, cfg, ctx):
    runs = _DOUBLE_SPACE_RE.findall(seg["target_text"])
    if runs:
        # De-dup while preserving order so each distinct run length gets
        # painted exactly once (the highlight regex inside ``_highlight``
        # will match every occurrence in the rendered cell).
        unique = list(dict.fromkeys(runs))
        return [_make_issue("double_spaces", seg,
                            "Target contains double (or more) spaces.",
                            span_target=unique)]
    return []


# Only the edge whitespace memoQ actually surfaces to the translator (as the
# floating grey dot): a regular space (U+0020) and a non-breaking space
# (U+00A0). Tabs, line breaks and exotic spaces (thin/narrow/hair/ideographic,
# zero-width, …) are deliberately excluded — they are invisible in memoQ, so
# flagging them produces warnings the reviewer can neither see nor fix.
_FLAGGABLE_EDGE_WS = (" ", "\u00a0")


def _check_whitespace_edges(seg, cfg, ctx):
    """Flag a *leading or trailing* regular space (U+0020) or non-breaking
    space (U+00A0) on the target — the only edge whitespace memoQ shows.

    The edge is judged on the marker-rendered ``target_text_edges`` (real
    inline chips shown as ``[id]`` / ``[id]…[/id]``, but memoQ
    character-formatting runs deliberately rendered WITHOUT a marker), NOT on
    the tag-free ``target_text``. memoQ renders special whitespace such as a
    tab as an inline tag (a brown "tab" chip). When such a tag sits at the
    segment edge with a real space next to it (e.g. a leading tab chip then a
    space), the tag-free reconstruction drops the chip and the space lands at
    position 0, looking like a leading edge space even though it is really
    *internal* whitespace between the tag and the first word — memoQ reports
    zero errors for it. Keeping the chip's `[id]` marker at the edge means the
    adjacent space is no longer treated as an edge space. A formatting run is
    invisible style (not a chip), so it emits no marker here and does NOT
    shield a space behind it — a space after a leading bold toggle still fires.
    A genuine stray edge space with no chip shielding it still fires. Falls
    back to ``target_text_display`` then ``target_text`` when no edge variant
    is supplied (manually built seg dicts)."""
    tgt = seg.get("target_text_edges")
    if not tgt:
        tgt = seg.get("target_text_display")
    if not tgt:
        tgt = seg.get("target_text") or ""
    if not tgt:
        return []
    side = []
    if tgt[0] in _FLAGGABLE_EDGE_WS:
        side.append("leading")
    if tgt[-1] in _FLAGGABLE_EDGE_WS:
        side.append("trailing")
    if not side:
        return []
    return [_make_issue("whitespace_edges", seg, f"Target has {' and '.join(side)} whitespace.")]


def _check_bracket_balance(seg, cfg, ctx):
    issues = []
    for opener, closer in _BRACKET_PAIRS:
        for label, text in (("source", seg["source_text"]), ("target", seg["target_text"])):
            n_open, n_close = text.count(opener), text.count(closer)
            if n_open != n_close:
                # Paint the offending side only, with whichever bracket
                # character is in excess so the user sees what's unbalanced.
                offender = opener if n_open > n_close else closer
                issues.append(_make_issue("bracket_balance", seg,
                                           f"Unbalanced '{opener}{closer}' in {label}: "
                                           f"{n_open} opening vs {n_close} closing.",
                                           span_source=offender if label == "source" else None,
                                           span_target=offender if label == "target" else None))
                break
    # Quote balance (straight double quotes only — even count expected).
    for label, text in (("source", seg["source_text"]), ("target", seg["target_text"])):
        if text.count('"') % 2 != 0:
            issues.append(_make_issue("bracket_balance", seg,
                                      f"Unbalanced double quotes in {label}.",
                                      span_source='"' if label == "source" else None,
                                      span_target='"' if label == "target" else None))
            break
    return issues


_ANNOTATION_TAG_NAMES = {"mrk", "sub"}


def _filter_content_tags(tags):
    """Strip annotation tags (``<mrk>``, ``<sub>``) before tag-count /
    tag-order comparison. memoQ auto-inserts ``<mrk>`` on the target for
    terminology highlights, comments, locked content and inline formatting
    (underline/colour) — none of which are translator-managed content
    placeholders, so they should not raise tag mismatches. Other QA tools
    (Xbench, Verifika, QA Distiller) exclude them by default for the same
    reason. ``tag_malformed`` and the renderer still see them; only the
    count/order checks ignore them.

    Also drops memoQ character-formatting runs (bold/italic/underline) flagged
    by ``_inline_tags`` (see ``_is_formatting_run``): memoQ renders those as
    styled text rather than as tag chips and lets translators split/merge them
    freely, so counting them produced phantom ``tag_count``/``tag_order``
    errors on segments memoQ reports clean."""
    out = []
    for t in tags:
        if t[0] in _ANNOTATION_TAG_NAMES:
            continue
        if len(t) > 2 and t[2]:  # formatting run → not a content chip
            continue
        out.append(t)
    return out


def _check_tag_count(seg, cfg, ctx):
    src = _filter_content_tags(seg["source_tags"])
    tgt = _filter_content_tags(seg["target_tags"])
    if len(src) == len(tgt):
        return []
    src_markers = _tag_markers(src)
    tgt_markers = _tag_markers(tgt)
    # Bag-difference (multiset) so duplicates on either side surface;
    # then dedup for highlight purposes (the highlighter paints each
    # distinct marker once, the bag-count stays in the message).
    from collections import Counter
    src_count = Counter(src_markers)
    tgt_count = Counter(tgt_markers)
    missing = list(dict.fromkeys(
        m for m in src_markers if src_count[m] > tgt_count[m]))
    extra = list(dict.fromkeys(
        m for m in tgt_markers if tgt_count[m] > src_count[m]))
    def _fmt(lst):
        return ", ".join(lst) if lst else "∅"
    if missing and extra:
        msg = (f"Tag count differs (source {len(src)}, target {len(tgt)}). "
               f"Missing in target: {_fmt(missing)} · Unexpected in target: {_fmt(extra)}.")
    elif missing:
        msg = (f"Tag count differs (source {len(src)}, target {len(tgt)}). "
               f"Missing in target: {_fmt(missing)}.")
    else:
        msg = (f"Tag count differs (source {len(src)}, target {len(tgt)}). "
               f"Unexpected in target: {_fmt(extra)}.")
    return [_make_issue("tag_count", seg, msg,
                         span_source=missing or None,
                         span_target=extra or None,
                         tag_sig_missing=tuple(missing),
                         tag_sig_extra=tuple(extra))]


def _check_tag_malformed(seg, cfg, ctx):
    """Detect leftover textual tag-like sequences in the visible text.

    Uses ``*_text_raw`` (pre-pseudo-tag-strip) because ``source_text`` /
    ``target_text`` have already had pseudo-tag fragments removed by
    ``_strip_pseudo_tags`` — running this regex against the cleaned text
    would always return zero matches and silently disable the check.
    """
    issues = []
    raw_pairs = (
        ("source", seg.get("source_text_raw", seg["source_text"])),
        ("target", seg.get("target_text_raw", seg["target_text"])),
    )
    for label, text in raw_pairs:
        if re.search(r"</?(?:ph|bpt|ept|g|x|it|bx|ex|mrk)\b[^>]*>", text or ""):
            issues.append(_make_issue("tag_malformed", seg, f"Possible broken inline tag detected in {label}."))
            break
    return issues


def _extract_set(regex: re.Pattern, text: str) -> List[str]:
    return regex.findall(text)


def _diff_spans(src_list, tgt_list):
    """Return (span_source, span_target) where each is the list of items that
    appear on one side but not the other, or ``None`` if that side has no
    extra item to paint. Returning the full list (not just the first) means
    every mismatching ID/URL/email/placeholder/number gets highlighted in
    its group color, so users don't miss a second mismatch in the same
    segment (e.g. both ``45678912-X→J`` and ``SM-RA→SM-AR`` painted)."""
    src_set, tgt_set = set(src_list), set(tgt_list)
    only_src = [t for t in src_list if t not in tgt_set]
    only_tgt = [t for t in tgt_list if t not in src_set]
    return (only_src or None, only_tgt or None)


def _check_urls(seg, cfg, ctx):
    src = sorted(_extract_set(_URL_RE, seg["source_text"]))
    tgt = sorted(_extract_set(_URL_RE, seg["target_text"]))
    if src != tgt:
        ss, st = _diff_spans(src, tgt)
        return [_make_issue("urls", seg,
                             f"URLs differ. Source: {src or '∅'} | Target: {tgt or '∅'}.",
                             span_source=ss, span_target=st)]
    return []


def _check_emails(seg, cfg, ctx):
    src = sorted(_extract_set(_EMAIL_RE, seg["source_text"]))
    tgt = sorted(_extract_set(_EMAIL_RE, seg["target_text"]))
    if src != tgt:
        ss, st = _diff_spans(src, tgt)
        return [_make_issue("emails", seg,
                             f"Emails differ. Source: {src or '∅'} | Target: {tgt or '∅'}.",
                             span_source=ss, span_target=st)]
    return []


def _check_placeholders(seg, cfg, ctx):
    src = sorted(_extract_set(_PLACEHOLDER_RE, seg["source_text"]))
    tgt = sorted(_extract_set(_PLACEHOLDER_RE, seg["target_text"]))
    if src != tgt:
        ss, st = _diff_spans(src, tgt)
        return [_make_issue("placeholders", seg,
                             f"Placeholders differ. Source: {src or '∅'} | Target: {tgt or '∅'}.",
                             span_source=ss, span_target=st,
                             tag_sig_missing=tuple(ss or ()),
                             tag_sig_extra=tuple(st or ()))]
    return []


def _check_repeated_punctuation(seg, cfg, ctx):
    tgt = seg["target_text"]
    if not tgt:
        return []
    matches = _REPEATED_PUNCT_RE.findall(tgt)
    if not matches:
        return []
    # De-duplicate while preserving order.
    seen = []
    for m in matches:
        if m not in seen:
            seen.append(m)
    sample = ", ".join(repr(s) for s in seen[:5])
    return [_make_issue("repeated_punctuation", seg,
                        f"Target contains repeated punctuation: {sample}.",
                        span_target=seen)]


def _check_control_characters(seg, cfg, ctx):
    tgt = seg["target_text"]
    if not tgt:
        return []
    found = sorted({c for c in tgt if _INVISIBLE_CHAR_RE.match(c)})
    if not found:
        return []
    names = ", ".join(_INVISIBLE_CHAR_NAMES.get(c, f"U+{ord(c):04X}") for c in found)
    return [_make_issue("control_characters", seg,
                        f"Target contains invisible/control characters: {names}.",
                        span_target=found)]


def _check_mixed_scripts(seg, cfg, ctx):
    """Detect a single word that mixes scripts that almost never co-occur.

    We deliberately ignore Latin+Greek combinations because biomedical /
    scientific terminology routinely mixes them in one token (e.g. ``TNFα``,
    ``IL2β``, ``Aβ-42``, ``IFN-γ``). The real bug pattern that this check
    targets is Latin↔Cyrillic confusion (look-alike letters from a bad PDF
    copy-paste) — flagged here together with the much rarer Greek↔Cyrillic
    mix.
    """
    tgt = seg["target_text"]
    if not tgt:
        return []
    issues = []
    seen_words = set()
    for word in _WORD_RE.findall(tgt):
        if word in seen_words:
            continue
        seen_words.add(word)
        has_latin = bool(_LATIN_RE.search(word))
        has_cyrillic = bool(_CYRILLIC_RE.search(word))
        has_greek = bool(_GREEK_RE.search(word))
        # Cyrillic is the marker of a real bug: Latin↔Cyrillic look-alike
        # confusion from a bad PDF copy-paste, or the much rarer Greek↔Cyrillic
        # mix. Plain Latin+Greek is intentionally NOT flagged because it is
        # the legitimate spelling pattern of biomedical / scientific tokens
        # (TNFα, IL2β, IFN-γ, Aβ-42, ...).
        scripts = []
        if has_latin and has_cyrillic and not has_greek:
            scripts = ["Latin", "Cyrillic"]
        elif has_greek and has_cyrillic and not has_latin:
            scripts = ["Greek", "Cyrillic"]
        elif has_latin and has_greek and has_cyrillic:
            scripts = ["Latin", "Greek", "Cyrillic"]
        if scripts:
            issues.append(_make_issue("mixed_scripts", seg,
                                       f"Word '{word}' mixes scripts: {' + '.join(scripts)}.",
                                       span_target=word))
    return issues


# --- Task #47 — Symbol mismatch -------------------------------------------
# Math / comparison / currency operators that should round-trip 1:1 between
# source and target. Deliberately excludes:
#   * ``%`` and ``°C`` / ``°F`` — already paired with their numeric value by
#     ``unit_mismatch``; including them here would double-flag.
#   * ``°`` bare — too noisy (degree-of-freedom, mathematical degree, ordinal).
#   * ``+`` / ``-`` — too overloaded (hyphens, list bullets, signed numbers
#     where the target locale legitimately drops the sign word).
# All members are stable across languages: a translator keeping the same
# meaning should keep the same operator.
_SYMBOL_LEXICON: List[str] = [
    "<", ">", "≤", "≥", "≠", "±", "×", "÷",
    "√", "∞", "‰",
    "$", "€", "£", "¥",
]
_SYMBOL_SET = set(_SYMBOL_LEXICON)


def _check_symbol_mismatch(seg, cfg, ctx):
    """Compare math / comparison / currency symbols between source and
    target as a multiset. Source ``<3`` vs target ``>3`` fires (operator
    flipped). Source ``±5%`` vs target ``+5%`` fires (missing ``±``).
    Source ``€100`` vs target ``$100`` fires (currency swapped).

    ``%`` and ``°C/°F`` are intentionally NOT in the lexicon — they're
    already paired with their numeric value by :func:`_check_unit_mismatch`,
    and re-flagging them here would double-flag the same issue.
    """
    src_text = seg["source_text"] or ""
    tgt_text = seg["target_text"] or ""
    src_syms = [c for c in src_text if c in _SYMBOL_SET]
    tgt_syms = [c for c in tgt_text if c in _SYMBOL_SET]
    if Counter(src_syms) == Counter(tgt_syms):
        return []
    src_extra = Counter(src_syms) - Counter(tgt_syms)
    tgt_extra = Counter(tgt_syms) - Counter(src_syms)
    if not src_extra and not tgt_extra:
        return []
    # Paint one offending symbol per side (the first one alphabetically so
    # the choice is deterministic across runs). Single-char span_source /
    # span_target use the string branch of the highlighter, which calls
    # ``re.sub(..., count=1)`` — only the first occurrence is painted, so
    # repeated symbols don't bleed into unrelated occurrences.
    src_paint = sorted(src_extra.elements())[0] if src_extra else None
    tgt_paint = sorted(tgt_extra.elements())[0] if tgt_extra else None
    msg = (f"Symbol mismatch. Source: {sorted(src_syms) or '∅'} | "
           f"Target: {sorted(tgt_syms) or '∅'}.")
    return [_make_issue("symbol_mismatch", seg, msg,
                        span_source=src_paint, span_target=tgt_paint)]


def _check_number_mismatch(seg, cfg, ctx):
    """Compare numeric tokens between source and target as a multiset of
    *canonical numeric values* — locale-correct decimal/thousand swaps
    (EN ``5.5`` ↔ ES ``5,5`` ↔ ES ``1.000`` ↔ EN ``1,000``) are now NOT
    flagged. Painted spans use the RAW tokens from each side so users see
    the real text in context."""
    src_lang = (ctx or {}).get("source_lang", "")
    tgt_lang = (ctx or {}).get("target_lang", "")
    src_raw = _NUMBER_RE.findall(seg["source_text"])
    tgt_raw = _NUMBER_RE.findall(seg["target_text"])
    src_norm = [_normalize_number_token(t, src_lang) for t in src_raw]
    tgt_norm = [_normalize_number_token(t, tgt_lang) for t in tgt_raw]
    if sorted(src_norm) == sorted(tgt_norm):
        return []
    # Raw-token guard: when the surface tokens themselves are identical
    # multisets, never flag number_mismatch — the user kept the source
    # numbers verbatim, and any locale-format issue is the responsibility
    # of number_format_mismatch (Warning) alone, not number_mismatch
    # (Error). This prevents the cross-locale false positive where the
    # same string ``1,000`` carries different canonical values under two
    # locale rules.
    if sorted(src_raw) == sorted(tgt_raw):
        return []
    # Multiset (Counter) subtraction so duplicate-count differences are
    # surfaced — e.g. source ``1 1 2`` vs target ``1 2 2`` paints the
    # extra ``1`` on the source side and the extra ``2`` on the target.
    src_extra = Counter(src_norm) - Counter(tgt_norm)
    tgt_extra = Counter(tgt_norm) - Counter(src_norm)
    only_src = _multiset_pick_raw(src_raw, src_norm, src_extra)
    only_tgt = _multiset_pick_raw(tgt_raw, tgt_norm, tgt_extra)
    return [_make_issue("number_mismatch", seg,
                         f"Numbers differ. Source: {src_raw or '∅'} | Target: {tgt_raw or '∅'}.",
                         span_source=only_src or None,
                         span_target=only_tgt or None)]


def _check_unit_mismatch(seg, cfg, ctx):
    """Compare (number, unit) pairs as a multiset of normalized units.

    Units are case-insensitive and the two micro symbols (U+00B5 / U+03BC)
    are folded together. When source and target both contain unit-bearing
    numbers but the unit multisets differ, fire one Error per segment with
    every mismatching unit highlighted on each side. When the two unequal
    units belong to the same magnitude family (mass / volume / molar /
    time), append a magnitude hint (e.g. ``mg → g = ×1000``) so reviewers
    immediately see the dosing impact.
    """
    src_text = seg["source_text"]
    tgt_text = seg["target_text"]
    src_pairs = _UNIT_PAIR_RE.findall(src_text)
    tgt_pairs = _UNIT_PAIR_RE.findall(tgt_text)
    if not src_pairs and not tgt_pairs:
        return []
    src_units_raw = [u for _, u in src_pairs]
    tgt_units_raw = [u for _, u in tgt_pairs]
    src_units_norm = [_normalize_unit(u) for u in src_units_raw]
    tgt_units_norm = [_normalize_unit(u) for u in tgt_units_raw]
    if sorted(src_units_norm) == sorted(tgt_units_norm):
        return []
    # Multiset subtraction preserves count — e.g. source ``5 mg and 10 mg``
    # vs target ``5 mg and 10 g`` correctly surfaces ONE extra ``mg`` on
    # the source side and ONE extra ``g`` on the target side, instead of
    # losing the count to set semantics.
    src_extra = Counter(src_units_norm) - Counter(tgt_units_norm)
    tgt_extra = Counter(tgt_units_norm) - Counter(src_units_norm)
    src_paint = _multiset_pick_raw(src_units_raw, src_units_norm, src_extra) or None
    tgt_paint = _multiset_pick_raw(tgt_units_raw, tgt_units_norm, tgt_extra) or None
    hints: List[str] = []
    for sn in sorted(src_extra.elements()):
        for tn in sorted(tgt_extra.elements()):
            h = _format_magnitude_hint(sn, tn)
            if h and h not in hints:
                hints.append(h)
    msg = (f"Unit mismatch. Source: {sorted(src_units_norm) or '∅'} | "
           f"Target: {sorted(tgt_units_norm) or '∅'}.")
    if hints:
        msg += f" Magnitude change: {', '.join(hints[:3])}."
    return [_make_issue("unit_mismatch", seg, msg,
                        span_source=src_paint, span_target=tgt_paint)]


def _check_number_format_mismatch(seg, cfg, ctx):
    """Warn when a target numeric token uses a decimal separator that's wrong
    for the target locale — e.g. ``5.5`` in an ES segment, or ``5,5`` in an
    EN segment. Conservative: only 1- or 2-digit fractions trigger; integer
    thousand groupings (``1,000`` / ``1.000``) are left alone because they
    are locale-correct under both conventions."""
    tgt_lang = (ctx or {}).get("target_lang", "")
    if not tgt_lang:
        return []
    bad = [t for t in _NUMBER_RE.findall(seg["target_text"])
           if _violates_number_locale(t, tgt_lang)]
    if not bad:
        return []
    expected = "," if _lang_uses_comma_decimal(tgt_lang) else "."
    deduped = list(dict.fromkeys(bad))
    sample = ", ".join(repr(t) for t in deduped[:5])
    return [_make_issue("number_format_mismatch", seg,
                        f"Target uses wrong decimal separator for locale "
                        f"'{tgt_lang}' (expected '{expected}'): {sample}.",
                        span_target=deduped)]


def _check_date_format_mismatch(seg, cfg, ctx):
    """Detect day/month order swaps (positions 1 and 2) when source and
    target contain the same number of dates *and* at least one day or month
    component is ≥13, which makes the swap unambiguous. Fully ambiguous
    cases (both ≤12) are silently skipped to avoid noisy false positives.
    Year (3rd component) must match exactly to be considered a swap."""
    src_matches = list(_DATE_RE.finditer(seg["source_text"]))
    tgt_matches = list(_DATE_RE.finditer(seg["target_text"]))
    if not src_matches or len(src_matches) != len(tgt_matches):
        return []
    # ctx language is read but not currently used for firing — the rule
    # below intentionally stays language-agnostic. See the docstring:
    # cross-language MDY↔DMY swaps with both components ≤12 are
    # legitimate translations and must not be flagged.
    _ = (ctx or {}).get("source_lang", ""), (ctx or {}).get("target_lang", "")
    src_paint: List[str] = []
    tgt_paint: List[str] = []
    for sm, tm in zip(src_matches, tgt_matches):
        sa, _sep, sb, sy = sm.group(1), sm.group(2), sm.group(3), sm.group(4)
        ta, _sep2, tb, ty = tm.group(1), tm.group(2), tm.group(3), tm.group(4)
        if sy != ty:
            continue
        if (sa, sb) == (ta, tb):
            continue
        try:
            ia, ib, ja, jb = int(sa), int(sb), int(ta), int(tb)
        except ValueError:
            continue
        # Positions 1 and 2 swapped, AND at least one component reveals
        # which is the day (≥13) — keeps the check unambiguous.
        if ia == jb and ib == ja and max(ia, ib) >= 13:
            src_paint.append(sm.group(0))
            tgt_paint.append(tm.group(0))
    if not src_paint:
        return []
    return [_make_issue("date_format_mismatch", seg,
                        f"Possible day/month swap: {src_paint} → {tgt_paint}.",
                        span_source=src_paint, span_target=tgt_paint)]


# --- Task #71: range/ratio structure mismatch -----------------------------
# A "structured numeric pattern" expresses a relationship between two
# numbers — a range (1-3, 1 to 3, 1–3, 1 a 3, 1 bis 3) or a ratio
# (1:3, 1/3, 1 out of 3, 1 von 3, 1 de cada 3, 1 sur 3, 1 su 3). The
# check is language-agnostic by design: it only compares the *kind* of
# structure and the canonical number pair on each side. If the source
# expresses a range and the target lost that structure (e.g. "1-3" →
# "1 und 3" / "1 and 3" / "1 y 3"), the check fires. Conjunctions are
# not recognised as ranges — that is exactly the failure mode we want
# to surface. The check is conservative: it skips segments where neither
# side has any structured pattern.

# Range with a Unicode dash between two digit runs. Bounded by
# non-letter / non-digit on both sides so it doesn't bite into product
# codes ("ABC-12-34"), dates (handled separately below), or
# CamelCase-with-digits identifiers.
_RANGE_DASH_RE = re.compile(
    r"(?<![\w./])(\d{1,4})\s*[\-\u2013\u2014]\s*(\d{1,4})(?![\w./])"
)
# Colon / slash ratio. Same boundary logic — avoids matching inside
# 12:30 (time), 2025/04 (date partial), HH:MM:SS chains, etc. We rule
# out times with a small post-check.
_COLON_SLASH_RATIO_RE = re.compile(
    r"(?<![\w./:])(\d{1,4})\s*([:/])\s*(\d{1,4})(?![\w./:])"
)
# Word ratio: "<num> <connector> <num>" — high-confidence connectors
# only, per language. The Spanish "de" alone is intentionally NOT in
# the list (too noisy: "página 3 de 5" → page numbering, not a ratio).
# "de cada" is included instead.
_WORD_RATIO_RE = re.compile(
    r"(?<!\d)(\d{1,4})\s+"
    r"(?:out\s+of|in\s+(?:every|each)|von|aus|de\s+cada|sur|su|op\s+de|na|"
    r"z|w|out|over|fuori\s+da|sobre)"
    r"\s+(\d{1,4})(?!\d)",
    re.IGNORECASE,
)


def _extract_range_ratio_patterns(text: str) -> List[Tuple[str, Tuple[int, int]]]:
    """Return ``[(kind, (a, b)), ...]`` for every structured numeric
    pattern found in ``text``. ``kind`` is ``"range"`` or ``"ratio"``;
    numbers are returned as ints in surface order (not sorted) so the
    direction is preserved (1-3 ≠ 3-1).

    Date spans are masked out beforehand so a date like ``15/03/2025``
    does not generate spurious ``15/3`` / ``3/2025`` ratios.
    """
    if not text or not any(c.isdigit() for c in text):
        return []
    # Mask date spans with spaces so the date components disappear from
    # the ratio/range scan without shifting downstream offsets.
    masked = text
    for m in _DATE_RE.finditer(text):
        s, e = m.start(), m.end()
        masked = masked[:s] + (" " * (e - s)) + masked[e:]
    out: List[Tuple[str, Tuple[int, int]]] = []
    for m in _RANGE_DASH_RE.finditer(masked):
        try:
            out.append(("range", (int(m.group(1)), int(m.group(2)))))
        except (TypeError, ValueError):
            continue
    for m in _COLON_SLASH_RATIO_RE.finditer(masked):
        sep = m.group(2)
        # Skip clock-style times when the separator is ":" and both
        # sides look like HH:MM components — i.e. both 2-digit zero-padded
        # AND in range. Loose forms like "1:3" stay as ratios; "1:30"
        # is treated as a ratio (you'd never write a time that way in
        # clinical copy, and false positives there were too costly).
        if sep == ":":
            ga, gb = m.group(1), m.group(3)
            try:
                a, b = int(ga), int(gb)
            except (TypeError, ValueError):
                continue
            if (len(ga) == 2 and len(gb) == 2
                    and 0 <= a <= 23 and 0 <= b <= 59):
                continue
        try:
            out.append(("ratio", (int(m.group(1)), int(m.group(3)))))
        except (TypeError, ValueError):
            continue
    for m in _WORD_RATIO_RE.finditer(masked):
        try:
            out.append(("ratio", (int(m.group(1)), int(m.group(2)))))
        except (TypeError, ValueError):
            continue
    return out


def _check_range_ratio_mismatch(seg, cfg, ctx):
    """Flag when source and target disagree on the kind of structured
    numeric pattern (range vs ratio) or on the number pair inside it.

    Mismatch examples that fire:
      * source ``1-3`` (range) vs target ``1 und 3`` (no structure).
      * source ``1 out of 3`` (ratio) vs target ``1:2`` (ratio with
        different numbers).
      * source ``2-5`` vs target ``5-2`` (reversed pair).

    Skipped (no fire):
      * Neither side has any structured pattern.
      * Same structure kind and same canonical pair on both sides
        (e.g. source ``1:3`` ↔ target ``1 von 3``).
    """
    src_text = seg.get("source_text") or ""
    tgt_text = seg.get("target_text") or ""
    src_pats = _extract_range_ratio_patterns(src_text)
    tgt_pats = _extract_range_ratio_patterns(tgt_text)
    if not src_pats and not tgt_pats:
        return []
    # Canonical multiset comparison. Ranges keep the order (1-3 ≠ 3-1)
    # because direction is semantically meaningful; ratios likewise
    # (1:3 ≠ 3:1 — "1 in 3" vs "3 in 1").
    src_key = sorted([(k, p) for k, p in src_pats])
    tgt_key = sorted([(k, p) for k, p in tgt_pats])
    if src_key == tgt_key:
        return []
    src_paint = [f"{a}{'-' if k == 'range' else ':'}{b}"
                 for k, (a, b) in src_pats] or None
    tgt_paint = [f"{a}{'-' if k == 'range' else ':'}{b}"
                 for k, (a, b) in tgt_pats] or None
    return [_make_issue(
        "range_ratio_mismatch", seg,
        f"Range / ratio structure differs. Source: {src_paint or '∅'} | "
        f"Target: {tgt_paint or '∅'}.",
        span_source=src_paint, span_target=tgt_paint,
    )]


# --- Task #71: polarity (negation flip) ------------------------------------
# Per-language negation-token sets used by ``polarity_mismatch``. Lists are
# small, stable and curated — they cover the negation particles that
# materially affect clinical meaning in the 8 European pharma languages
# the team works with. Documents in a language NOT in this map are
# silently skipped for this check and one notice is surfaced per run
# listing the skipped languages.
#
# French note: ``ne`` is intentionally absent so the standard French
# ``ne … pas`` construction is counted ONCE (via ``pas``), not twice.
# English note: contractions (n't endings) are listed explicitly so
# tokens like "don't", "isn't", "can't" match without needing apostrophe
# tokenization gymnastics.
_POLARITY_NEGATIONS: Dict[str, frozenset] = {
    "en": frozenset({
        "not", "no", "never", "none", "nothing", "nobody", "neither",
        "nor", "without", "cannot",
        "isn't", "aren't", "wasn't", "weren't", "don't", "doesn't",
        "didn't", "won't", "wouldn't", "shouldn't", "can't",
        "couldn't", "hasn't", "haven't", "hadn't", "mustn't",
        "mightn't", "needn't", "shan't",
    }),
    "de": frozenset({
        "nicht", "kein", "keine", "keinen", "keiner", "keinem",
        "keines", "niemals", "nie", "niemand", "weder", "ohne",
        "nichts", "nirgends", "nirgendwo",
    }),
    "es": frozenset({
        "no", "nunca", "ningún", "ninguna", "ningunos", "ningunas",
        "ninguno", "nadie", "nada", "ni", "sin", "jamás", "tampoco",
    }),
    "fr": frozenset({
        # "ne" omitted on purpose — see module note above.
        "pas", "non", "jamais", "aucun", "aucune", "rien", "personne",
        "sans", "ni", "nul", "nulle", "nullement", "guère", "plus",
    }),
    "it": frozenset({
        "non", "no", "mai", "nessuno", "nessuna", "niente", "nulla",
        "né", "senza", "neanche", "neppure", "nemmeno",
    }),
    "pt": frozenset({
        "não", "nao", "nunca", "nenhum", "nenhuma", "nenhuns",
        "nenhumas", "ninguém", "ninguem", "nada", "nem", "sem",
        "jamais", "tampouco",
    }),
    "nl": frozenset({
        "niet", "geen", "nooit", "niets", "niemand", "noch", "zonder",
        "nergens",
    }),
    "pl": frozenset({
        "nie", "nigdy", "nikt", "nic", "ani", "bez", "żaden", "żadna",
        "żadne", "żadnego", "żadnej",
    }),
}

# Tokeniser that keeps apostrophes (straight + curly) and hyphens inside
# a token so English contractions like "don't" / "isn't" match the
# negation set verbatim, and Polish/German compounds stay together.
_POLARITY_TOKEN_RE = re.compile(
    r"[^\W\d_]+(?:[\u2019'\-][^\W\d_]+)*", re.UNICODE,
)


def _normalize_polarity_lang(lang: str) -> Optional[str]:
    """Return the 2-letter language head if it has a negation list,
    else ``None``. ``pt-BR`` → ``pt``; ``en_US`` → ``en``; ``ja`` →
    ``None`` (no list).
    """
    if not lang:
        return None
    head = lang.split("-")[0].split("_")[0].lower()
    return head if head in _POLARITY_NEGATIONS else None


def _count_negations(text: str, lang_head: str) -> int:
    """Whole-word, case-insensitive count of negation tokens in *text*
    using the per-language set in :data:`_POLARITY_NEGATIONS`.

    Apostrophe normalisation: curly U+2019, U+2018 and U+02BC are
    folded to ASCII ``'`` before the set lookup so contractions written
    with typographic apostrophes (default in Word / InDesign) still
    match — e.g. ``isn’t`` is counted just like ``isn't``.
    """
    if not text:
        return 0
    neg = _POLARITY_NEGATIONS.get(lang_head)
    if not neg:
        return 0
    count = 0
    for tok in _POLARITY_TOKEN_RE.findall(text):
        key = tok.lower().translate(_APOSTROPHE_FOLD)
        if key in neg:
            count += 1
    return count


# Folded-apostrophe table for polarity contraction matching (Task #71
# architect review). All three typographic apostrophe variants map to
# the ASCII ``'`` that the negation sets use verbatim.
_APOSTROPHE_FOLD = str.maketrans({"\u2019": "'", "\u2018": "'", "\u02bc": "'"})


def _check_polarity_mismatch(seg, cfg, ctx):
    """Flag a polarity *flip*: one side carries negation token(s) and
    the other carries none — a strong heuristic for accidental
    affirmation↔negation swaps.

    This is a presence/absence test, NOT a count comparison. Romance
    and Slavic negative concord (Spanish "no … ningún … ni …", etc.)
    expresses ONE logical negation with several negative words, so raw
    counts legitimately differ for the same meaning; comparing counts
    produced massive false positives. Presence vs absence is the only
    count-based signal that reliably points at a real flip without
    semantic analysis.

    Both source and target languages must be in
    :data:`_POLARITY_NEGATIONS`; otherwise the segment is silently
    skipped and the unsupported language is recorded in ``ctx`` for a
    single end-of-run notice. The check does NOT claim to detect
    *which* negation flipped; reviewers inspect the segment manually.
    """
    src_raw_lang = (ctx or {}).get("source_lang", "") or ""
    tgt_raw_lang = (ctx or {}).get("target_lang", "") or ""
    src_lang = _normalize_polarity_lang(src_raw_lang)
    tgt_lang = _normalize_polarity_lang(tgt_raw_lang)
    if src_lang is None or tgt_lang is None:
        # Stash the unsupported language(s) so run_qa_checks can surface
        # one notice listing them all at the end of the run.
        unsupported = ctx.setdefault("_polarity_unsupported_langs", set())
        if src_lang is None and src_raw_lang:
            unsupported.add(src_raw_lang)
        if tgt_lang is None and tgt_raw_lang:
            unsupported.add(tgt_raw_lang)
        return []
    src_count = _count_negations(seg.get("source_text") or "", src_lang)
    tgt_count = _count_negations(seg.get("target_text") or "", tgt_lang)
    # Only flag a polarity *flip*: one side carries negation(s) and the
    # other carries none. Romance/Slavic negative concord (Spanish
    # "no … ningún … ni …", etc.) makes raw counts legitimately differ
    # (1 vs 2, 2 vs 8) for the SAME meaning, so comparing counts produced
    # massive false positives. A presence/absence mismatch is the only
    # count-based signal that reliably points at a real
    # affirmation↔negation error without semantic analysis.
    if (src_count == 0) == (tgt_count == 0):
        return []
    return [_make_issue(
        "polarity_mismatch", seg,
        f"Negation on only one side (source: {src_count}, target: {tgt_count}). "
        f"Possible polarity error — verify negations manually.",
    )]


def _has_letter(text: str) -> bool:
    for c in text:
        if c.isalpha():
            return True
    return False


def _check_untranslated_segment(seg, cfg, ctx):
    src = (seg["source_text"] or "").strip()
    tgt = (seg["target_text"] or "").strip()
    if not src or not tgt:
        return []
    if len(src) < _UNTRANSLATED_MIN_LEN:
        return []
    if not _has_letter(src):
        return []
    if src == tgt:
        return [_make_issue("untranslated_segment", seg,
                             "Target is identical to source (likely untranslated).")]
    return []


def _last_meaningful_char(text: str) -> str:
    """Return the last non-whitespace character of text, or '' if none."""
    for c in reversed(text or ""):
        if not c.isspace():
            return c
    return ""


def _check_alphanum_id_mismatch(seg, cfg, ctx):
    """Compare mixed alphanumeric IDs (subject codes, lot numbers, etc.)
    between source and target — they should match exactly (case-sensitive)
    because they are typically untranslatable opaque identifiers. Pure
    digit tokens are handled by ``number_mismatch`` and explicitly skipped
    inside ``_extract_alphanum_ids``."""
    src = seg["source_text"] or ""
    tgt = seg["target_text"] or ""
    src_ids = sorted(_extract_alphanum_ids(src))
    tgt_ids = sorted(_extract_alphanum_ids(tgt))
    if src_ids == tgt_ids:
        return []
    ss, st = _diff_spans(src_ids, tgt_ids)
    return [_make_issue(
        "alphanum_id_mismatch", seg,
        f"Alphanumeric IDs differ. Source: {src_ids or '∅'} | Target: {tgt_ids or '∅'}.",
        span_source=ss, span_target=st,
    )]


def _check_length_ratio(seg, cfg, ctx):
    """Flag suspiciously short/long target relative to source. Only fires
    above a minimum source length so short labels don't trigger false
    positives."""
    src = (seg["source_text"] or "").strip()
    tgt = (seg["target_text"] or "").strip()
    if not src or not tgt:
        return []
    min_chars = int(cfg.get("length_ratio_min_chars", _LENGTH_RATIO_MIN_CHARS))
    low = float(cfg.get("length_ratio_low", _LENGTH_RATIO_LOW))
    high = float(cfg.get("length_ratio_high", _LENGTH_RATIO_HIGH))
    if len(src) < min_chars:
        return []
    ratio = len(tgt) / max(1, len(src))
    if ratio < low:
        return [_make_issue("length_ratio", seg,
            f"Target unusually short: {len(tgt)} chars vs source {len(src)} "
            f"(ratio {ratio:.2f}, expected ≥ {low}).")]
    if ratio > high:
        return [_make_issue("length_ratio", seg,
            f"Target unusually long: {len(tgt)} chars vs source {len(src)} "
            f"(ratio {ratio:.2f}, expected ≤ {high}).")]
    return []


def _check_repeated_words(seg, cfg, ctx):
    """Detect adjacent duplicated words in the target (e.g. "the the",
    "el el"). Case-insensitive, skips digit-only tokens."""
    tgt = seg["target_text"] or ""
    if not tgt:
        return []
    # Use finditer (not findall) so we capture the exact matched text —
    # including any internal whitespace (\t, \n, multiple spaces). The
    # previous ``f"{w} {w}"`` construction assumed a single ASCII space
    # and silently failed to highlight when the duplicate was separated
    # by a tab or a newline.
    matches = list(_REPEATED_WORD_RE.finditer(tgt))
    if not matches:
        return []
    seen_lower = set()
    spans = []
    for m in matches:
        wl = m.group(1).lower()
        if wl in seen_lower:
            continue
        seen_lower.add(wl)
        spans.append(m.group(0))
    sample = ", ".join(repr(s) for s in spans[:5])
    return [_make_issue("repeated_words", seg,
        f"Target contains adjacent repeated word(s): {sample}.",
        span_target=spans)]


def _check_first_letter_case(seg, cfg, ctx):
    """Flag a first-letter capitalization mismatch between source and target.
    Only triggers when both first characters are letters."""
    src = (seg["source_text"] or "").lstrip()
    tgt = (seg["target_text"] or "").lstrip()
    if not src or not tgt:
        return []
    sc, tc = src[0], tgt[0]
    if not (sc.isalpha() and tc.isalpha()):
        return []
    if sc.isupper() and tc.islower():
        return [_make_issue("first_letter_case", seg,
            f"Source starts uppercase ('{sc}') but target starts lowercase ('{tc}').",
            span_source=sc, span_target=tc)]
    if sc.islower() and tc.isupper():
        return [_make_issue("first_letter_case", seg,
            f"Source starts lowercase ('{sc}') but target starts uppercase ('{tc}').",
            span_source=sc, span_target=tc)]
    return []


def _tag_pairing_broken(tags):
    """True iff a closing inline tag (``ept``) appears *before* its own
    opening tag (``bpt``) with the same id within this segment — the
    "close before open" inversion the QA tab reports as a broken pair.

    Only ids that have BOTH halves present in this segment are judged. memoQ
    legitimately splits a formatting run across segment boundaries, leaving a
    lone ``bpt`` or lone ``ept`` behind; those partnerless tags are ignored
    so cross-segment formatting never looks broken (a missing half is
    ``tag_count``'s concern, not ordering).

    Crucially this can NEVER fire on a *valid repositioning* of a well-formed
    ``bpt … ept`` pair (the translator moving bold/links to fit target word
    order): a complete pair always keeps its opener before its closer no
    matter where in the segment it is moved."""
    first_bpt: Dict[str, int] = {}
    first_ept: Dict[str, int] = {}
    for pos, (name, tid, *_) in enumerate(tags):
        if name == "bpt" and tid not in first_bpt:
            first_bpt[tid] = pos
        elif name == "ept" and tid not in first_ept:
            first_ept[tid] = pos
    for tid, ept_pos in first_ept.items():
        bpt_pos = first_bpt.get(tid)
        if bpt_pos is not None and ept_pos < bpt_pos:
            return True
    return False


def _check_tag_order(seg, cfg, ctx):
    """Report only *genuinely broken* inline-tag pairing in the target — a
    closing tag that appears before its matching opening tag — NOT legitimate
    repositioning.

    memoQ stores character formatting (bold, italic, links, …) as paired
    ``bpt``/``ept`` inline tags. Translators routinely move those pairs to a
    different position so the formatting lands on the right target words, and
    memoQ reports zero errors for that. The previous exact-sequence
    comparison flagged every such reposition as a false positive, so the
    check now fires only when the target's pairing is actually inverted
    (``_tag_pairing_broken``) relative to a well-formed source.

    Skipped when tag counts differ (``tag_count`` already reports that)."""
    src = _filter_content_tags(seg["source_tags"])
    tgt = _filter_content_tags(seg["target_tags"])
    if len(src) != len(tgt) or not src:
        return []
    if not _tag_pairing_broken(tgt) or _tag_pairing_broken(src):
        return []
    # Render compactly for the message: "bpt#1, ept#1".
    def _fmt(seq):
        return ", ".join(f"{n}#{i}" if i else n for n, i, *_ in seq)
    # Marker sequence (same `[id]` form as _element_text_with_markers) for the
    # highlight + cross-segment signature.
    src_marker_seq = tuple(_tag_markers(src))
    tgt_marker_seq = tuple(_tag_markers(tgt))
    return [_make_issue("tag_order", seg,
        "Tag pairing is broken in the target: a closing tag appears before "
        f"its opening tag. Source: [{_fmt(src)}] | Target: [{_fmt(tgt)}].",
        span_target=list(dict.fromkeys(tgt_marker_seq)) or None,
        tag_sig_missing=src_marker_seq,
        tag_sig_extra=tgt_marker_seq)]


def _check_final_punctuation_mismatch(seg, cfg, ctx):
    src = seg["source_text"] or ""
    tgt = seg["target_text"] or ""
    if not src.strip() or not tgt.strip():
        return []
    src_last = _last_meaningful_char(src)
    tgt_last = _last_meaningful_char(tgt)
    src_is_punct = src_last in _FINAL_PUNCT_CHARS
    tgt_is_punct = tgt_last in _FINAL_PUNCT_CHARS
    if not src_is_punct and not tgt_is_punct:
        return []
    # Treat fullwidth/CJK variants of '.', '!', '?' as equivalent to their ASCII versions.
    equiv = {"。": ".", "．": ".", "！": "!", "？": "?"}
    s = equiv.get(src_last, src_last) if src_is_punct else ""
    t = equiv.get(tgt_last, tgt_last) if tgt_is_punct else ""
    if s == t:
        return []
    if not src_is_punct:
        msg = f"Source has no final punctuation but target ends with '{tgt_last}'."
    elif not tgt_is_punct:
        msg = f"Source ends with '{src_last}' but target has no final punctuation."
    else:
        msg = f"Final punctuation differs: source ends with '{src_last}', target with '{tgt_last}'."
    # Paint the offending punctuation only on the side(s) where it exists.
    # If a side has no final punctuation, leave it un-highlighted — there is
    # literally nothing to mark there.
    return [_make_issue("final_punctuation_mismatch", seg, msg,
                        span_source=src_last if src_is_punct else None,
                        span_target=tgt_last if tgt_is_punct else None)]


# -----------------------------------------------------------------------------
# Public entry point
# -----------------------------------------------------------------------------

def _normalize_config(config: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    cfg = {**DEFAULT_CHECK_CONFIG, "enabled_checks": set(DEFAULT_CHECK_CONFIG["enabled_checks"])}
    if not config:
        return cfg
    for key, val in config.items():
        if key == "enabled_checks" and val is not None:
            cfg[key] = set(val)
        else:
            cfg[key] = val
    return cfg


def _check_spellcheck(seg, cfg, ctx):
    """Spell-check the target text. OPT-IN — pre-loaded dictionary lives in ctx."""
    target = seg.get("target_text") or ""
    if not target.strip():
        return []
    dictionary = ctx.get("_spellcheck_dict")
    if dictionary is None:
        # No dictionary available → check is silently skipped for this run
        # (a notice has already been queued at run_qa_checks level).
        return []
    try:
        from spellcheck import spell_check_text
    except Exception:
        return []
    ignore = ctx.get("_spellcheck_ignore") or set()
    misspelled = spell_check_text(target, dictionary, ignore_words=ignore)
    if not misspelled:
        return []
    cap = 20
    visible = misspelled[:cap]
    suffix = f" (… and {len(misspelled) - cap} more)" if len(misspelled) > cap else ""
    msg = (f"{len(misspelled)} possible misspelling(s) in target: "
           f"{', '.join(visible)}{suffix}")
    return [_make_issue("spellcheck", seg, msg, span_target=visible)]


def _check_confusable_pairs(seg, cfg, ctx):
    """Flag any target token that belongs to a user-supplied confusable group.

    Groups come exclusively from ``cfg["confusable_pairs_custom"]`` — the
    user supplies them via the textarea or .xlsx upload in the QA tab. No
    built-in language packs are shipped. If no custom groups are loaded
    the check is a no-op regardless of target language.

    Matching is whole-word and Unicode-aware (re-uses ``_term_pattern``).
    One issue per *distinct* matched member, painting the first
    occurrence of that token in the target.
    """
    target = seg.get("target_text") or ""
    if not target.strip():
        return []

    custom = cfg.get("confusable_pairs_custom") or []
    if not custom:
        return []

    groups: List[Tuple[str, ...]] = []
    for g in custom:
        if isinstance(g, (list, tuple)) and len(g) >= 2:
            groups.append(tuple(str(m) for m in g))

    if not groups:
        return []

    issues: List[Dict[str, Any]] = []
    seen_in_seg: set = set()
    for group in groups:
        members = [m for m in group if m]
        if len(members) < 2:
            continue
        for member in members:
            key = (member.lower(), tuple(sorted(m.lower() for m in members)))
            if key in seen_in_seg:
                continue
            pat = _term_pattern(member, case_sensitive=False)
            if not pat.search(target):
                continue
            seen_in_seg.add(key)
            others = [m for m in members if m.lower() != member.lower()]
            others_disp = ", ".join(f"'{o}'" for o in others)
            msg = (f"'{member}' detected — confirm you didn't mean {others_disp}.")
            issues.append(_make_issue(
                "confusable_pairs", seg, msg, span_target=member,
            ))
    return issues


# ---------------------------------------------------------------------------
# Cross-segment checks (Task #51).
# Per-segment checks see one segment at a time. Some checks need the whole
# document — e.g. inconsistent_translation groups segments by source and
# fires when their targets diverge. Run after the per-segment loop.
# ---------------------------------------------------------------------------

def highlight_diff(text_a: str, text_b: str, *,
                   css_class: Optional[str] = None) -> Tuple[str, str]:
    """Return (html_a, html_b) with word-level differences wrapped in red.

    Single source of truth for word-level diff rendering — used by the
    Anonymizer's Duplicates tab, the QA tab's `inconsistent_translation`
    card, and the HTML report. Pure helper, no Streamlit deps.

    css_class
        When ``None`` (default) each diff word is wrapped in an inline-
        styled ``<span>`` — matches the legacy ``app.highlight_diff``
        output that the Duplicates tab depends on.
        When ``"qa-span"`` the spans also get ``class="qa-span"`` and a
        ``data-token`` attribute so the HTML report's interactive
        toolbar (Ignore-by-token, group filter, Undo) can act on them.
    """
    import difflib
    def _esc(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    base_style = (
        "background:#ffcccc;color:#c0392b;font-weight:600;"
        "padding:0 2px;border-radius:2px;"
    )
    if css_class == "qa-span":
        wrap_tpl = (
            f'<span class="qa-span" data-token="{{tok}}" '
            f'style="{base_style}">{{w}}</span>'
        )
    else:
        wrap_tpl = f'<span style="{base_style}">{{w}}</span>'

    def _wrap(chunk: str) -> str:
        if css_class == "qa-span":
            tok = chunk.lower().replace('"', "&quot;").replace("'", "&#39;")
            return wrap_tpl.format(w=chunk, tok=tok)
        return wrap_tpl.format(w=chunk)

    words_a = (text_a or "").split()
    words_b = (text_b or "").split()
    sm = difflib.SequenceMatcher(None, words_a, words_b)
    parts_a: List[str] = []
    parts_b: List[str] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        chunk_a = _esc(" ".join(words_a[i1:i2]))
        chunk_b = _esc(" ".join(words_b[j1:j2]))
        if tag == "equal":
            parts_a.append(chunk_a)
            parts_b.append(chunk_b)
        elif tag == "replace":
            parts_a.append(_wrap(chunk_a))
            parts_b.append(_wrap(chunk_b))
        elif tag == "delete":
            parts_a.append(_wrap(chunk_a))
        elif tag == "insert":
            parts_b.append(_wrap(chunk_b))
    return " ".join(parts_a), " ".join(parts_b)


# Back-compat alias used by older internal call sites + tests.
_highlight_diff_html = highlight_diff


def _check_inconsistent_translation(segments: List[Dict[str, Any]],
                                     cfg: Dict[str, Any],
                                     ctx: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Cross-segment check with three patterns + pair-level dedup (Task #55).

    Patterns emitted:
      * ``axis="target"`` — same/similar SOURCE, different TARGET (original
        Task #51 behaviour).
      * ``axis="source"`` — same/similar TARGET, different SOURCE (new:
        a single translation reused for multiple distinct source texts).
      * ``axis="both"``   — fuzzy path only. Neither side identical but
        BOTH source and target are above the similarity threshold:
        likely a near-duplicate translation pair that needs a human eye.

    Threshold semantics match the Duplicates tab:
      * ``1.0``  → exact match only (fast hash-grouping; no "axis=both").
      * ``<1.0`` → enables fuzzy via ``difflib.SequenceMatcher`` on the
        full string (O(n^2) over candidate segments).

    Pair-level dedup: every (a, b) pair fires at most ONE issue, even if
    multiple patterns would match. Priority — exact-axis (``target`` or
    ``source``) > ``fuzzy-both`` > fuzzy single-axis.

    Length guard: ``inconsistent_translation_min_chars`` (default 8) is
    applied to the *grouping axis* — source length for ``axis=target``,
    target length for ``axis=source``, both for ``axis=both``. Stops the
    check from going crazy on 1-2 word UI strings ("OK", "Cancel", etc).

    Issue payload (renderer-visible):
      * ``mismatch_axis``        — "target" | "source" | "both"
      * ``reference_segment_id`` — lowest segment id of the pair
      * ``reference_source``     — reference's source text (for the
                                   stacked comparison block in the QA
                                   card and HTML report)
      * ``reference_target``     — reference's target text
      * ``diff_against``         — kept for back-compat (= reference_target)
      * ``sibling_segment_ids``  — sorted ``[ref_id, other_id]``
    """
    try:
        threshold = float(cfg.get("inconsistent_translation_threshold", 1.00))
    except (TypeError, ValueError):
        threshold = 1.0
    threshold = max(0.5, min(1.0, threshold))

    try:
        min_chars = int(cfg.get("inconsistent_translation_min_chars", 8))
    except (TypeError, ValueError):
        min_chars = 8
    min_chars = max(0, min_chars)

    # Only consider segments that have BOTH non-empty source and target —
    # an empty target is already covered by `empty_target`, and we don't
    # want to fire on garbage rows.
    candidates = []
    for seg in segments:
        src = (seg.get("source_text") or "").strip()
        tgt = (seg.get("target_text") or "").strip()
        if src and tgt:
            candidates.append(seg)

    # Pair dedup table. Key: frozenset({ref_id, other_id}). Value: dict
    # with the strongest signal seen so far for that pair.
    _PRIORITY = {
        "exact-target": 4, "exact-source": 4,
        "fuzzy-both": 3,
        "fuzzy-target": 2, "fuzzy-source": 2,
    }
    pairs: Dict[frozenset, Dict[str, Any]] = {}

    def _record(ref_seg: Dict[str, Any], other_seg: Dict[str, Any],
                kind: str, group_size: int) -> None:
        # Reference is always the lower segment id of the pair so the
        # message reads "Reference (seg X)" with X < Y.
        if other_seg["id"] < ref_seg["id"]:
            ref_seg, other_seg = other_seg, ref_seg
        key = frozenset({ref_seg["id"], other_seg["id"]})
        existing = pairs.get(key)
        if existing and _PRIORITY[existing["_kind"]] >= _PRIORITY[kind]:
            return
        pairs[key] = {
            "_kind": kind,
            "ref": ref_seg,
            "other": other_seg,
            "group_size": group_size,
        }

    # Pass 1 — exact SOURCE grouping → axis=target
    groups_src: Dict[str, List[Dict[str, Any]]] = {}
    for seg in candidates:
        src = (seg.get("source_text") or "").strip()
        if len(src) < min_chars:
            continue
        groups_src.setdefault(src, []).append(seg)
    for segs in groups_src.values():
        if len(segs) < 2:
            continue
        segs.sort(key=lambda s: s["id"])
        ref = segs[0]
        ref_tgt = (ref.get("target_text") or "").strip()
        diverging = [s for s in segs[1:]
                     if (s.get("target_text") or "").strip() != ref_tgt]
        if not diverging:
            continue
        n = len(segs)
        for other in diverging:
            _record(ref, other, "exact-target", n)

    # Pass 2 — exact TARGET grouping → axis=source (new in Task #55)
    groups_tgt: Dict[str, List[Dict[str, Any]]] = {}
    for seg in candidates:
        tgt = (seg.get("target_text") or "").strip()
        if len(tgt) < min_chars:
            continue
        groups_tgt.setdefault(tgt, []).append(seg)
    for segs in groups_tgt.values():
        if len(segs) < 2:
            continue
        segs.sort(key=lambda s: s["id"])
        ref = segs[0]
        ref_src = (ref.get("source_text") or "").strip()
        diverging = [s for s in segs[1:]
                     if (s.get("source_text") or "").strip() != ref_src]
        if not diverging:
            continue
        n = len(segs)
        for other in diverging:
            _record(ref, other, "exact-source", n)

    # Pass 3 — fuzzy pair scan (O(n^2)). Guardrail: skip on very large
    # files to keep QA runs from blowing up; user is warned in the UI.
    _FUZZY_MAX_CANDIDATES = 2000
    if threshold < 1.0 and len(candidates) <= _FUZZY_MAX_CANDIDATES:
        import difflib
        n_cand = len(candidates)
        for i in range(n_cand):
            a = candidates[i]
            src_a = (a.get("source_text") or "").strip()
            tgt_a = (a.get("target_text") or "").strip()
            for j in range(i + 1, n_cand):
                b = candidates[j]
                key = frozenset({a["id"], b["id"]})
                existing = pairs.get(key)
                # Exact-axis already wins; nothing fuzzy can upgrade it.
                if existing and _PRIORITY[existing["_kind"]] >= 4:
                    continue
                src_b = (b.get("source_text") or "").strip()
                tgt_b = (b.get("target_text") or "").strip()
                src_same = (src_a == src_b)
                tgt_same = (tgt_a == tgt_b)
                if src_same and tgt_same:
                    continue
                # Only compute ratios that may matter, with length guard
                # applied to BOTH segments on the relevant axis.
                src_sim = False
                tgt_sim = False
                if (not src_same
                        and len(src_a) >= min_chars
                        and len(src_b) >= min_chars):
                    if difflib.SequenceMatcher(None, src_a, src_b).ratio() >= threshold:
                        src_sim = True
                if (not tgt_same
                        and len(tgt_a) >= min_chars
                        and len(tgt_b) >= min_chars):
                    if difflib.SequenceMatcher(None, tgt_a, tgt_b).ratio() >= threshold:
                        tgt_sim = True

                # axis=both: neither side identical, both fuzzy-similar.
                if src_sim and tgt_sim:
                    _record(a, b, "fuzzy-both", 2)
                # axis=target: fuzzy source, different target. Note we
                # don't require `tgt_same is False` separately — if
                # targets were identical there'd be no inconsistency.
                elif src_sim and not tgt_same:
                    _record(a, b, "fuzzy-target", 2)
                # axis=source: fuzzy target, different source.
                elif tgt_sim and not src_same:
                    _record(a, b, "fuzzy-source", 2)

    # Materialize pairs into issue dicts.
    issues: List[Dict[str, Any]] = []
    for pair in pairs.values():
        kind = pair["_kind"]
        ref = pair["ref"]
        other = pair["other"]
        ref_src = (ref.get("source_text") or "").strip()
        ref_tgt = (ref.get("target_text") or "").strip()
        n = pair["group_size"]
        if kind == "exact-target":
            axis = "target"
            msg = (f"Same source appears in {n} segments with different "
                   f"targets. Reference (seg {ref['id']}). Target differs.")
        elif kind == "exact-source":
            axis = "source"
            msg = (f"Same target appears in {n} segments with different "
                   f"sources. Reference (seg {ref['id']}). Source differs.")
        elif kind == "fuzzy-target":
            axis = "target"
            msg = (f"Near-duplicate source (fuzzy match) with different "
                   f"target. Reference (seg {ref['id']}). Target differs.")
        elif kind == "fuzzy-source":
            axis = "source"
            msg = (f"Near-duplicate target (fuzzy match) with different "
                   f"source. Reference (seg {ref['id']}). Source differs.")
        else:  # fuzzy-both
            axis = "both"
            msg = (f"Near-duplicate pair: both source and target differ "
                   f"from reference (seg {ref['id']}).")
        issue = _make_issue("inconsistent_translation", other, msg)
        issue["mismatch_axis"] = axis
        issue["reference_segment_id"] = ref["id"]
        issue["reference_source"] = ref_src
        issue["reference_target"] = ref_tgt
        # Back-compat: legacy renderers/tests still read `diff_against`.
        issue["diff_against"] = ref_tgt
        issue["sibling_segment_ids"] = sorted([ref["id"], other["id"]])
        issues.append(issue)

    issues.sort(key=lambda i: (i["segment_id"], i.get("reference_segment_id") or 0))
    return issues


# Task #66 — ALLUPPERCASE casing check. Token regex requires 3+ uppercase
# letters (incl. Latin-Extended A/B for ÁÉÍÓÚÑÇÄÖÜ etc.) bounded by
# non-word chars on both sides. Common clinical/pharma acronyms are
# whitelisted to keep the signal high — they're already-uppercase by
# convention and almost never get downcased in translation.
_ALLCAPS_TOKEN_RE = re.compile(
    r"(?<!\w)[A-ZÀ-ÖØ-Þ\u0100-\u017F]{3,}(?!\w)", re.UNICODE)
_ALLCAPS_WORD_RE = re.compile(
    r"(?<!\w)[^\W\d_]{3,}(?!\w)", re.UNICODE)
_ALLCAPS_ACRONYM_WHITELIST = frozenset({
    "FDA", "EMA", "NCT", "CDC", "IRB", "WHO", "EU", "US", "USA", "UK",
    "ID", "PDF", "HTML", "XML", "URL", "API", "HIV", "DNA", "RNA", "PCR",
    "MRI", "EKG", "ECG", "HIPAA", "GDPR", "ICH", "GCP", "AE", "SAE", "ADR",
    "BMI", "SAP", "TMF", "CRF", "PK", "PD", "COVID", "SARS",
    "OMS", "AEMPS", "BOE", "DOUE", "UE", "PCT", "ISO", "IEC", "NIH",
})


def _check_alluppercase_mismatch(seg, cfg, ctx):
    """Flag ALLUPPERCASE asymmetry between source and target.

    Two firing modes:
      1. **Paired** — the same word (case-insensitive match) appears
         ALLCAPS on one side and not the other.  Best signal; pairs the
         exact spans so the diff is unambiguous.
      2. **Stylistic asymmetry** — one side has 1+ ALLCAPS non-acronym
         tokens and the other side has zero ALLCAPS tokens (after
         acronym whitelist).  Catches the cross-language case where the
         tokens don't share letters (e.g. ``Introduction`` →
         ``INTRODUCCIÓN``) and only fires when mode 1 produced nothing
         for the segment.

    Acronyms (``FDA``, ``NCT``, ``WHO`` …) are ignored on both sides:
    they're already-uppercase by convention and never get downcased.
    """
    src = seg["source_text"]
    tgt = seg["target_text"]
    if not src or not tgt:
        return []
    src_caps = [w for w in _ALLCAPS_TOKEN_RE.findall(src)
                if w not in _ALLCAPS_ACRONYM_WHITELIST]
    tgt_caps = [w for w in _ALLCAPS_TOKEN_RE.findall(tgt)
                if w not in _ALLCAPS_ACRONYM_WHITELIST]
    if not src_caps and not tgt_caps:
        return []
    issues: List[Dict[str, Any]] = []
    seen_pairs: set = set()
    src_words = {m.group().lower(): m.group()
                 for m in _ALLCAPS_WORD_RE.finditer(src)}
    tgt_words = {m.group().lower(): m.group()
                 for m in _ALLCAPS_WORD_RE.finditer(tgt)}
    paired_src_lower: set = set()
    paired_tgt_lower: set = set()
    # Mode 1: paired tokens
    for w in src_caps:
        wl = w.lower()
        if wl in tgt_words:
            t = tgt_words[wl]
            paired_src_lower.add(wl)
            paired_tgt_lower.add(wl)
            if not t.isupper() and (w, t) not in seen_pairs:
                seen_pairs.add((w, t))
                issues.append(_make_issue("alluppercase_mismatch", seg,
                    f"'{w}' is ALLCAPS in source but rendered as '{t}' "
                    f"in target.",
                    span_source=w, span_target=t))
    for w in tgt_caps:
        wl = w.lower()
        if wl in src_words and wl not in paired_src_lower:
            s = src_words[wl]
            paired_tgt_lower.add(wl)
            if not s.isupper() and (s, w) not in seen_pairs:
                seen_pairs.add((s, w))
                issues.append(_make_issue("alluppercase_mismatch", seg,
                    f"'{w}' is ALLCAPS in target but rendered as '{s}' "
                    f"in source.",
                    span_source=s, span_target=w))
    # Mode 2: stylistic asymmetry (only when mode 1 produced nothing).
    # Catches Introduction → INTRODUCCIÓN where the letters don't match
    # across languages so mode 1 can't pair them.
    if not issues:
        unpaired_src = [w for w in src_caps if w.lower() not in paired_src_lower]
        unpaired_tgt = [w for w in tgt_caps if w.lower() not in paired_tgt_lower]
        if unpaired_tgt and not unpaired_src:
            unique: List[str] = []
            for w in unpaired_tgt:
                if w not in unique:
                    unique.append(w)
            issues.append(_make_issue("alluppercase_mismatch", seg,
                f"Target uses ALLCAPS ({', '.join(repr(w) for w in unique)}) "
                f"but source has no ALLCAPS counterpart.",
                span_source=None, span_target=unique))
        elif unpaired_src and not unpaired_tgt:
            unique = []
            for w in unpaired_src:
                if w not in unique:
                    unique.append(w)
            issues.append(_make_issue("alluppercase_mismatch", seg,
                f"Source uses ALLCAPS ({', '.join(repr(w) for w in unique)}) "
                f"but target has no ALLCAPS counterpart.",
                span_source=unique, span_target=None))
    return issues


# CamelCase: at least one internal lowercase→UPPERCASE boundary. Matches
# "PharmaExample", "iPhone", "JavaScript", "memoQ", "iOS".  Excludes
# ALLCAPS tokens (no lower→upper boundary) and pure lowercase tokens.
# Tokens that contain digits are skipped — those are alphanumeric IDs
# already handled by `_check_alphanum_id_mismatch`.
_CAMELCASE_TOKEN_RE = re.compile(r"(?<!\w)[A-Za-z]*[a-z][A-Z][A-Za-z]*(?!\w)")


def _check_camelcase_mismatch(seg, cfg, ctx):
    """Flag CamelCase tokens (lowercase→UPPERCASE boundary inside a
    word) present in source but missing literally in target. Case-
    sensitive because CamelCase tokens are identifiers / product
    names that must be preserved verbatim. Tokens containing digits
    fall through to ``alphanum_id_mismatch``."""
    src = seg["source_text"]
    tgt = seg["target_text"]
    if not src or not tgt:
        return []
    src_cc = _CAMELCASE_TOKEN_RE.findall(src)
    if not src_cc:
        return []
    seen: set = set()
    missing: List[str] = []
    for tok in src_cc:
        if any(ch.isdigit() for ch in tok):
            continue
        if tok in seen:
            continue
        seen.add(tok)
        if tok not in tgt:
            missing.append(tok)
    if not missing:
        return []
    return [_make_issue("camelcase_mismatch", seg,
        f"CamelCase token(s) in source not found literally in target: "
        f"{', '.join(repr(t) for t in missing)}.",
        span_source=missing, span_target=None)]


# ---------------------------------------------------------------------------
# Task #61 — User-supplied custom regex patterns
# ---------------------------------------------------------------------------

try:
    import regex as _regex_mod  # PyPI ``regex`` package (transitive via spaCy)
    _HAS_REGEX_MOD = True
except Exception:  # pragma: no cover - defensive; spaCy hard-requires it
    _regex_mod = None
    _HAS_REGEX_MOD = False


def parse_custom_regex_patterns(text: str, case_sensitive: bool = False
                                ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Parse the QA Configuration textarea into compiled regex patterns.

    Each non-empty, non-comment line (``#`` prefix after .strip()) is
    compiled twice:

    1. With the stdlib ``re`` module — to surface familiar ``re.error``
       messages to the user (Python-style escape errors, unbalanced
       parens, etc).
    2. With the third-party ``regex`` module (PyPI) — used at runtime
       inside :func:`_safe_regex_findall` because it supports a native
       per-call ``timeout=`` argument that the stdlib does not. The
       ``regex`` module is a strict superset of ``re``, so any pattern
       that compiles with ``re`` also compiles with ``regex``; if the
       second compile fails (extremely unlikely) the pattern is still
       accepted and the runtime falls back to the stdlib match (no
       enforced timeout for that one pattern).

    Returns ``(patterns, errors)`` where each pattern dict carries
    ``line``, ``raw``, ``pattern`` (stdlib compile), ``regex_pattern``
    (PyPI compile or ``None``). The function never raises — invalid
    lines are reported via the ``errors`` channel.
    """
    patterns: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    if not text:
        return patterns, errors
    re_flags = re.UNICODE if case_sensitive else (re.IGNORECASE | re.UNICODE)
    if _HAS_REGEX_MOD:
        rx_flags = _regex_mod.UNICODE if case_sensitive else (
            _regex_mod.IGNORECASE | _regex_mod.UNICODE)
    for idx, raw_line in enumerate(text.replace("\r", "\n").split("\n"), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        try:
            compiled = re.compile(line, re_flags)
        except re.error as exc:
            errors.append({"line": idx, "raw": line, "error": str(exc)})
            continue
        rx_compiled = None
        if _HAS_REGEX_MOD:
            try:
                rx_compiled = _regex_mod.compile(line, rx_flags)
            except Exception:  # pragma: no cover - very rare compat gap
                rx_compiled = None
        patterns.append({"line": idx, "raw": line, "pattern": compiled,
                         "regex_pattern": rx_compiled})
    return patterns, errors


class _RegexTimeoutError(Exception):
    """Raised when a user-supplied regex exceeds its per-segment budget."""


def _safe_regex_findall(entry: Any, text: str,
                        budget_seconds: float = 0.25) -> List[Any]:
    """Run ``findall(text)`` under an *enforced* wallclock guard.

    ``entry`` may be either a pattern dict from
    :func:`parse_custom_regex_patterns` (preferred — carries the
    pre-compiled ``regex`` module pattern that supports native
    timeouts) or a bare stdlib ``re.Pattern`` (used by tests for the
    bare-bones contract).

    The PyPI ``regex`` module exposes ``Pattern.findall(text,
    timeout=…)`` which, on exceeding the budget, raises
    ``TimeoutError``. We translate that to ``_RegexTimeoutError`` so
    callers have a single sentinel to catch regardless of which engine
    actually ran. CPython's stdlib ``re`` holds the GIL during matching
    (so a worker-thread guard does not help), which is precisely why we
    use ``regex`` for the enforced timeout path.

    Fallback: if neither the entry nor the global ``regex`` module is
    available, run stdlib ``re`` unguarded — accepted only as a
    defensive last resort; the parse step always tries to populate
    ``regex_pattern`` so this path is essentially unreachable in
    production.
    """
    rx_pattern = None
    re_pattern = None
    if isinstance(entry, dict):
        rx_pattern = entry.get("regex_pattern")
        re_pattern = entry.get("pattern")
    else:
        re_pattern = entry

    if rx_pattern is not None:
        try:
            return rx_pattern.findall(text, timeout=max(0.0, float(budget_seconds)))
        except TimeoutError as exc:
            raise _RegexTimeoutError() from exc

    if re_pattern is None:
        return []
    return re_pattern.findall(text)


def _flatten_findall(matches: List[Any]) -> List[str]:
    """Coerce ``re.findall`` output to a flat list of non-empty strings.

    ``findall`` returns tuples when the pattern has 2+ groups; strings
    otherwise. We pick the first non-empty group for tuples so the
    highlighter has something concrete to paint.
    """
    out: List[str] = []
    for m in matches:
        if isinstance(m, tuple):
            picked = next((g for g in m if g), "")
        else:
            picked = m
        if picked:
            out.append(picked)
    return out


def _check_custom_forbidden_regex(seg, cfg, ctx):
    patterns = cfg.get("custom_forbidden_regex_patterns") or []
    if not patterns:
        return []
    target = seg.get("target_text") or ""
    if not target:
        return []
    budget = float(cfg.get("custom_regex_timeout_seconds") or 0.25)
    issues: List[Dict[str, Any]] = []
    for entry in patterns:
        raw = entry.get("raw", "")
        if entry.get("pattern") is None and entry.get("regex_pattern") is None:
            continue
        try:
            matches = _safe_regex_findall(entry, target, budget_seconds=budget)
        except _RegexTimeoutError:
            ctx.setdefault("_custom_regex_timeout_segs", set()).add(seg["id"])
            continue
        if not matches:
            continue
        spans = _flatten_findall(matches) or None
        msg = (f"Custom forbidden pattern matched target: /{raw}/ "
               f"({len(matches)} match(es)).")
        issues.append(_make_issue(
            "custom_forbidden_regex", seg, msg,
            span_target=spans,
        ))
    return issues


def _check_custom_required_regex(seg, cfg, ctx):
    patterns = cfg.get("custom_required_regex_patterns") or []
    if not patterns:
        return []
    source = seg.get("source_text") or ""
    target = seg.get("target_text") or ""
    if not source:
        return []
    budget = float(cfg.get("custom_regex_timeout_seconds") or 0.25)
    issues: List[Dict[str, Any]] = []
    for entry in patterns:
        raw = entry.get("raw", "")
        if entry.get("pattern") is None and entry.get("regex_pattern") is None:
            continue
        try:
            src_matches = _safe_regex_findall(entry, source, budget_seconds=budget)
            tgt_matches = _safe_regex_findall(entry, target, budget_seconds=budget)
        except _RegexTimeoutError:
            ctx.setdefault("_custom_regex_timeout_segs", set()).add(seg["id"])
            continue
        if not src_matches:
            # Pattern absent from source -> no expectation on target.
            continue
        if len(src_matches) == len(tgt_matches):
            continue
        src_spans = _flatten_findall(src_matches) or None
        tgt_spans = _flatten_findall(tgt_matches) or None
        msg = (f"Custom required pattern /{raw}/: source has "
               f"{len(src_matches)} match(es), target has "
               f"{len(tgt_matches)}.")
        issues.append(_make_issue(
            "custom_required_regex", seg, msg,
            span_source=src_spans,
            span_target=tgt_spans,
        ))
    return issues


# --- Task #71: cross-segment extensions ----------------------------------
# Three cross-document scans that piggyback on existing categories:
#   * Hyphenation inconsistency → emits under ``inconsistent_translation``
#     with ``mismatch_axis="hyphenation"``.
#   * Date format inconsistency  → emits under ``date_format_mismatch``.
#   * Number format inconsistency → emits under ``number_format_mismatch``.
# All three only look at TARGET text (translator output is what we're
# auditing for internal consistency) and ignore the source.

_HYPHENATED_TOKEN_RE = re.compile(
    r"[^\W\d_]{2,}(?:-[^\W\d_]{2,})+", re.UNICODE,
)
_PLAIN_WORD_RE = re.compile(r"[^\W\d_]{4,}", re.UNICODE)


def _check_hyphenation_inconsistency(segments: List[Dict[str, Any]],
                                     cfg: Dict[str, Any],
                                     ctx: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Detect the same compound term appearing hyphenated in one
    segment and unhyphenated (or hyphenated differently) in another —
    a frequent translator-side inconsistency, especially in DE/NL/SV
    compound nouns. Emits under ``inconsistent_translation`` with
    ``mismatch_axis="hyphenation"``.

    Pair-level dedup: at most one issue per (segA, segB, normalized
    token) triple. Reference is always the lowest segment id of the
    pair, mirroring the existing ``_check_inconsistent_translation``
    convention so renderers reuse the same stacked layout.
    """
    # Pass 1 — index every hyphenated token by its hyphen-stripped form.
    hyph_index: Dict[str, List[Tuple[int, str, Dict[str, Any]]]] = {}
    for seg in segments:
        tgt = (seg.get("target_text") or "").strip()
        if not tgt or "-" not in tgt:
            continue
        seen_local: set = set()
        for m in _HYPHENATED_TOKEN_RE.finditer(tgt):
            tok = m.group(0)
            norm = tok.replace("-", "").lower()
            if len(norm) < 6 or norm in seen_local:
                continue
            seen_local.add(norm)
            hyph_index.setdefault(norm, []).append((seg["id"], tok, seg))
    if not hyph_index:
        return []

    # Pass 2 — for each indexed norm, scan all segments for an
    # unhyphenated plain-word match. Track only segments not already
    # contributing the hyphenated variant.
    pair_seen: set = set()
    issues: List[Dict[str, Any]] = []
    for norm, hyph_occurrences in hyph_index.items():
        hyph_seg_ids = {sid for sid, _, _ in hyph_occurrences}
        plain_occurrences: List[Tuple[int, str, Dict[str, Any]]] = []
        for seg in segments:
            sid = seg["id"]
            if sid in hyph_seg_ids:
                continue
            tgt = (seg.get("target_text") or "")
            if not tgt:
                continue
            # Fast pre-filter to avoid the regex on segments that
            # cannot possibly contain the normalized token.
            if norm not in tgt.lower():
                continue
            for m in _PLAIN_WORD_RE.finditer(tgt):
                if m.group(0).lower() == norm:
                    plain_occurrences.append((sid, m.group(0), seg))
                    break
        if not plain_occurrences:
            continue
        # Build issue pairs: each plain occurrence vs the lowest-id
        # hyphenated occurrence (= reference).
        ref_sid, ref_tok, ref_seg = min(hyph_occurrences, key=lambda t: t[0])
        ref_src = (ref_seg.get("source_text") or "").strip()
        ref_tgt = (ref_seg.get("target_text") or "").strip()
        for sid, tok, seg in plain_occurrences:
            # Order pair so reference is the lower id.
            if sid < ref_sid:
                a_sid, a_seg, a_tok = sid, seg, tok
                b_sid, b_seg, b_tok = ref_sid, ref_seg, ref_tok
            else:
                a_sid, a_seg, a_tok = ref_sid, ref_seg, ref_tok
                b_sid, b_seg, b_tok = sid, seg, tok
            key = (a_sid, b_sid, norm)
            if key in pair_seen:
                continue
            pair_seen.add(key)
            msg = (f"Hyphenation inconsistency: '{a_tok}' (seg {a_sid}) "
                   f"vs '{b_tok}' (seg {b_sid}).")
            issue = _make_issue("inconsistent_translation", b_seg, msg)
            issue["mismatch_axis"] = "hyphenation"
            issue["reference_segment_id"] = a_sid
            issue["reference_source"] = (a_seg.get("source_text") or "").strip()
            issue["reference_target"] = (a_seg.get("target_text") or "").strip()
            issue["diff_against"] = issue["reference_target"]
            issue["sibling_segment_ids"] = sorted([a_sid, b_sid])
            issues.append(issue)
    issues.sort(key=lambda i: (i["segment_id"], i.get("reference_segment_id") or 0))
    return issues


def _date_separator_of(token_text: str) -> Optional[str]:
    """Return the separator char of a date token matched by ``_DATE_RE``
    (``/``, ``.`` or ``-``), else None."""
    m = _DATE_RE.match(token_text)
    return m.group(2) if m else None


def _check_date_format_inconsistency_cross(segments: List[Dict[str, Any]],
                                           cfg: Dict[str, Any],
                                           ctx: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Cross-segment: flag target segments whose date separator differs
    from the document-wide majority. Same category as
    ``date_format_mismatch`` so it surfaces in the same card.

    Only fires when at least two distinct separators appear in the
    target across the document (otherwise there is nothing to
    inconsistently mix).
    """
    per_seg: List[Tuple[Dict[str, Any], List[str], List[str]]] = []
    sep_counts: "Counter[str]" = Counter()
    for seg in segments:
        tgt = (seg.get("target_text") or "")
        if not tgt or not any(c.isdigit() for c in tgt):
            continue
        seps: List[str] = []
        raws: List[str] = []
        for m in _DATE_RE.finditer(tgt):
            seps.append(m.group(2))
            raws.append(m.group(0))
        if seps:
            per_seg.append((seg, seps, raws))
            sep_counts.update(seps)
    if len(sep_counts) < 2:
        return []
    # Require a STRICT majority (top count > second count) before
    # flagging outliers; in a tie there is no canonical majority to
    # compare against, so emitting issues would be arbitrary.
    top_two = sep_counts.most_common(2)
    if len(top_two) >= 2 and top_two[0][1] == top_two[1][1]:
        return []
    majority = top_two[0][0]
    issues: List[Dict[str, Any]] = []
    for seg, seps, raws in per_seg:
        odd = [(s, r) for s, r in zip(seps, raws) if s != majority]
        if not odd:
            continue
        odd_raws = [r for _, r in odd]
        odd_seps = sorted({s for s, _ in odd})
        issues.append(_make_issue(
            "date_format_mismatch", seg,
            f"Date format inconsistent across document: segment uses "
            f"separator(s) {odd_seps} (e.g. {odd_raws[0]}) while the "
            f"majority of the document uses '{majority}'.",
            span_target=odd_raws,
        ))
    return issues


def _decimal_separator_of(token: str) -> Optional[str]:
    """Return the *decimal* separator char of a numeric token, or None
    when the token is unambiguous integer / unrecognised / has only a
    3-digit suffix (true thousand-grouping → no decimal info)."""
    if "." in token and "," in token:
        return "." if token.rfind(".") > token.rfind(",") else ","
    if "." in token:
        suffix = token.rsplit(".", 1)[1]
        if suffix.isdigit() and len(suffix) in (1, 2):
            return "."
        if suffix.isdigit() and len(suffix) >= 4:
            return "."
        return None
    if "," in token:
        suffix = token.rsplit(",", 1)[1]
        if suffix.isdigit() and len(suffix) in (1, 2):
            return ","
        if suffix.isdigit() and len(suffix) >= 4:
            return ","
        return None
    return None


def _check_number_format_inconsistency_cross(segments: List[Dict[str, Any]],
                                             cfg: Dict[str, Any],
                                             ctx: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Cross-segment: flag target segments whose decimal separator
    differs from the document-wide majority. Emits under
    ``number_format_mismatch``.

    Only ``.`` and ``,`` are considered — and only when the token is
    unambiguously a decimal (1- or 2-digit suffix, or 4+-digit suffix).
    Pure integers, 3-digit-suffix tokens (genuine ambiguity), times,
    dates and IDs are skipped because they cannot characterise the
    document's locale convention reliably.
    """
    per_seg: List[Tuple[Dict[str, Any], List[str], List[str]]] = []
    sep_counts: "Counter[str]" = Counter()
    for seg in segments:
        tgt = (seg.get("target_text") or "")
        if not tgt or not any(c.isdigit() for c in tgt):
            continue
        # Mask date spans so date separators do not leak into the
        # numeric-format characterisation.
        masked = tgt
        for m in _DATE_RE.finditer(tgt):
            s, e = m.start(), m.end()
            masked = masked[:s] + (" " * (e - s)) + masked[e:]
        seps: List[str] = []
        raws: List[str] = []
        for m in _NUMBER_RE.finditer(masked):
            tok = m.group(0)
            sep = _decimal_separator_of(tok)
            if sep is None:
                continue
            seps.append(sep)
            raws.append(tok)
        if seps:
            per_seg.append((seg, seps, raws))
            sep_counts.update(seps)
    if len(sep_counts) < 2:
        return []
    # Strict-majority guard (architect review): only flag when the top
    # count is strictly greater than the runner-up. Ties have no
    # canonical "document convention" to compare segments against.
    top_two = sep_counts.most_common(2)
    if len(top_two) >= 2 and top_two[0][1] == top_two[1][1]:
        return []
    majority = top_two[0][0]
    issues: List[Dict[str, Any]] = []
    for seg, seps, raws in per_seg:
        odd = [(s, r) for s, r in zip(seps, raws) if s != majority]
        if not odd:
            continue
        odd_raws = [r for _, r in odd]
        issues.append(_make_issue(
            "number_format_mismatch", seg,
            f"Number format inconsistent across document: segment uses "
            f"decimal separator '{odd[0][0]}' (e.g. {odd_raws[0]}) while "
            f"the majority of the document uses '{majority}'.",
            span_target=odd_raws,
        ))
    return issues


_PER_SEGMENT_CHECKS: List[Tuple[str, Callable]] = [
    ("empty_target", _check_empty_target),
    ("untranslated_segment", _check_untranslated_segment),
    ("control_characters", _check_control_characters),
    ("mixed_scripts", _check_mixed_scripts),
    ("length_ratio", _check_length_ratio),
    ("repeated_words", _check_repeated_words),
    ("forbidden_terms", _check_forbidden_terms),
    ("glossary_violation", _check_glossary_violation),
    ("bracket_balance", _check_bracket_balance),
    ("tag_count", _check_tag_count),
    ("tag_order", _check_tag_order),
    ("tag_malformed", _check_tag_malformed),
    ("urls", _check_urls),
    ("emails", _check_emails),
    ("placeholders", _check_placeholders),
    ("number_mismatch", _check_number_mismatch),
    ("range_ratio_mismatch", _check_range_ratio_mismatch),
    ("number_format_mismatch", _check_number_format_mismatch),
    ("unit_mismatch", _check_unit_mismatch),
    ("date_format_mismatch", _check_date_format_mismatch),
    ("alphanum_id_mismatch", _check_alphanum_id_mismatch),
    ("polarity_mismatch", _check_polarity_mismatch),
    ("symbol_mismatch", _check_symbol_mismatch),
    ("repeated_punctuation", _check_repeated_punctuation),
    ("final_punctuation_mismatch", _check_final_punctuation_mismatch),
    ("first_letter_case", _check_first_letter_case),
    ("alluppercase_mismatch", _check_alluppercase_mismatch),
    ("camelcase_mismatch", _check_camelcase_mismatch),
    ("custom_forbidden_regex", _check_custom_forbidden_regex),
    ("custom_required_regex", _check_custom_required_regex),
    ("double_spaces", _check_double_spaces),
    ("whitespace_edges", _check_whitespace_edges),
    ("spellcheck", _check_spellcheck),
    ("confusable_pairs", _check_confusable_pairs),
]


def _prepare_spellcheck_context(target_lang: str, cfg: Dict[str, Any],
                                  ctx: Dict[str, Any]) -> Optional[str]:
    """Resolve the target dictionary for the run. Returns a notice string
    when spell-check had to be skipped, or None on success."""
    try:
        import spellcheck as _sc
    except Exception as exc:
        return f"Spell-check skipped: spylls library not available ({exc})."

    norm = _sc.normalize_lang_code(target_lang)
    ctx["_spellcheck_lang"] = norm or target_lang
    if norm is None:
        return (f"Spell-check skipped: target language '{target_lang}' is not in the "
                f"bundled set ({', '.join(_sc.BUNDLED_LANGUAGES)}).")
    dictionary = _sc.get_dictionary(norm)
    if dictionary is None:
        reason = _sc.get_negative_reason(norm) or "dictionary unavailable"
        return f"Spell-check skipped: no dictionary for '{norm}' ({reason})."
    ctx["_spellcheck_dict"] = dictionary
    raw_ignore = cfg.get("spellcheck_ignore") or []
    if isinstance(raw_ignore, str):
        raw_ignore = [w for w in re.split(r"[\r\n,;]+", raw_ignore) if w.strip()]
    ctx["_spellcheck_ignore"] = {str(w).strip().lower() for w in raw_ignore if str(w).strip()}
    return None


def run_qa_checks(xml_content: bytes, filename: str,
                  config: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Run all enabled QA checks and return a structured report."""
    cfg = _normalize_config(config)
    fmt, source_lang, target_lang, segments = _extract_segments(xml_content, filename)

    enabled = cfg["enabled_checks"]
    ctx: Dict[str, Any] = {"format": fmt, "source_lang": source_lang, "target_lang": target_lang}
    notices: List[str] = []

    if "spellcheck" in enabled:
        notice = _prepare_spellcheck_context(target_lang, cfg, ctx)
        if notice:
            notices.append(notice)

    issues_by_cat: Dict[str, List[Dict[str, Any]]] = {cid: [] for cid in ALL_CHECK_IDS}

    for seg in segments:
        for cat_id, fn in _PER_SEGMENT_CHECKS:
            if cat_id not in enabled:
                continue
            for issue in fn(seg, cfg, ctx):
                issues_by_cat[cat_id].append(issue)

    # Cross-segment checks (Task #51) — run after the per-segment loop
    # since they need to compare segments against each other.
    if "inconsistent_translation" in enabled:
        for issue in _check_inconsistent_translation(segments, cfg, ctx):
            issues_by_cat["inconsistent_translation"].append(issue)
        # Task #71 — hyphenation inconsistency rides under the same
        # category with mismatch_axis="hyphenation".
        for issue in _check_hyphenation_inconsistency(segments, cfg, ctx):
            issues_by_cat["inconsistent_translation"].append(issue)

    # Task #71 — cross-segment date / number format inconsistency
    # extensions to the existing per-segment checks. Same category ids
    # so issues surface in the same QA cards.
    if "date_format_mismatch" in enabled:
        for issue in _check_date_format_inconsistency_cross(segments, cfg, ctx):
            issues_by_cat["date_format_mismatch"].append(issue)
    if "number_format_mismatch" in enabled:
        for issue in _check_number_format_inconsistency_cross(segments, cfg, ctx):
            issues_by_cat["number_format_mismatch"].append(issue)

    # Task #71 — surface a single notice when the polarity check was
    # skipped for segments whose source or target language has no
    # bundled negation list. Listed languages are the raw values read
    # from the file headers (so users can see exactly what we saw).
    if "polarity_mismatch" in enabled:
        unsupported = ctx.get("_polarity_unsupported_langs") or set()
        if unsupported:
            langs_sorted = ", ".join(sorted(unsupported))
            supported = ", ".join(sorted(_POLARITY_NEGATIONS.keys()))
            notices.append(
                f"Polarity check skipped for language(s) without a "
                f"bundled negation list: {langs_sorted}. Supported "
                f"languages: {supported}."
            )

    # Task #61 — surface a single summary notice if any custom regex
    # pattern timed out on at least one segment, so the user knows
    # something was silently skipped rather than failing.
    _timeouts = ctx.get("_custom_regex_timeout_segs") or set()
    if _timeouts:
        notices.append(
            f"{len(_timeouts)} segment(s) skipped one or more custom regex "
            f"patterns due to timeout (likely catastrophic backtracking)."
        )

    # Task #61 hardening: surface a notice when the optional ``regex``
    # module is unavailable AND the user supplied custom patterns, so the
    # silent loss of the enforced timeout guarantee is never invisible.
    if not _HAS_REGEX_MOD and (cfg.get("custom_forbidden_regex_patterns")
                               or cfg.get("custom_required_regex_patterns")):
        notices.append(
            "Custom regex timeout guard unavailable (the optional `regex` "
            "package is not installed); patterns ran without an enforced "
            "per-segment deadline."
        )

    # Build categories result preserving CATEGORY_METADATA order
    categories: Dict[str, Any] = {}
    total = high = low = 0
    for cat_id, meta in CATEGORY_METADATA.items():
        if cat_id not in enabled:
            continue
        issues = issues_by_cat.get(cat_id, [])
        if not issues:
            continue
        categories[cat_id] = {
            "label": meta["label"],
            "icon": meta["icon"],
            "group": meta["group"],
            "severity": meta["severity"],
            "issues": issues,
            "count": len(issues),
        }
        total += len(issues)
        if meta["severity"] == SEVERITY_HIGH:
            high += len(issues)
        else:
            low += len(issues)

    # Task #54 — lightweight segment payload for the HTML report's
    # "Show context" feature. Only id + cleaned source/target text;
    # XML elements / tag tuples are intentionally dropped so the
    # serialized JSON stays small enough to embed in the report.
    segments_data = [
        {
            "id": s["id"],
            "source": s.get("source_text") or "",
            "target": s.get("target_text") or "",
        }
        for s in segments
    ]

    return {
        "filename": filename,
        "format": fmt,
        "source_lang": source_lang,
        "target_lang": target_lang,
        "notices": notices,
        "segment_count": len(segments),
        "categories": categories,
        "segments_data": segments_data,
        "summary": {
            "total": total,
            "high": high,
            "low": low,
        },
    }


# -----------------------------------------------------------------------------
# Cleaned-file preparation (manual overrides only — auto-fix removed Task #61)
# -----------------------------------------------------------------------------

def _replace_target_with_plain_text(target_element, new_text: str) -> None:
    """Replace the entire content of a target element with plain text.

    Inline tags inside the original target are lost by design.
    """
    if target_element is None:
        return
    for child in list(target_element):
        target_element.remove(child)
    target_element.text = new_text or ""


def prepare_qa_download(xml_content: bytes, filename: str, *,
                        target_overrides: Optional[Dict[str, str]] = None) -> bytes:
    """Re-parse the original file, apply manual target overrides, and serialize back.

    Auto-fixes were removed in Task #61 — the translator does every change
    by hand via the inline editor (``target_overrides``). If no overrides
    are supplied, the file is round-tripped unchanged.

    Parameters
    ----------
    target_overrides
        Optional mapping ``{segment_id (str) -> new_target_text (str)}``.
        Overridden segments are replaced as plain text (inline tags inside
        the segment are lost — caller surfaces a warning to the user).
    """
    fmt = _detect_format(xml_content, filename)
    xml_norm = _normalize_xml_input(xml_content)
    parser = etree.XMLParser(remove_blank_text=False, strip_cdata=False)
    tree = etree.fromstring(xml_norm, parser=parser)

    overrides = {str(k): v for k, v in (target_overrides or {}).items()}

    if fmt == "tmx":
        _, target_lang = _autodetect_tmx_languages(tree)
        for idx, tu in enumerate(tree.xpath("//tu"), start=1):
            seg_id = str(idx)
            if seg_id in overrides:
                tgt_seg = _get_tmx_tuv_by_lang(tu, target_lang)
                _replace_target_with_plain_text(tgt_seg, overrides[seg_id])
    else:
        for idx, tu in enumerate(tree.xpath("//xliff:trans-unit", namespaces=XLIFF_NS), start=1):
            seg_id = str(idx)
            if seg_id in overrides:
                targets = tu.xpath("xliff:target", namespaces=XLIFF_NS)
                if targets:
                    _replace_target_with_plain_text(targets[0], overrides[seg_id])

    out = etree.tostring(tree, xml_declaration=True, encoding="UTF-8")
    out = out.replace(
        b"<?xml version='1.0' encoding='UTF-8'?>",
        b'<?xml version="1.0" encoding="UTF-8"?>',
    )
    return out


# -----------------------------------------------------------------------------
# Report export
# -----------------------------------------------------------------------------

def _flatten_issues(results: Dict[str, Any],
                    target_overrides: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    overrides = {str(k): v for k, v in (target_overrides or {}).items()}
    flat = []
    for cat_id, cat in results.get("categories", {}).items():
        for issue in cat["issues"]:
            seg_id = str(issue["segment_id"])
            original_target = issue["target"] or ""
            edited_target = overrides.get(seg_id, "")
            is_edited = (seg_id in overrides
                         and overrides[seg_id] != original_target)
            flat.append({
                "category_id": cat_id,
                "category": cat["label"],
                "group": cat["group"],
                "severity": issue["severity"],
                "segment": issue["segment_id"],
                "message": issue["message"],
                "note": issue.get("note", ""),
                "source": issue["source"],
                "target": original_target,
                "edited_target": edited_target if is_edited else "",
                "is_edited": is_edited,
                # Pass-through of the in-app highlight payload so the HTML
                # export can render the same red-background + underline as
                # the QA tab. ``span_target`` is either:
                #   - None / "" -> no highlight
                #   - a single str -> red block highlight (legacy)
                #   - a list[str] -> per-token highlight (spell-check)
                "span_source": issue.get("span_source"),
                "span_target": issue.get("span_target"),
                # Task #51 — for inconsistent_translation, the renderer
                # word-diffs the target against the reference target and
                # paints both cells with the Content group background.
                # Source cell does NOT get an internal diff (per spec the
                # source is by definition the same / near-same as the
                # reference). Empty / absent for every other category.
                "diff_against": issue.get("diff_against"),
                "reference_segment_id": issue.get("reference_segment_id"),
                "reference_source": issue.get("reference_source"),
                "reference_target": issue.get("reference_target"),
                "mismatch_axis": issue.get("mismatch_axis"),
                "sibling_segment_ids": issue.get("sibling_segment_ids"),
                # Tag-error grouping signature — see _make_issue.
                "tag_sig_missing": issue.get("tag_sig_missing"),
                "tag_sig_extra": issue.get("tag_sig_extra"),
            })
    flat.sort(key=lambda r: (r["segment"], r["category_id"]))
    return flat


def export_qa_report(results: Dict[str, Any], fmt: str = "csv", *,
                     target_overrides: Optional[Dict[str, str]] = None) -> bytes:
    """Export the QA results as CSV or HTML.

    target_overrides
        Optional mapping of ``segment_id`` (str) → edited target text. When
        provided, every row whose segment was manually edited in the QA tab
        will display both the original and the edited target text, and the
        HTML export will tag the row with an "EDITED" badge so reviewers can
        see at a glance which segments diverge from the QA snapshot. The
        function stays backward-compatible when called with no overrides.
    """
    fmt = (fmt or "csv").lower()
    rows = _flatten_issues(results, target_overrides=target_overrides)
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
        writer.writerow(["Segment", "Category", "Group", "Severity", "Message",
                         "Note", "Source", "Target", "Edited target"])
        for r in rows:
            writer.writerow([r["segment"], r["category"], r["group"], r["severity"],
                             r["message"], r.get("note", ""),
                             r["source"], r["target"],
                             r["edited_target"]])
        return buf.getvalue().encode("utf-8-sig")

    if fmt == "html":
        def esc(s: str) -> str:
            return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        def esc_attr(s: str) -> str:
            # Attribute-safe escaping: also escape both quote variants so a
            # crafted span/filename containing ' or " cannot break out of
            # the attribute and inject markup. Used everywhere a value is
            # interpolated inside an HTML attribute (data-term, data-key,
            # data-row, …) to harden the downloaded report against XSS.
            return (esc(s)
                    .replace('"', "&quot;")
                    .replace("'", "&#39;"))
        # Task #68 — severity colors / emojis are no longer rendered.
        # Kept as inert empty maps in case downstream extensions need them.
        sev_color = {}
        sev_emoji = {}

        # Mirror app.py's _qa_highlight: highlight color comes from the
        # category's group palette (5 groups, one color each — Task #58) so the HTML
        # report matches what the user saw in the QA tab. Spell-check uses
        # the same color but with an underline to disambiguate per-token
        # matches at a glance.
        #
        # Task #44: every painted span carries class="qa-span" and a
        # `data-token` attribute (the lowercased matched text). The
        # toolbar JS uses these for per-token selection ("click a
        # highlighted term, then Ignore") and for the un-paint/hide
        # logic at apply() time.
        def _highlight(text: str, span, category_id: Optional[str] = None) -> str:
            escaped = esc(text or "")
            if not span:
                return escaped
            bg, fg = get_highlight_color(category_id)

            def _wrap(matched: str, *, underline: bool) -> str:
                # `matched` is HTML-escaped (the regex ran against the
                # already-escaped text). We use the same escaped form
                # for the data-token value so JS comparisons stay
                # consistent — JS reads getAttribute('data-token'),
                # which decodes attribute escapes once, and matched
                # span textContent decodes the inner HTML once too, so
                # both produce the same string after a single decode.
                token_attr = esc_attr(matched.lower())
                deco = "text-decoration:underline;" if underline else ""
                # Task #65 — for double_spaces the matched payload is a run
                # of 2+ space characters (ASCII *or* a non-breaking / narrow /
                # thin / other Unicode space). HTML collapses consecutive
                # whitespace visually, so the painted span was invisible.
                # Replace each space with a middle dot `·` ONLY for display
                # (see dotify_double_space); the data-token keeps the real
                # spaces so the JS Ignore-by-token feature still works.
                display = matched
                if category_id == "double_spaces":
                    display = dotify_double_space(matched)
                return (f'<span class="qa-span" data-token="{token_attr}" '
                        f'style="background:{bg};color:{fg};'
                        f'font-weight:600;padding:0 2px;border-radius:2px;'
                        f'{deco}">{display}</span>')

            if isinstance(span, (list, tuple)):
                tokens = [s for s in dict.fromkeys(span) if s]
                if not tokens:
                    return escaped
                tokens.sort(key=len, reverse=True)
                # Per-category boundary — see get_highlight_boundary above.
                # Mirrors app.py:_qa_highlight so the HTML report matches
                # the in-app QA panel for short numeric tokens, IDs, URLs.
                left, right = get_highlight_boundary(category_id)
                pattern = re.compile(
                    left + r"(?:" +
                    "|".join(re.escape(esc(t)) for t in tokens) +
                    r")" + right,
                    re.IGNORECASE | re.UNICODE,
                )
                return pattern.sub(
                    lambda m: _wrap(m.group(0), underline=True), escaped)
            span_escaped = esc(str(span))
            if not span_escaped:
                return escaped
            # count=1: mark only the first occurrence so short spans (single
            # letter / digit / 2-3 char unit) don't paint incidental repeats
            # elsewhere in the segment. Mirrors app.py:_qa_highlight.
            pattern = re.compile(re.escape(span_escaped), re.IGNORECASE)
            # Some categories' span semantics are "last occurrence" rather
            # than "first" — e.g. final_punctuation_mismatch points at the
            # trailing '.' / '!' / '?', and a sentence like
            # "1.2.3.4." would otherwise paint the FIRST '.' instead of
            # the trailing one the user is being told about.
            if category_id in LAST_OCCURRENCE_CATEGORIES:
                hits = list(pattern.finditer(escaped))
                if not hits:
                    return escaped
                m = hits[-1]
                return (escaped[:m.start()]
                        + _wrap(m.group(0), underline=False)
                        + escaped[m.end():])
            return pattern.sub(
                lambda m: _wrap(m.group(0), underline=False), escaped, count=1)
        # Color legend (one swatch per group, in CATEGORY_METADATA order)
        # and HTML layout: table-layout:fixed + per-column widths via
        # <colgroup> + word-wrap:break-word so long URLs / IDs don't blow
        # the 7-column grid out of proportion.
        # Task #45: swatches double as group filter toggles. Each carries
        # a `data-group` attribute; clicking it toggles `qa-group-active`
        # and apply() hides rows whose data-group is not in the active
        # set. Multi-select = OR (union). When zero swatches are active,
        # all rows are visible (current behavior). The "Clear groups"
        # button next to the legend resets all selections at once.
        legend_swatches = []
        for _grp, (_bg, _fg) in GROUP_HIGHLIGHT_COLORS.items():
            legend_swatches.append(
                f'<button type="button" class="qa-legend-swatch" '
                f'data-group="{esc(_grp)}" '
                f'style="background:{_bg};color:{_fg};font-weight:600;'
                f'padding:2px 8px;border-radius:3px;margin:2px 4px 2px 0;'
                f'font-size:11px;border:2px solid transparent;cursor:pointer;'
                f'font-family:inherit;">{esc(_grp)}</button>'
            )
        legend_html = (
            '<div id="qa-legend" style="margin:10px 0 14px 0;padding:8px 12px;'
            'background:#f5f7f8;border:1px solid #e0e3e4;border-radius:6px;'
            'font-size:12px;display:flex;flex-wrap:wrap;align-items:center;gap:4px;">'
            '<strong style="color:#1a5488;margin-right:6px;">Highlight legend:</strong>'
            + "".join(legend_swatches)
            + '<span style="flex:1;"></span>'
            '<button type="button" data-qa-action="clear-groups" id="qa-clear-groups-btn" '
            'disabled title="No group filter active" '
            'style="background:#fff;color:#1a5488;border:1px solid #1a5488;'
            'padding:3px 10px;border-radius:4px;font-weight:600;font-size:11px;'
            'cursor:not-allowed;opacity:0.5;font-family:inherit;">Clear groups</button>'
            '</div>'
        )

        # Interactive toolbar (Task #44 — per-error selection).
        # The reviewer clicks any highlighted span inside a Source/Target
        # cell to select that exact term, then presses "Ignore" to dismiss
        # it. The Ignore button is disabled until something is selected.
        # apply() un-paints the ignored token everywhere it appears; rows
        # whose only highlighted content was the ignored token are hidden,
        # rows that still have other findings stay visible (just without
        # the ignored span painted) so other issues remain reviewable.
        # "Undo last" pops the most recent Ignore (LIFO). Choices are NOT
        # persisted across reopens — every fresh load starts with all
        # spans painted and visible. "Copy ignored terms" puts the
        # currently-ignored set on the clipboard. "Export filtered HTML"
        # downloads a clean copy that mirrors the on-screen state.
        toolbar_html = (
            '<div id="qa-toolbar" style="position:sticky;top:0;z-index:10;'
            'margin:10px 0 14px 0;padding:10px 14px;background:#1a5488;'
            'color:#fff;border-radius:6px;font-size:13px;'
            'display:flex;flex-wrap:wrap;align-items:center;gap:10px;">'
            '<strong style="font-size:14px;">Review:</strong>'
            '<span id="qa-counter" style="background:#0e7bc0;padding:3px 10px;'
            'border-radius:12px;font-weight:600;">0 / 0 visible</span>'
            '<span id="qa-selection" style="font-size:12px;font-style:italic;'
            'opacity:0.85;">No term selected</span>'
            '<span style="flex:1;"></span>'
            '<button type="button" data-qa-action="ignore" id="qa-ignore-btn" disabled '
            'title="Click a highlighted term first" '
            'style="background:#f4cce0;color:#a01060;border:none;padding:6px 12px;'
            'border-radius:4px;font-weight:600;cursor:not-allowed;'
            'opacity:0.5;pointer-events:none;">Ignore selected</button>'
            '<button type="button" data-qa-action="undo" id="qa-undo-btn" disabled '
            'style="background:#7cb4db;color:#130e45;border:none;padding:6px 12px;'
            'border-radius:4px;font-weight:600;cursor:pointer;'
            'opacity:0.5;pointer-events:none;">Undo last</button>'
            '<button type="button" data-qa-action="copy" '
            'style="background:#fff;color:#1a5488;border:1px solid #1a5488;'
            'padding:6px 12px;border-radius:4px;font-weight:600;cursor:pointer;">Copy ignored terms</button>'
            '<button type="button" data-qa-action="export" '
            'style="background:#fff;color:#1a5488;border:1px solid #1a5488;'
            'padding:6px 12px;border-radius:4px;font-weight:600;cursor:pointer;">Export filtered HTML</button>'
            # Sort toggle — when OFF rows render in original order (by
            # segment id). When ON (default), qa-rows are re-grouped by
            # tag/group following CATEGORY_METADATA order (Content →
            # Numeric Elements → Tags → Terminology → Spelling), and
            # inside each group by the user-approved category order.
            # The toggle only re-arranges the DOM; it stacks cleanly
            # with Ignore and Group filters which just hide rows.
            # Task #68 — renamed from "Sort by severity"; HIGH/LOW
            # bucketing dropped (severity no longer surfaced in the UI).
            '<button type="button" data-qa-action="sort" id="qa-sort-btn" '
            'aria-pressed="true" '
            'title="Group rows by tag (Content → Numeric Elements → Tags → Terminology → Spelling)" '
            'style="background:#0e7bc0;color:#fff;border:1px solid #0e7bc0;'
            'padding:6px 12px;border-radius:4px;font-weight:600;cursor:pointer;'
            'box-shadow:inset 0 0 0 2px #fff;">'
            'Sort by groups ✓</button>'
            # Task #54 — context selector. Sets how many segments before/
            # after a flagged segment the per-row context-icon button reveals.
            # Lives in the toolbar (not the per-row button) so the
            # reviewer picks the depth once and applies it everywhere.
            '<label for="qa-context-n" style="font-size:11px;opacity:0.9;'
            'margin-left:6px;">Context:</label>'
            '<select id="qa-context-n" '
            'style="background:#fff;color:#1a5488;border:none;padding:5px 8px;'
            'border-radius:4px;font-weight:600;cursor:pointer;font-family:inherit;'
            'font-size:12px;">'
            '<option value="1">±1</option>'
            '<option value="2" selected>±2</option>'
            '<option value="3">±3</option>'
            '<option value="5">±5</option>'
            '</select>'
            '</div>'
            '<div id="qa-help" style="margin:0 0 12px 0;font-size:11px;color:#5e5f6b;">'
            'Pick a highlight + <strong style="color:#a01060;">Ignore selected</strong> to hide it everywhere · '
            '<strong style="color:#1a5488;">Undo last</strong> reverts · '
            'click legend colors to filter by group (<strong style="color:#1a5488;">Clear groups</strong> to reset) · '
            'click the <strong style="color:#1a5488;">context icon</strong> next to a segment to see its neighbors. '
            'All choices reset on reopen.'
            '</div>'
        )
        parts = [
            "<!doctype html><html><head><meta charset='utf-8'><title>QA Report</title>",
            "<style>",
            "body{font-family:Arial,sans-serif;background:#fff;color:#222;padding:20px;}",
            "table{border-collapse:collapse;width:100%;font-size:13px;table-layout:fixed;word-wrap:break-word;overflow-wrap:anywhere;white-space:normal;}",
            "th,td{border:1px solid #bcbdbe;padding:6px 8px;text-align:left;vertical-align:top;"
            "word-wrap:break-word;overflow-wrap:anywhere;white-space:normal;}",
            "th{background:#e0e3e4;color:#1a5488;}",
            "tr:nth-child(even){background:#f5f7f8;}",
            ".edited-badge{display:inline-block;background:#1a5488;color:#fff;padding:1px 6px;"
            "border-radius:3px;font-size:11px;font-weight:700;margin-left:6px;}",
            ".tgt-original{color:#888;text-decoration:line-through;display:block;margin-bottom:4px;}",
            ".tgt-edited{color:#1a5488;font-weight:600;display:block;}",
            ".tgt-label{display:inline-block;font-size:10px;color:#1a5488;font-weight:700;"
            "text-transform:uppercase;margin-right:4px;}",
            # Glossary reviewer notes (column C of the user-supplied
            # glossary). Lighter / smaller than the main message so it
            # reads as auxiliary context without competing for attention.
            ".qa-note{margin-top:4px;padding:4px 8px;background:#fff8e1;"
            "border-left:3px solid #f0b429;border-radius:3px;font-size:11px;"
            "color:#5e5f6b;line-height:1.35;}",
            # Task #44: clickable highlighted spans. The reviewer clicks a
            # span to select it, then uses the toolbar's Ignore button. A
            # selected span gets a dashed dark outline so the choice is
            # unambiguous before pressing Ignore.
            ".qa-span{cursor:pointer;}",
            ".qa-span:hover{outline:1px dashed #130e45;outline-offset:1px;}",
            ".qa-span.qa-selected{outline:2px dashed #130e45;outline-offset:2px;"
            "box-shadow:0 0 0 2px rgba(19,14,69,0.15);}",
            # Task #45: legend swatches as filter toggles. Inactive = thin
            # transparent border (so size doesn't jump on toggle); active =
            # solid dark border + subtle shadow + checkmark prefix via
            # ::before, matching the qa-selected visual language.
            ".qa-legend-swatch:hover{outline:1px dashed #130e45;outline-offset:1px;}",
            ".qa-legend-swatch.qa-group-active{border-color:#130e45 !important;"
            "box-shadow:0 0 0 2px rgba(19,14,69,0.15);}",
            ".qa-legend-swatch.qa-group-active::before{content:'\\2713  ';font-weight:700;}",
            # Task #54 — "Show context" UI. The per-row button sits in
            # the Segment cell next to the segment number. When toggled,
            # JS inserts a sibling row with a small inner table listing
            # the N segments before, the flagged one, and N after.
            # Visual icon button: a small SVG of three
            # stacked horizontal bars with the middle one highlighted in
            # the brand blue, immediately readable as "this segment +
            # its neighbors". Hover lifts the border + adds a soft
            # shadow; active state inverts colors so the user sees at a
            # glance which row is currently expanded.
            ".qa-ctx-btn{background:#fff;border:1px solid #bcbdbe;"
            "border-radius:4px;padding:2px 4px;margin-left:6px;"
            "cursor:pointer;line-height:0;font-family:inherit;"
            "vertical-align:middle;display:inline-flex;align-items:center;"
            "justify-content:center;width:22px;height:18px;"
            "transition:background .15s,border-color .15s,box-shadow .15s;}",
            # pointer-events:none on the icon so clicks always land on
            # the <button> itself (defense in depth — the JS handler
            # also walks up via closest('[data-qa-action]')).
            ".qa-ctx-btn svg,.qa-ctx-btn svg *{display:block;pointer-events:none;}",
            ".qa-ctx-btn .qa-ctx-bar-side{fill:#7cb4db;}",
            ".qa-ctx-btn .qa-ctx-bar-mid{fill:#1a5488;}",
            ".qa-ctx-btn:hover{border-color:#1a5488;"
            "box-shadow:0 1px 3px rgba(26,84,136,0.25);}",
            ".qa-ctx-btn:hover .qa-ctx-bar-side{fill:#5e98c4;}",
            ".qa-ctx-btn.qa-ctx-active{background:#1a5488;border-color:#134277;}",
            ".qa-ctx-btn.qa-ctx-active .qa-ctx-bar-side{fill:#7cb4db;}",
            ".qa-ctx-btn.qa-ctx-active .qa-ctx-bar-mid{fill:#fff;}",
            # Task #69 — tag-error group summary row. Tan-tinted left
            # border picks up the Tags palette so the summary reads as
            # "a Tags-group rollup". The "Show all" button is the same
            # pill-button as the per-row context icon for visual
            # consistency. `qa-tag-child` rows start hidden via inline
            # style="display:none" set in Python and are toggled by JS.
            ".qa-tag-group-summary{background:#fff8ef;}",
            ".qa-tag-group-summary > td{border-left:3px solid #6b4423;}",
            ".qa-tag-summary-hint{margin-top:6px;}",
            ".qa-tag-show-btn{background:#e8d4b8;color:#6b4423;"
            "border:1px solid #c8a878;border-radius:3px;padding:2px 8px;"
            "font-size:11px;font-weight:600;cursor:pointer;"
            "font-family:inherit;}",
            ".qa-tag-show-btn:hover{background:#dcc29c;}",
            ".qa-tag-child{background:#fffaf3;}",
            ".qa-context-row > td.qa-ctx-cell{background:#f5f7f8;padding:6px 12px;}",
            ".qa-ctx-inner{width:100%;font-size:12px;border:1px solid #d0d3d4;"
            "border-collapse:collapse;table-layout:fixed;}",
            ".qa-ctx-inner col.qa-ctx-col-id{width:6%;}",
            ".qa-ctx-inner col.qa-ctx-col-text{width:47%;}",
            ".qa-ctx-inner th{background:#e0e3e4;color:#5e5f6b;padding:4px 6px;"
            "font-size:11px;text-align:left;}",
            ".qa-ctx-inner td{padding:4px 6px;color:#5e5f6b;font-style:italic;"
            "vertical-align:top;word-wrap:break-word;overflow-wrap:anywhere;}",
            ".qa-ctx-inner td.qa-ctx-id{font-style:normal;color:#1a5488;font-weight:600;}",
            # Task #54 — flagged segment row sits between the before/
            # after siblings and is visually marked so the reviewer
            # can immediately see where the issue lives in the local
            # narrative. Solid Content-group background, dark border,
            # bold non-italic text + arrow indicator in the id cell.
            ".qa-ctx-inner tr.qa-ctx-flagged > td{background:#ffe8e8;color:#1a1a1a;"
            "font-style:normal;font-weight:600;border-top:2px solid #c0392b;"
            "border-bottom:2px solid #c0392b;}",
            ".qa-ctx-inner tr.qa-ctx-flagged > td.qa-ctx-id{color:#c0392b;"
            "white-space:nowrap;}",
            ".qa-ctx-inner tr.qa-ctx-flagged > td.qa-ctx-id::after{"
            "content:' \\2190';font-weight:700;font-size:10px;"
            "color:#c0392b;letter-spacing:0.3px;}",
            "@media print{#qa-toolbar,#qa-help,#qa-clear-groups-btn,.qa-ctx-btn{display:none !important;} "
            ".qa-span,.qa-legend-swatch{cursor:default;outline:none !important;"
            "box-shadow:none !important;}}",
            "h1{color:#1a5488;}",
            "</style></head><body>",
            f"<h1>QA Report: {esc(results.get('filename', ''))}</h1>",
            f"<p><strong>Format:</strong> {esc(results.get('format', ''))}",
            f" &middot; <strong>Source:</strong> {esc(results.get('source_lang', ''))}",
            f" &middot; <strong>Target:</strong> {esc(results.get('target_lang', ''))}",
            f" &middot; <strong>Segments:</strong> {results.get('segment_count', 0)}</p>",
            toolbar_html,
            legend_html,
        ]
        s = results.get("summary", {})
        edited_segs = sorted({str(r["segment"]) for r in rows if r["is_edited"]},
                             key=lambda x: (len(x), x))
        # Task #68 — severity breakdown dropped from the report. Show the
        # plain issue total so the reviewer still has a headline figure.
        totals_line = f"<p><strong>Total:</strong> {s.get('total', len(rows))} issue(s)"
        if edited_segs:
            totals_line += (
                f" &middot; <span class='edited-badge'>EDITED</span> "
                f"{len(edited_segs)} segment(s) manually edited"
            )
        totals_line += "</p>"
        parts.append(totals_line)
        # Task #68 — dropped the Level/Severity column (HIGH/LOW are no
        # longer surfaced in the UI). 5 columns now, widths re-balanced to
        # total 100%: Seg 4 / Grp 5 / Msg 22 / Src 34 / Tgt 35.
        # Group cell renders "Num. Elements" instead of "Numeric Elements"
        # so the 5% column fits the longest label without wrapping; the
        # internal group name (data-group, GROUP_RANK, palette, filter UI)
        # is unchanged.
        parts.append(
            "<table id='qa-table'>"
            "<colgroup>"
            "<col style='width:4%'/>"   # Segment (narrow — only id + ctx icon)
            "<col style='width:5%'/>"   # Group (short labels; "Numeric Elements" rendered as "Num. Elements")
            "<col style='width:22%'/>"  # Message
            "<col style='width:34%'/>"  # Source
            "<col style='width:35%'/>"  # Target
            "</colgroup>"
            "<thead><tr><th>Segment</th>"
            "<th>Group</th><th>Message</th><th>Source</th><th>Target</th>"
            "</tr></thead><tbody>"
        )

        # Task #69 — tag-error grouping. Repeated `tag_count`,
        # `tag_order`, `placeholders` issues that share the same
        # (missing, extra) signature collapse into a single summary
        # row with a lazy "Show all ▾" expander. Children are stashed
        # in `qa-tag-groups-data` and only injected into the DOM when
        # the reviewer clicks the button — keeps the initial paint
        # tight on big reports while preserving full detail on demand.
        _TAG_GROUPABLE = {"tag_count", "tag_order", "placeholders"}
        _tag_sig_count: Dict[Tuple, int] = {}
        for r in rows:
            cid = r.get("category_id")
            if cid not in _TAG_GROUPABLE:
                continue
            sig = (cid, r.get("tag_sig_missing"), r.get("tag_sig_extra"))
            _tag_sig_count[sig] = _tag_sig_count.get(sig, 0) + 1
        # Groups with ≥ 2 collapse. Assigned a stable group id by
        # first-encounter order so the JSON payload keys are simple.
        _tag_group_id: Dict[Tuple, int] = {}
        _tag_groups_data: Dict[str, List[str]] = {}
        _tag_group_seen: Dict[Tuple, bool] = {}
        for sig, n in _tag_sig_count.items():
            if n >= 2:
                _tag_group_id[sig] = len(_tag_group_id)

        def _row_html(idx, r):
            # Local capture of the per-row rendering (extracted from
            # the original inline loop) so it can be reused both for
            # singleton rows in the main pass and for collapsed
            # children stashed in the JSON payload.
            ctx_icon_svg = (
                "<svg width='12' height='12' viewBox='0 0 12 12' "
                "aria-hidden='true' focusable='false'>"
                "<rect class='qa-ctx-bar-side' x='1' y='1' width='10' "
                "height='2' rx='1'/>"
                "<rect class='qa-ctx-bar-mid' x='1' y='5' width='10' "
                "height='2' rx='1'/>"
                "<rect class='qa-ctx-bar-side' x='1' y='9' width='10' "
                "height='2' rx='1'/>"
                "</svg>"
            )
            seg_cell = (
                f"{r['segment']}"
                f"<button type='button' class='qa-ctx-btn' "
                f"data-qa-action='ctx' data-row='{idx}' "
                f"data-segment='{r['segment']}' "
                f"aria-label='Show surrounding segments' "
                f"title='Show surrounding segments'>{ctx_icon_svg}</button>"
            )
            cid = r.get("category_id")
            if (cid == "inconsistent_translation"
                    and r.get("reference_segment_id") is not None):
                content_bg, content_fg = GROUP_HIGHLIGHT_COLORS["Content"]
                axis = r.get("mismatch_axis") or "target"
                ref_src_txt = r.get("reference_source") or ""
                ref_tgt_txt = r.get("reference_target") or ""
                this_src_txt = r["source"] or ""
                this_tgt_txt = r["target"] or ""
                ref_id = r.get("reference_segment_id")
                if axis == "source":
                    ref_src_html, this_src_html = highlight_diff(
                        ref_src_txt, this_src_txt, css_class="qa-span")
                    ref_tgt_html = esc(ref_tgt_txt)
                    this_tgt_html = esc(this_tgt_txt)
                elif axis == "both":
                    ref_src_html, this_src_html = highlight_diff(
                        ref_src_txt, this_src_txt, css_class="qa-span")
                    ref_tgt_html, this_tgt_html = highlight_diff(
                        ref_tgt_txt, this_tgt_txt, css_class="qa-span")
                else:
                    ref_src_html = esc(ref_src_txt)
                    this_src_html = esc(this_src_txt)
                    ref_tgt_html, this_tgt_html = highlight_diff(
                        ref_tgt_txt, this_tgt_txt, css_class="qa-span")

                def _stack(ref_html: str, this_html: str) -> str:
                    return (
                        f'<div style="background:{content_bg};'
                        f'border-left:3px solid {content_fg};'
                        f'padding:6px 8px;border-radius:3px;">'
                        f'<div style="font-size:0.72rem;color:#5e5f6b;'
                        f'font-weight:600;margin-bottom:2px;">seg {ref_id}</div>'
                        f'<div style="margin-bottom:6px;">{ref_html}</div>'
                        f'<div style="font-size:0.72rem;color:#5e5f6b;'
                        f'font-weight:600;margin-bottom:2px;">seg {r["segment"]}</div>'
                        f'<div>{this_html}</div>'
                        f'</div>'
                    )
                source_cell = _stack(ref_src_html, this_src_html)
                highlighted_target = _stack(ref_tgt_html, this_tgt_html)
            else:
                source_cell = _highlight(r["source"], r.get("span_source"), cid)
                highlighted_target = _highlight(r["target"], r.get("span_target"), cid)
            local_seg_cell = seg_cell
            if r["is_edited"]:
                local_seg_cell = seg_cell + "<span class='edited-badge'>EDITED</span>"
                target_cell = (
                    f"<span class='tgt-label'>Original</span>"
                    f"<span class='tgt-original'>{highlighted_target}</span>"
                    f"<span class='tgt-label'>Edited</span>"
                    f"<span class='tgt-edited'>{esc(r['edited_target'])}</span>"
                )
            else:
                target_cell = highlighted_target
            return (
                f"<tr class='qa-row' data-row='{idx}' data-segment='{r['segment']}' "
                f"data-group=\"{esc(r['group'])}\" "
                f"data-group-rank='{GROUP_RANK.get(r['group'], 99)}' "
                f"data-cat-rank='{CATEGORY_RANK.get(r.get('category_id', ''), 999)}'>"
                f"<td>{local_seg_cell}</td>"
                f"<td>{esc('Num. Elements' if r['group'] == 'Numeric Elements' else r['group'])}</td>"
                f"<td>{esc(r['message'])}"
                f"{('<div class=' + chr(34) + 'qa-note' + chr(34) + '><strong>Note:</strong> ' + esc(r['note']) + '</div>') if r.get('note') else ''}"
                f"</td>"
                f"<td>{source_cell}</td>"
                f"<td>{target_cell}</td>"
                f"</tr>"
            )

        def _summary_message(r, count):
            cid = r.get("category_id")
            miss = r.get("tag_sig_missing") or ()
            extra = r.get("tag_sig_extra") or ()
            def _fmt(seq):
                return ", ".join(seq) if seq else "∅"
            if cid == "tag_order":
                head = (f"Tag pairing is broken in the target: source [{_fmt(miss)}] vs "
                        f"target [{_fmt(extra)}]")
            elif cid == "placeholders":
                bits = []
                if miss: bits.append(f"missing {_fmt(miss)}")
                if extra: bits.append(f"unexpected {_fmt(extra)}")
                head = "Placeholders differ" + (": " + " · ".join(bits) if bits else "")
            else:  # tag_count
                bits = []
                if miss: bits.append(f"missing {_fmt(miss)} in target")
                if extra: bits.append(f"unexpected {_fmt(extra)} in target")
                head = "Tag count differs" + (": " + " · ".join(bits) if bits else "")
            return f"{head} — {count} segments affected."

        for idx, r in enumerate(rows):
            cid = r.get("category_id")
            sig = (cid, r.get("tag_sig_missing"), r.get("tag_sig_extra"))
            gid = _tag_group_id.get(sig) if cid in _TAG_GROUPABLE else None
            if gid is not None:
                # Stash the child row HTML for lazy injection. The
                # main pass emits only the summary row, at first
                # occurrence of the signature.
                _tag_groups_data.setdefault(str(gid), []).append(
                    _row_html(idx, r))
                if _tag_group_seen.get(sig):
                    continue
                _tag_group_seen[sig] = True
                summary_msg = _summary_message(r, _tag_sig_count[sig])
                # Summary row: same data-row / data-segment as the
                # first child so context/sort/filter behave naturally.
                # Adds qa-tag-group-summary + data-tag-group so JS can
                # find children + Show all button can toggle them.
                btn = (
                    f"<button type='button' class='qa-tag-show-btn' "
                    f"data-qa-action='tag-show' data-tag-group='{gid}' "
                    f"data-count='{_tag_sig_count[sig]}'>"
                    f"Show all {_tag_sig_count[sig]} ▾</button>"
                )
                parts.append(
                    f"<tr class='qa-row qa-tag-group-summary' "
                    f"data-row='{idx}' data-segment='{r['segment']}' "
                    f"data-group=\"{esc(r['group'])}\" "
                    f"data-group-rank='{GROUP_RANK.get(r['group'], 99)}' "
                    f"data-cat-rank='{CATEGORY_RANK.get(cid, 999)}' "
                    f"data-tag-group='{gid}'>"
                    f"<td>{r['segment']} +{_tag_sig_count[sig]-1}</td>"
                    f"<td>{esc('Num. Elements' if r['group'] == 'Numeric Elements' else r['group'])}</td>"
                    f"<td>{esc(summary_msg)}<div class='qa-tag-summary-hint'>{btn}</div></td>"
                    f"<td>{_highlight(r['source'], r.get('span_source'), cid)}</td>"
                    f"<td>{_highlight(r['target'], r.get('span_target'), cid)}</td>"
                    f"</tr>"
                )
                continue
            parts.append(_row_html(idx, r))
        # Emit the lazy children payload (always; empty dict when
        # there are no collapsed groups).
        _tg_json = json.dumps(_tag_groups_data, ensure_ascii=False).replace("</", "<\\/")
        parts.append(
            "<script id='qa-tag-groups-data' type='application/json'>"
            + _tg_json
            + "</script>"
        )
        parts.append("</tbody></table>")

        # Task #54 — embed cleaned segments payload as JSON so the
        # toolbar's "Show context" feature can fetch neighbors entirely
        # client-side. Escape `</` so a stray "</script>" inside any
        # segment text can't terminate the script tag early.
        segments_data = results.get("segments_data") or []
        segments_json = (json.dumps(segments_data, ensure_ascii=False)
                         .replace("</", "<\\/"))
        parts.append(
            "<script id='qa-segments-data' type='application/json'>"
            + segments_json
            + "</script>"
        )

        # Task #44 — per-error Ignore. The reviewer clicks any
        # `.qa-span` inside Source/Target to select a specific token,
        # then presses the toolbar Ignore button. apply() walks every
        # row, uses each cell's snapshot to re-derive the visible HTML,
        # un-paints spans whose data-token is in `ignored`, and hides
        # any row whose source+target ended up with zero remaining
        # painted spans (i.e. that row's only finding was the ignored
        # term). Rows that still have other findings stay visible.
        # No browser storage — every reopen starts clean. Undo is a
        # LIFO stack of the active session's ignore actions.
        js = r"""
<script>
(function(){
  var ignored = {};       // tokenKey (from data-token) -> human label
  var undoStack = [];     // LIFO of {key, term}
  var selected = null;    // currently-selected tokenKey
  // Task #45: legend group filter. When non-empty, only rows whose
  // data-group is in this set stay visible. Empty = no group filter
  // (current behavior — all rows visible). Multi-select is OR/union.
  var activeGroups = {};  // groupName -> true

  // Sort-by-tags toggle. false = original DOM order (by segment id).
  // true = qa-rows re-grouped by tag/group (Content → Numeric Elements
  // → Tags → Terminology → Spelling) and inside each group by the
  // category order baked into data-cat-rank. Context rows follow their
  // parent. Task #68 — default ON so the tag-grouped view leads from
  // first paint; HIGH/LOW bucketing dropped along with the severity UI.
  var sortByTags = true;
  // Cached original ordering captured once at startup. Each entry is
  // a <tr.qa-row> in render order. Used to restore when the toggle
  // goes back to OFF and as the stable ordering source for sorting.
  var origOrder = [];

  // Indices of the Source / Target cells inside each row (Seg, Grp,
  // Msg, Src, Tgt — Level column dropped in Task #68).
  var SRC_IDX = 3, TGT_IDX = 4;

  // Task #54: load the embedded segments_data payload once. Used by
  // the per-row context-icon toggle to render the N segments before/after
  // a flagged segment without round-tripping anything.
  var SEG_BY_ID = {};
  try {
    var dataNode = document.getElementById('qa-segments-data');
    if (dataNode) {
      var arr = JSON.parse(dataNode.textContent || '[]');
      for (var k = 0; k < arr.length; k++) {
        SEG_BY_ID[arr[k].id] = arr[k];
      }
    }
  } catch (e) { SEG_BY_ID = {}; }
  function escHtml(s){
    return String(s == null ? '' : s)
      .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
      .replace(/"/g,'&quot;').replace(/'/g,'&#39;');
  }
  function truncText(s, n){
    s = String(s == null ? '' : s);
    if (s.length <= n) return s;
    return s.slice(0, n) + '\u2026';
  }

  // Snapshot original innerHTML of every Source/Target cell so apply()
  // can restore + re-derive on every change without losing un-painted
  // tokens when the user hits Undo. Keyed by data-row (stable row id)
  // so apply() keeps working after reorder() shuffles DOM order.
  var snap = {};

  function rows(){ return document.querySelectorAll('#qa-table tbody tr.qa-row'); }

  function snapshot(){
    snap = {};
    origOrder = [];
    rows().forEach(function(tr){
      var cells = tr.children;
      var key = tr.getAttribute('data-row');
      snap[key] = {
        src: cells[SRC_IDX] ? cells[SRC_IDX].innerHTML : '',
        tgt: cells[TGT_IDX] ? cells[TGT_IDX].innerHTML : ''
      };
      // origOrder only stores the row reference (in original render
      // order). The matching context row is resolved fresh at reorder
      // time so rows opened *after* snapshot also follow their parent.
      origOrder.push(tr);
    });
  }

  // Re-arrange the table body to either the original order (OFF) or
  // grouped by severity (ON). Uses tbody.appendChild() which MOVES
  // existing nodes (preserves event listeners + snapshot mapping by
  // data-row + ignore/group state). Context rows opened via the
  // per-row context-icon are looked up live via data-context-of so
  // rows opened after snapshot still follow their parent.
  function reorder(){
    if (!origOrder.length) return;
    var tbody = origOrder[0].parentNode;
    if (!tbody) return;
    var seq = origOrder.slice();
    if (sortByTags) {
      seq.sort(function(a, b){
        // Task #68 — primary key: tag/group rank (Content → Numeric
        // Elements → Tags → Terminology → Spelling). Within a group,
        // sort by data-cat-rank (CATEGORY_METADATA insertion order) so
        // the per-group order matches the Configuration expander.
        var ga = parseInt(a.getAttribute('data-group-rank'), 10);
        var gb = parseInt(b.getAttribute('data-group-rank'), 10);
        if (isNaN(ga)) ga = 99;
        if (isNaN(gb)) gb = 99;
        if (ga !== gb) return ga - gb;
        var ca = parseInt(a.getAttribute('data-cat-rank'), 10);
        var cb = parseInt(b.getAttribute('data-cat-rank'), 10);
        if (isNaN(ca)) ca = 999;
        if (isNaN(cb)) cb = 999;
        if (ca !== cb) return ca - cb;
        // Stable within (group, category): keep original render order
        // (segment id ascending).
        return origOrder.indexOf(a) - origOrder.indexOf(b);
      });
    }
    // Task #69 — children of a tag-group summary follow it inside
    // reorder() so they stay adjacent regardless of where the summary
    // ends up after a sort toggle. Children themselves drop out of
    // the sort sequence (filtered above).
    seq = seq.filter(function(tr){
      return !(tr.classList && tr.classList.contains('qa-tag-child'));
    });
    seq.forEach(function(tr){
      tbody.appendChild(tr);
      var key = tr.getAttribute('data-row');
      if (key) {
        var ctx = tbody.querySelector(
          'tr.qa-context-row[data-context-of="' + key + '"]');
        if (ctx) tbody.appendChild(ctx);
      }
      // Re-attach this summary's children right after it so the
      // expanded view stays clustered under its summary.
      if (tr.classList && tr.classList.contains('qa-tag-group-summary')) {
        var gid = tr.getAttribute('data-tag-group') || '';
        if (gid !== '') {
          var kids = tbody.querySelectorAll(
            'tr.qa-tag-child[data-tag-group="' + gid + '"]');
          kids.forEach(function(k){ tbody.appendChild(k); });
        }
      }
    });
  }

  // Returns {html, remaining}: html with any .qa-span whose data-token
  // is currently ignored replaced by its plain text content; remaining
  // = number of .qa-span elements still painted afterwards.
  function strip(html){
    var tmp = document.createElement('div');
    tmp.innerHTML = html;
    var spans = tmp.querySelectorAll('.qa-span');
    var remaining = 0;
    spans.forEach(function(s){
      var tok = s.getAttribute('data-token') || '';
      if (tok && Object.prototype.hasOwnProperty.call(ignored, tok)) {
        s.parentNode.replaceChild(document.createTextNode(s.textContent), s);
      } else {
        remaining++;
      }
    });
    return {html: tmp.innerHTML, remaining: remaining};
  }

  function apply(){
    var total = 0, visible = 0;
    var hasIgnored = Object.keys(ignored).length > 0;
    var groupKeys = Object.keys(activeGroups);
    var hasGroupFilter = groupKeys.length > 0;
    rows().forEach(function(tr){
      total++;
      var key = tr.getAttribute('data-row');
      var s = snap[key] || {src:'', tgt:''};
      var cells = tr.children;
      var src = strip(s.src);
      var tgt = strip(s.tgt);
      if (cells[SRC_IDX]) cells[SRC_IDX].innerHTML = src.html;
      if (cells[TGT_IDX]) cells[TGT_IDX].innerHTML = tgt.html;
      // Hide when EITHER (a) ignored-term rule strips all spans on a
      // row that originally had highlights, OR (b) Task #45 group
      // filter is active and this row's group is not in the active
      // set. The two filters stack — a row must pass both to show.
      var hadHighlights = (s.src.indexOf('qa-span') !== -1)
                        || (s.tgt.indexOf('qa-span') !== -1);
      var hideByIgnore = hasIgnored && hadHighlights
                        && (src.remaining + tgt.remaining) === 0;
      var grp = tr.getAttribute('data-group') || '';
      var hideByGroup = hasGroupFilter
                        && !Object.prototype.hasOwnProperty.call(activeGroups, grp);
      // Task #69 — keep collapsed tag-group children hidden across
      // apply() refreshes. Without this, ignore/group-filter clicks
      // would silently re-show children of any group whose summary is
      // still collapsed (`data-tag-open != "1"`).
      var hideByCollapse = false;
      if (tr.classList && tr.classList.contains('qa-tag-child')) {
        var gid = tr.getAttribute('data-tag-group') || '';
        var sum = document.querySelector(
          'tr.qa-tag-group-summary[data-tag-group="' + gid + '"]');
        if (!sum || sum.getAttribute('data-tag-open') !== '1') {
          hideByCollapse = true;
        }
      }
      var hide = hideByIgnore || hideByGroup || hideByCollapse;
      tr.style.display = hide ? 'none' : '';
      // Task #54: a context row sits as a sibling <tr> next to its
      // qa-row. When the qa-row is hidden by Ignore/Group, the context
      // row must follow suit so the layout stays consistent. Resolve
      // by stable data-row (not loop index) so this keeps working
      // after the Sort-by-severity toggle re-orders the table.
      if (key) {
        var ctxRow = document.querySelector(
          'tr.qa-context-row[data-context-of="' + key + '"]');
        if (ctxRow) ctxRow.style.display = hide ? 'none' : '';
      }
      if (!hide) visible++;
    });
    // Re-mark the currently-selected token (if any of its spans remain).
    if (selected) {
      var stillThere = false;
      document.querySelectorAll('.qa-span').forEach(function(s){
        if ((s.getAttribute('data-token') || '') === selected) {
          s.classList.add('qa-selected');
          stillThere = true;
        }
      });
      if (!stillThere) selected = null;
    }
    var counter = document.getElementById('qa-counter');
    if (counter) counter.textContent = visible + ' / ' + total + ' visible';
    var sel = document.getElementById('qa-selection');
    if (sel) {
      if (selected) {
        var sample = document.querySelector('.qa-span.qa-selected');
        var label = sample ? sample.textContent : selected;
        if (label.length > 40) label = label.slice(0, 40) + '\u2026';
        sel.textContent = 'Selected: \u201c' + label + '\u201d';
        sel.style.fontStyle = 'normal';
        sel.style.opacity = '1';
      } else {
        sel.textContent = 'No term selected';
        sel.style.fontStyle = 'italic';
        sel.style.opacity = '0.85';
      }
    }
    var ignoreBtn = document.getElementById('qa-ignore-btn');
    if (ignoreBtn) {
      var noSel = !selected;
      ignoreBtn.disabled = noSel;
      ignoreBtn.style.opacity = noSel ? '0.5' : '1';
      ignoreBtn.style.pointerEvents = noSel ? 'none' : 'auto';
      ignoreBtn.style.cursor = noSel ? 'not-allowed' : 'pointer';
      ignoreBtn.title = noSel ? 'Click a highlighted term first' : '';
    }
    var undoBtn = document.getElementById('qa-undo-btn');
    if (undoBtn) {
      var empty = undoStack.length === 0;
      undoBtn.disabled = empty;
      undoBtn.style.opacity = empty ? '0.5' : '1';
      undoBtn.style.pointerEvents = empty ? 'none' : 'auto';
      undoBtn.style.cursor = empty ? 'not-allowed' : 'pointer';
    }
    // Task #45: sync the Clear groups button + swatch active classes.
    var clearBtn = document.getElementById('qa-clear-groups-btn');
    if (clearBtn) {
      clearBtn.disabled = !hasGroupFilter;
      clearBtn.style.opacity = hasGroupFilter ? '1' : '0.5';
      clearBtn.style.cursor = hasGroupFilter ? 'pointer' : 'not-allowed';
      clearBtn.title = hasGroupFilter
        ? 'Reset group filter'
        : 'No group filter active';
    }
    document.querySelectorAll('.qa-legend-swatch').forEach(function(b){
      var g = b.getAttribute('data-group') || '';
      if (Object.prototype.hasOwnProperty.call(activeGroups, g)) {
        b.classList.add('qa-group-active');
      } else {
        b.classList.remove('qa-group-active');
      }
    });
  }

  function clearSelection(){
    document.querySelectorAll('.qa-span.qa-selected').forEach(function(s){
      s.classList.remove('qa-selected');
    });
    selected = null;
  }

  document.addEventListener('click', function(ev){
    var t = ev.target;
    if (!t || !t.getAttribute) return;

    // Task #45: legend swatch click → toggle that group in the active
    // filter set. Multi-select OR. Independent of the per-term Ignore
    // selection (clicking a swatch never touches `selected`).
    var swatch = t.closest ? t.closest('.qa-legend-swatch') : null;
    if (swatch) {
      var grp = swatch.getAttribute('data-group') || '';
      if (!grp) return;
      ev.preventDefault();
      if (Object.prototype.hasOwnProperty.call(activeGroups, grp)) {
        delete activeGroups[grp];
      } else {
        activeGroups[grp] = true;
      }
      apply();
      return;
    }

    // Span click → toggle selection. Use closest() so a click on a
    // child element inside the span (none today, but future-proof)
    // still resolves to the span itself.
    var span = t.closest ? t.closest('.qa-span') : null;
    if (span) {
      var tok = span.getAttribute('data-token') || '';
      if (!tok) return;
      ev.preventDefault();
      if (selected === tok) {
        clearSelection();
      } else {
        clearSelection();
        selected = tok;
        // Mark every visible occurrence so the user sees their pick.
        document.querySelectorAll('.qa-span').forEach(function(s){
          if ((s.getAttribute('data-token') || '') === tok) {
            s.classList.add('qa-selected');
          }
        });
      }
      apply();
      return;
    }

    // Walk up to the nearest element carrying data-qa-action so that
    // clicks landing on inner DOM (e.g. the <svg>/<rect> inside the
    // ctx button, or any future nested icon) still resolve to the
    // owning button. Without this, target===<rect> returns no action
    // and the click silently no-ops — that was the "sometimes opens,
    // sometimes doesn't" lag the reviewer reported.
    var actEl = t.closest ? t.closest('[data-qa-action]') : null;
    if (!actEl) return;
    t = actEl;
    var act = t.getAttribute('data-qa-action');
    if (!act) return;

    if (act === 'ctx') {
      // Task #54 — toggle the per-row context block. Click again on
      // the same button removes it. Selecting a different row leaves
      // existing context blocks untouched (multi-open is fine).
      ev.preventDefault();
      var rowIdx = t.getAttribute('data-row');
      var segId = parseInt(t.getAttribute('data-segment') || '0', 10);
      if (!rowIdx || !segId) return;
      var existing = document.querySelector(
        'tr.qa-context-row[data-context-of="' + rowIdx + '"]');
      if (existing) {
        existing.parentNode.removeChild(existing);
        t.classList.remove('qa-ctx-active');
        return;
      }
      var sel = document.getElementById('qa-context-n');
      var n = parseInt((sel && sel.value) || '2', 10);
      // Allowed depths exposed in the toolbar: 1 / 2 / 3 / 5. Anything
      // else (missing toolbar in the exported clone, manual DOM edit,
      // browser default) falls back to 2.
      if ([1,2,3,5].indexOf(n) === -1) n = 2;
      var flagged = SEG_BY_ID[segId] || null;
      var before = [], after = [];
      for (var i = segId - n; i < segId; i++) {
        if (SEG_BY_ID[i]) before.push(SEG_BY_ID[i]);
      }
      for (var j = segId + 1; j <= segId + n; j++) {
        if (SEG_BY_ID[j]) after.push(SEG_BY_ID[j]);
      }
      if (!flagged && !before.length && !after.length) {
        // Edge case: only one segment in the file. Render a tiny
        // notice so the click doesn't look broken.
        var notice = document.createElement('tr');
        notice.className = 'qa-context-row';
        notice.setAttribute('data-context-of', rowIdx);
        notice.innerHTML = "<td colspan='5' class='qa-ctx-cell' "
          + "style='font-style:italic;color:#888;'>"
          + "No surrounding segments available.</td>";
        var hostRow0 = document.querySelector(
          'tr.qa-row[data-row="' + rowIdx + '"]');
        if (hostRow0 && hostRow0.parentNode) {
          hostRow0.parentNode.insertBefore(notice, hostRow0.nextSibling);
          t.classList.add('qa-ctx-active');
        }
        return;
      }
      function ctxRowHtml(s, label){
        return '<tr class="qa-ctx-' + label + '">'
          + '<td class="qa-ctx-id">' + s.id + '</td>'
          + '<td>' + escHtml(truncText(s.source, 400)) + '</td>'
          + '<td>' + escHtml(truncText(s.target, 400)) + '</td>'
          + '</tr>';
      }
      var inner = '<table class="qa-ctx-inner">'
        + '<colgroup><col class="qa-ctx-col-id"/>'
        + '<col class="qa-ctx-col-text"/><col class="qa-ctx-col-text"/></colgroup>'
        + '<thead><tr><th>#</th><th>Source</th><th>Target</th></tr></thead>'
        + '<tbody>';
      for (var bi = 0; bi < before.length; bi++) inner += ctxRowHtml(before[bi], 'before');
      // Flagged segment sits between before/after with a visual marker
      // (qa-ctx-flagged class — see CSS) so the reviewer can locate
      // the issue inside the local narrative at a glance.
      if (flagged) inner += ctxRowHtml(flagged, 'flagged');
      for (var ai = 0; ai < after.length; ai++) inner += ctxRowHtml(after[ai], 'after');
      inner += '</tbody></table>';
      var ctxTr = document.createElement('tr');
      ctxTr.className = 'qa-context-row';
      ctxTr.setAttribute('data-context-of', rowIdx);
      ctxTr.innerHTML = "<td colspan='5' class='qa-ctx-cell'>" + inner + "</td>";
      var hostRow = document.querySelector(
        'tr.qa-row[data-row="' + rowIdx + '"]');
      if (hostRow && hostRow.parentNode) {
        hostRow.parentNode.insertBefore(ctxTr, hostRow.nextSibling);
        t.classList.add('qa-ctx-active');
      }
      return;
    }

    if (act === 'tag-show') {
      // Task #69 — lazy expand of a tag-error signature group.
      // Children HTML is stashed in `#qa-tag-groups-data`; on first
      // click we parse, inject the new rows after the summary, and
      // snapshot their Src/Tgt cells so Ignore/group-filter keep
      // working. On subsequent clicks we just toggle visibility.
      ev.preventDefault();
      var gid = t.getAttribute('data-tag-group');
      if (gid == null) return;
      var summary = t.closest ? t.closest('tr.qa-tag-group-summary') : null;
      if (!summary) return;
      var alreadyOpen = summary.getAttribute('data-tag-open') === '1';
      if (!alreadyOpen) {
        if (!summary.getAttribute('data-tag-injected')) {
          var dataNode = document.getElementById('qa-tag-groups-data');
          var groups = {};
          try { groups = JSON.parse(dataNode ? (dataNode.textContent || '{}') : '{}'); }
          catch (e) { groups = {}; }
          var childHtmls = groups[gid] || [];
          if (childHtmls.length) {
            var tmp = document.createElement('tbody');
            tmp.innerHTML = childHtmls.join('');
            // Walk forward; each new child is inserted right after the
            // current anchor so the final DOM order matches the JSON
            // payload order.
            var nodes = Array.prototype.slice.call(tmp.children);
            var anchor = summary;
            for (var i = 0; i < nodes.length; i++) {
              var child = nodes[i];
              child.classList.add('qa-tag-child');
              child.setAttribute('data-tag-group', gid);
              summary.parentNode.insertBefore(child, anchor.nextSibling);
              anchor = child;
              // Snapshot the new row so apply() can re-derive its
              // highlighted Src/Tgt on every refresh.
              var cells2 = child.children;
              var key2 = child.getAttribute('data-row');
              if (key2) {
                snap[key2] = {
                  src: cells2[SRC_IDX] ? cells2[SRC_IDX].innerHTML : '',
                  tgt: cells2[TGT_IDX] ? cells2[TGT_IDX].innerHTML : ''
                };
                origOrder.push(child);
              }
            }
          }
          summary.setAttribute('data-tag-injected', '1');
        }
        // Reveal all children of this group.
        document.querySelectorAll(
          'tr.qa-tag-child[data-tag-group="' + gid + '"]'
        ).forEach(function(c){ c.style.display = ''; });
        summary.setAttribute('data-tag-open', '1');
        var n = t.getAttribute('data-count') || '';
        t.textContent = 'Hide ' + n + ' \u25b4';
      } else {
        document.querySelectorAll(
          'tr.qa-tag-child[data-tag-group="' + gid + '"]'
        ).forEach(function(c){ c.style.display = 'none'; });
        summary.setAttribute('data-tag-open', '0');
        var n2 = t.getAttribute('data-count') || '';
        t.textContent = 'Show all ' + n2 + ' \u25be';
      }
      apply();
      return;
    }

    if (act === 'ignore') {
      if (!selected) return;
      var k = selected;
      var sample = document.querySelector('.qa-span.qa-selected');
      var term = sample ? sample.textContent : k;
      if (!Object.prototype.hasOwnProperty.call(ignored, k)) {
        ignored[k] = term;
        undoStack.push({key: k, term: term});
      }
      clearSelection();
    } else if (act === 'undo') {
      var last = undoStack.pop();
      if (last) delete ignored[last.key];
    } else if (act === 'clear-groups') {
      // Task #45: reset all group-filter selections at once. Does NOT
      // touch the per-term Ignore set or the undo stack — group filter
      // and term-Ignore are independent.
      activeGroups = {};
    } else if (act === 'copy') {
      // Copy the active ignored set in the order they were dismissed.
      var seen = {};
      var active = [];
      undoStack.forEach(function(e){
        if (Object.prototype.hasOwnProperty.call(ignored, e.key) && !seen[e.key]) {
          seen[e.key] = true;
          active.push(ignored[e.key] || e.term || e.key);
        }
      });
      var txt = active.join('\n');
      var label = active.length
        ? 'Copied ' + active.length + ' term(s)'
        : 'Nothing ignored yet';
      function flash(){
        t.textContent = label;
        setTimeout(function(){ t.textContent = 'Copy ignored terms'; }, 1800);
      }
      if (navigator.clipboard && navigator.clipboard.writeText) {
        navigator.clipboard.writeText(txt).then(flash, flash);
      } else {
        var ta = document.createElement('textarea');
        ta.value = txt; document.body.appendChild(ta); ta.select();
        try { document.execCommand('copy'); } catch(e) {}
        document.body.removeChild(ta);
        flash();
      }
      return;
    } else if (act === 'export') {
      // Build a filtered HTML clone that mirrors what the reviewer
      // sees: hidden rows removed, un-painted tokens stripped.
      var clone = document.documentElement.cloneNode(true);
      var tb = clone.querySelector('#qa-toolbar'); if (tb) tb.remove();
      var help = clone.querySelector('#qa-help'); if (help) help.remove();
      var hasIg = Object.keys(ignored).length > 0;
      clone.querySelectorAll('#qa-table tbody tr.qa-row').forEach(function(tr){
        var cells = tr.children;
        var hadHighlights = false;
        var rem = 0;
        [cells[SRC_IDX], cells[TGT_IDX]].forEach(function(c){
          if (!c) return;
          var spans = c.querySelectorAll('.qa-span');
          if (spans.length) hadHighlights = true;
          spans.forEach(function(s){
            var tok = s.getAttribute('data-token') || '';
            if (tok && Object.prototype.hasOwnProperty.call(ignored, tok)) {
              s.parentNode.replaceChild(
                clone.ownerDocument.createTextNode(s.textContent), s);
            } else {
              rem++;
            }
          });
        });
        if (hasIg && hadHighlights && rem === 0) tr.remove();
      });
      // Strip span interactivity from the exported file (no JS in clone).
      clone.querySelectorAll('.qa-span').forEach(function(s){
        s.classList.remove('qa-selected');
        s.style.cursor = 'default';
      });
      // Task #54 — preserve the context feature in the exported file
      // so the recipient can still expand/collapse surrounding
      // segments. We KEEP the per-row context-icon buttons, the embedded
      // qa-segments-data JSON block, and the toggle script. Other
      // toolbar interactivity (Ignore / Undo / etc.) is already gone
      // because we removed #qa-toolbar above; the click handler is
      // harmless without those buttons. Strip ONLY non-data scripts
      // would risk breaking the JSON payload, so we leave all scripts
      // intact.
      var html = '<!doctype html>' + clone.outerHTML;
      var blob = new Blob([html], {type:'text/html'});
      var url = URL.createObjectURL(blob);
      var a = document.createElement('a');
      a.href = url; a.download = 'QA_Report_filtered.html';
      document.body.appendChild(a); a.click();
      setTimeout(function(){ URL.revokeObjectURL(url); a.remove(); }, 100);
      return;
    } else if (act === 'sort') {
      // Toggle tag/group sorting. Pure DOM re-arrangement — does not
      // touch ignored / activeGroups / snap, so Ignore + Group filters
      // keep working unchanged after the reorder.
      sortByTags = !sortByTags;
      var sb = document.getElementById('qa-sort-btn');
      if (sb) {
        sb.setAttribute('aria-pressed', sortByTags ? 'true' : 'false');
        if (sortByTags) {
          sb.style.background = '#0e7bc0';
          sb.style.color = '#fff';
          sb.style.border = '1px solid #0e7bc0';
          sb.style.boxShadow = 'inset 0 0 0 2px #fff';
          sb.textContent = 'Sort by groups ✓';
        } else {
          sb.style.background = '#fff';
          sb.style.color = '#1a5488';
          sb.style.border = '1px solid #1a5488';
          sb.style.boxShadow = 'none';
          sb.textContent = 'Sort by groups';
        }
      }
      reorder();
      return;
    } else {
      return;
    }
    apply();
  });

  snapshot();
  // Task #68 — default-ON tag sort: apply the initial reorder so the
  // first paint matches the button's pressed state.
  if (sortByTags) reorder();
  apply();
})();
</script>
"""
        parts.append(js)
        parts.append("</body></html>")
        return "".join(parts).encode("utf-8")

    raise ValueError(f"Unsupported report format: {fmt}")

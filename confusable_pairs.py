"""Confusable-word pairs (real-word errors) — user-supplied only.

Hunspell's spell-check only flags words that are not in the dictionary.
That misses *real-word errors* — words that are spelled correctly but
used in place of a similar one. This module parses user-supplied
confusable groups (textarea or .xlsx) so the QA pipeline can surface
each occurrence as a Warning for the translator to confirm.

Public API
----------
- ``CONFUSABLE_PAIRS``: ``Dict[str, List[Tuple[str, ...]]]`` — empty
  dict; kept for backward compatibility. No built-in language packs
  are shipped; users supply groups via textarea or .xlsx.
- ``SUPPORTED_LANGUAGES``: empty list, kept for backward compatibility.
- ``get_groups_for_language(lang_code)`` — always returns ``[]`` (kept
  callable so external imports don't break).
- ``parse_custom_pairs(text)`` — parse the textarea custom format
  (``palabra1|palabra2|palabra3`` per line, ``#`` comments allowed).
- ``parse_pairs_xlsx(file_bytes)`` — parse an .xlsx where each row is
  one confusable group (one column per member).
"""
from __future__ import annotations

import io
import re
from typing import Dict, List, Tuple


CONFUSABLE_PAIRS: Dict[str, List[Tuple[str, ...]]] = {}

SUPPORTED_LANGUAGES: List[str] = []


def get_groups_for_language(lang_code: str) -> List[Tuple[str, ...]]:
    """No built-in groups. Always returns []. Kept for backward compat."""
    return []


_CUSTOM_SEP_RE = re.compile(r"[|/]")


def parse_custom_pairs(text: str) -> List[Tuple[str, ...]]:
    """Parse the custom-pairs textarea.

    Format: one group per line, members separated by ``|`` or ``/``.
    Lines starting with ``#`` and blank lines are ignored. Members are
    stripped; groups with fewer than 2 distinct members are dropped.
    """
    if not text:
        return []
    out: List[Tuple[str, ...]] = []
    seen = set()
    for raw in text.replace("\r", "\n").split("\n"):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        members = [m.strip() for m in _CUSTOM_SEP_RE.split(line) if m.strip()]
        # Dedupe inside the group case-insensitively but preserve original casing.
        uniq: List[str] = []
        seen_inner = set()
        for m in members:
            k = m.lower()
            if k not in seen_inner:
                seen_inner.add(k)
                uniq.append(m)
        if len(uniq) < 2:
            continue
        key = tuple(sorted(k for k in seen_inner))
        if key in seen:
            continue
        seen.add(key)
        out.append(tuple(uniq))
    return out


def parse_pairs_xlsx(file_bytes: bytes, filename: str = "") -> List[Tuple[str, ...]]:
    """Parse an .xlsx workbook where each row is a confusable group.

    First sheet only, one column per member, empty cells skipped, rows
    with fewer than 2 non-empty cells dropped. Header rows are NOT
    auto-detected here (the format is positional). The ``filename`` arg
    is accepted but unused — kept for forward-compat with call-sites.
    Uses openpyxl (already a project dep). Legacy ``.xls`` (BIFF) is not
    supported; modern Excel exports ``.xlsx`` by default.
    """
    if not file_bytes:
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.active
    if ws is None:
        return []
    out: List[Tuple[str, ...]] = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        members: List[str] = []
        seen_inner = set()
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            k = s.lower()
            if k in seen_inner:
                continue
            seen_inner.add(k)
            members.append(s)
        if len(members) < 2:
            continue
        key = tuple(sorted(seen_inner))
        if key in seen:
            continue
        seen.add(key)
        out.append(tuple(members))
    return out

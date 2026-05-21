import streamlit as st
import zipfile
import io
import re
import difflib
import time
import hashlib
from datetime import date
from typing import List, Dict
from anonymizer import MQXLIFFAnonymizer, load_dictionary_terms
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import qa_checker

st.set_page_config(
    page_title="Anonymizer",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded"
)

CUSTOM_CSS = """
<style>
    .main {
        background-color: #ffffff !important;
    }
    .stApp {
        background-color: #ffffff !important;
    }
    .stApp, .stApp p, .stApp span, .stApp label, .stApp div {
        color: #130e45 !important;
    }
    
    /* Global scrollbar styles */
    *::-webkit-scrollbar {
        width: 16px !important;
        height: 16px !important;
    }
    *::-webkit-scrollbar-track {
        background: #d0d3d4 !important;
        border-radius: 8px !important;
    }
    *::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0) !important;
        border-radius: 8px !important;
        border: 2px solid #d0d3d4 !important;
    }
    *::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488) !important;
    }
    
    /* Firefox scrollbar */
    * {
        scrollbar-width: auto;
        scrollbar-color: #1a5488 #d0d3d4;
    }
    [data-testid="stSidebar"] {
        background-color: #e0e3e4 !important;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] span, [data-testid="stSidebar"] label {
        color: #130e45 !important;
    }
    h1, h2, h3 {
        color: #1a5488 !important;
    }
    .stButton > button,
    [data-testid="stBaseButton-primary"] {
        background-color: #0e7bc0 !important;
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 800;
    }
    .stButton > button p,
    .stButton > button span,
    .stButton > button div,
    [data-testid="stBaseButton-primary"] p,
    [data-testid="stBaseButton-primary"] span {
        color: white !important;
        background-color: transparent !important;
        padding: 0 !important;
    }
    [data-testid="stBaseButton-primary"] {
        font-weight: 900 !important;
        font-size: 1.1rem !important;
    }
    .stButton > button:hover,
    [data-testid="stBaseButton-primary"]:hover {
        background-color: #134277 !important;
    }
    .stButton > button:disabled,
    .stButton > button[disabled],
    .stDownloadButton > button:disabled,
    .stDownloadButton > button[disabled],
    [data-testid="stBaseButton-primary"]:disabled,
    [data-testid="stBaseButton-secondary"]:disabled {
        background-color: #d9dcdd !important;
        color: #8a8d92 !important;
        cursor: not-allowed !important;
        opacity: 0.7 !important;
    }
    .stButton > button:disabled p,
    .stButton > button:disabled span,
    .stButton > button:disabled div,
    .stDownloadButton > button:disabled p,
    .stDownloadButton > button:disabled span,
    .stDownloadButton > button:disabled div,
    [data-testid="stBaseButton-primary"]:disabled p,
    [data-testid="stBaseButton-primary"]:disabled span,
    [data-testid="stBaseButton-secondary"]:disabled p,
    [data-testid="stBaseButton-secondary"]:disabled span {
        color: #8a8d92 !important;
    }
    .stButton > button:disabled:hover,
    .stDownloadButton > button:disabled:hover,
    [data-testid="stBaseButton-primary"]:disabled:hover,
    [data-testid="stBaseButton-secondary"]:disabled:hover {
        background-color: #d9dcdd !important;
    }
    .stat-card {
        background-color: #e0e3e4;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        border-left: 4px solid #1a5488;
    }
    .stat-card-safe-regex { border-left-color: #0e7bc0; }
    .stat-card-regex-ct { border-left-color: #6f42c1; }
    .stat-card-presidio { border-left-color: #e67e22; }
    .stat-card-biomedical { border-left-color: #e74c3c; }
    .stat-card-proper-names { border-left-color: #28a745; }
    .stat-card-dictionary { border-left-color: #17a2b8; }
    .stat-number {
        font-size: 2rem;
        font-weight: bold;
        color: #130e45;
    }
    .stat-label {
        color: #5e5f6b;
        font-size: 0.9rem;
    }
    .sidebar-divider {
        border: none;
        border-top: 1px solid #c0c3c4;
        margin: 0.8rem 0;
    }
    .app-footer {
        text-align: center;
        padding: 1.5rem 0 0.5rem 0;
        color: #8a8b96 !important;
        font-size: 0.8rem;
        border-top: 1px solid #e0e3e4;
        margin-top: 2rem;
    }
    .app-footer p, .app-footer span {
        color: #8a8b96 !important;
    }
    .preview-box {
        background-color: #f8f9fa;
        border: 1px solid #bcbdbe;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .preview-container {
        max-height: 500px;
        overflow-y: auto;
        padding-right: 10px;
        margin: 1rem 0;
        border: 1px solid #e0e3e4;
        border-radius: 8px;
        background-color: #fafafa;
    }
    .preview-container::-webkit-scrollbar {
        width: 18px;
    }
    .preview-container::-webkit-scrollbar-track {
        background: #d0d3d4;
        border-radius: 9px;
        box-shadow: inset 0 0 3px rgba(0,0,0,0.2);
    }
    .preview-container::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 9px;
        border: 3px solid #d0d3d4;
        box-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    .preview-container::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    div[data-testid="stExpander"] {
        max-height: 600px;
        overflow-y: auto;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar {
        width: 20px;
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-track {
        background: #c8cbcc;
        border-radius: 10px;
        box-shadow: inset 0 0 4px rgba(0,0,0,0.25);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #1a5488, #0e7bc0);
        border-radius: 10px;
        border: 3px solid #c8cbcc;
        box-shadow: 0 2px 5px rgba(0,0,0,0.35);
    }
    div[data-testid="stExpander"]::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #134277, #1a5488);
    }
    .exclude-badge {
        background-color: #dc3545;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 4px;
        font-size: 0.75rem;
        font-weight: bold;
        margin-left: 0.5rem;
    }
    .excluded-segment {
        border: 2px dashed #dc3545 !important;
        background-color: #fff0f0 !important;
    }
    .before-text {
        color: #dc3545;
        background-color: #ffe6e6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .after-text {
        color: #28a745;
        background-color: #e6ffe6;
        padding: 0.5rem;
        border-radius: 4px;
    }
    .section-header {
        background-color: #7cb4db;
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #e0e3e4;
        border-left: 4px solid #0e7bc0;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    div[data-testid="stFileUploader"] {
        background-color: #f8f9fa;
        border: 2px dashed #7cb4db;
        border-radius: 10px;
        padding: 1rem;
    }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def count_words(text: str) -> int:
    """Count words in text, ignoring XML tags and whitespace."""
    if not text:
        return 0
    import re
    clean = re.sub(r'<[^>]+>', ' ', text)
    clean = re.sub(r'\s+', ' ', clean).strip()
    words = [w for w in clean.split() if len(w) > 0]
    return len(words)


def segment_word_count(preview: dict) -> int:
    """Returns the max word count between source and target."""
    source_words = count_words(preview.get('source_before', ''))
    target_words = count_words(preview.get('target_before', ''))
    return max(source_words, target_words)


def is_junk_segment(preview: dict, min_words_junk: int = 3) -> bool:
    """Detects junk/short original segments that pollute TM databases.
    Returns True if the segment should be excluded."""
    import re
    source = preview.get('source_before', '').strip()
    target = preview.get('target_before', '').strip()
    
    if not source and not target:
        return True
    
    for text in [source, target]:
        if not text:
            continue
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r'\s+', ' ', clean).strip()
        if not clean:
            continue
        if re.fullmatch(r'[\d\s\-.,;:!?¿¡()\[\]{}/\\|@#$%^&*+=<>~`"\'°ºª•–—…\u2022\u2013\u2014\u2026]+', clean):
            return True
        words = [w for w in clean.split() if len(w) > 0]
        if len(words) < min_words_junk:
            return True
    
    return False


def render_stat_card(label: str, value: int, col, css_class: str = ""):
    with col:
        st.markdown(f"""
        <div class="stat-card {css_class}">
            <div class="stat-number">{value}</div>
            <div class="stat-label">{label}</div>
        </div>
        """, unsafe_allow_html=True)


def _normalize_lang_code(code: str) -> str:
    """Normalize language code to standard format (e.g., de-de -> de-DE, en -> en)."""
    parts = code.strip().split("-")
    if len(parts) == 2:
        return f"{parts[0].lower()}-{parts[1].upper()}"
    return parts[0].lower()


def strip_inline_tags(text: str) -> str:
    if not text:
        return text
    cleaned = re.sub(r'<[^>]+>', '', text)
    cleaned = re.sub(r'\{/?(\d+)?\}', '', cleaned)
    cleaned = re.sub(r'  +', ' ', cleaned)
    return cleaned.strip()


def _html_escape(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# Single source of truth for the word-level diff renderer lives in
# qa_checker so both the Anonymizer Duplicates tab and the QA tab's
# `inconsistent_translation` card paint the same way (Task #51).
from qa_checker import highlight_diff  # noqa: E402


def _qa_highlight_search(html: str, pattern) -> str:
    """Overlay the user's QA search query on top of already-rendered HTML.

    ``html`` is the output of :func:`_qa_highlight` (so it may already
    contain ``<span>`` wrappers for the QA finding). ``pattern`` is the
    compiled regex from the search bar (or ``None`` / ``"invalid"`` when
    no search is active). Matches inside text nodes are wrapped in a
    yellow ``<mark>`` so the reviewer sees exactly where their query
    landed without losing the QA finding's color.

    The implementation walks the HTML by splitting on tag boundaries
    (``<…>``) and only paints inside text segments — so we never inject
    ``<mark>`` inside an attribute value or break an existing tag.
    """
    if not pattern or pattern == "invalid" or not html:
        return html
    wrap = (
        '<mark style="background:#fff59d;color:inherit;'
        'padding:0 1px;border-radius:2px;">{m}</mark>'
    )
    # First split on tag boundaries so we never paint inside attributes.
    parts = re.split(r"(<[^>]+>)", html)
    for i, part in enumerate(parts):
        if not part or part.startswith("<"):
            continue
        # Then split each text node on HTML entities (`&amp;`, `&lt;`, …)
        # so the regex never lands inside an entity reference and breaks
        # it (e.g. searching for "amp" or "&" must not corrupt `&amp;`).
        sub = re.split(r"(&[A-Za-z]+;|&#\d+;|&#x[0-9A-Fa-f]+;)", part)
        for j, chunk in enumerate(sub):
            if not chunk or (chunk.startswith("&") and chunk.endswith(";")):
                continue
            sub[j] = pattern.sub(lambda m: wrap.format(m=m.group(0)), chunk)
        parts[i] = "".join(sub)
    return "".join(parts)


def _qa_highlight(text: str, span=None, category_id: str = None) -> str:
    """HTML-escape ``text`` and wrap any occurrence of ``span`` in a highlight.

    ``span`` may be:
      * ``None`` / empty   -> plain escaped text.
      * a single ``str``   -> block highlight using the category's group
        color (Task #36: 8 group colors, one per QA group).
      * a ``list[str]``    -> per-token block highlight + underline (used
        by spell-check). Same group color as the block highlight; the
        extra underline disambiguates per-token matches at a glance.

    ``category_id`` selects the highlight color via
    :func:`qa_checker.get_highlight_color`. When omitted, falls back to
    the Content red palette so legacy callers keep their old look.
    """
    escaped = _html_escape(text or "")
    if not span:
        return escaped
    bg, fg = qa_checker.get_highlight_color(category_id)
    block_tpl = (
        f'<span style="background:{bg};color:{fg};font-weight:600;'
        f'padding:0 2px;border-radius:2px;">{{w}}</span>'
    )
    wavy_tpl = (
        f'<span style="background:{bg};color:{fg};font-weight:600;'
        f'padding:0 2px;border-radius:2px;text-decoration:underline;">{{w}}</span>'
    )
    # Task #65 — double_spaces highlight: ASCII spaces collapse in HTML,
    # so the painted span was invisible. Render the matched payload with
    # each space swapped for a middle dot `·` so the reviewer sees
    # exactly where (and how many) extra spaces live in the target.
    def _format_match(matched: str) -> str:
        if category_id == "double_spaces":
            return matched.replace(" ", "·")
        return matched
    if isinstance(span, (list, tuple)):
        tokens = [s for s in dict.fromkeys(span) if s]  # ordered dedupe
        if not tokens:
            return escaped
        tokens.sort(key=len, reverse=True)
        # Per-category boundary so short tokens don't paint inside larger
        # ones — e.g. "2" inside "28005" (number_mismatch), "AB1" inside
        # "AB12" (alphanum_id_mismatch), "http://a.com" inside
        # "http://a.com/path" (urls). spellcheck keeps its letter-only
        # boundary so accents/digit-suffixed words behave as expected.
        # See qa_checker.get_highlight_boundary for the rule.
        left, right = qa_checker.get_highlight_boundary(category_id)
        pattern = re.compile(
            left + r"(?:" +
            "|".join(re.escape(_html_escape(t)) for t in tokens) +
            r")" + right,
            re.IGNORECASE | re.UNICODE,
        )
        return pattern.sub(lambda m: wavy_tpl.format(w=_format_match(m.group(0))), escaped)
    span_escaped = _html_escape(span)
    if not span_escaped:
        return escaped
    # Mark only the FIRST occurrence (count=1). A global sub would also paint
    # incidental repeats — e.g. span="t" highlighting every "t" in "título del
    # estudio", or span="2" highlighting the "2" of "28005" in "2 March, CP
    # 28005". Checks paint the offending span and it almost always coincides
    # with the first occurrence (leading letter, differing number, unique ID).
    pattern = re.compile(re.escape(span_escaped), re.IGNORECASE)
    # Some checks (e.g. final_punctuation_mismatch) point at the LAST
    # occurrence of the span — the trailing '.' / '!' / '?'. Painting
    # the first match in a string like "1.2.3.4." would highlight the
    # wrong character. See qa_checker.LAST_OCCURRENCE_CATEGORIES.
    if category_id in qa_checker.LAST_OCCURRENCE_CATEGORIES:
        hits = list(pattern.finditer(escaped))
        if not hits:
            return escaped
        m = hits[-1]
        return (escaped[:m.start()]
                + block_tpl.format(w=_format_match(m.group(0)))
                + escaped[m.end():])
    return pattern.sub(lambda m: block_tpl.format(w=_format_match(m.group(0))), escaped, count=1)


def _qa_legend_html() -> str:
    """Render the Task #36 highlight-color legend (one swatch per QA group).

    Used at the top of the QA results panel so users can see at a glance
    which color belongs to which QA group.
    """
    swatches = []
    for grp, (bg, fg) in qa_checker.GROUP_HIGHLIGHT_COLORS.items():
        swatches.append(
            f'<span style="display:inline-block;background:{bg};'
            f'color:{fg};font-weight:600;padding:2px 8px;border-radius:3px;'
            f'margin:2px 4px 2px 0;font-size:0.78rem;">{_html_escape(grp)}</span>'
        )
    return (
        '<div style="margin:6px 0 10px 0;padding:8px 12px;background:#f5f7f8;'
        'border:1px solid #e0e3e4;border-radius:6px;font-size:0.85rem;">'
        '<strong style="color:#1a5488;">Highlight legend:</strong> '
        + "".join(swatches) + '</div>'
    )


_QA_SEG_FILTER_RE = __import__("re").compile(r"^[\d,\-\s]+$")


def _parse_segment_filter(query: str):
    """Return a set of segment IDs if *query* is a numeric segment-id filter
    (e.g. "12", "12,15,20", "10-20", "5, 8-11, 30"), else ``None`` so the
    caller falls back to a free-text search.

    Accepts digits, commas, hyphens (ranges) and whitespace; everything else
    means "treat as text search". Invalid ranges (start > end, empty parts)
    are silently dropped.
    """
    if not query:
        return None
    q = query.strip()
    if not q or not _QA_SEG_FILTER_RE.match(q):
        return None
    ids = set()
    for chunk in q.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            parts = [p.strip() for p in chunk.split("-") if p.strip()]
            if len(parts) != 2 or not all(p.isdigit() for p in parts):
                continue
            a, b = int(parts[0]), int(parts[1])
            if a > b:
                a, b = b, a
            ids.update(range(a, b + 1))
        elif chunk.isdigit():
            ids.add(int(chunk))
    return ids or None


def _qa_total_badge(count: int) -> str:
    """Render a neutral total-issue badge for the QA summary line.

    Task #68 dropped the HIGH/LOW severity surface: the per-card and
    summary-line pills are replaced by a single total. Brand-blue pill so
    it visually matches the rest of the QA toolbar without re-introducing
    a severity color.
    """
    return (f'<span style="background:#cce0ff;color:#1a5488;'
            f'border:1px solid #7cb4db;'
            f'padding:2px 10px;border-radius:12px;font-size:0.78rem;font-weight:700;'
            f'letter-spacing:0.02em;">TOTAL · {count}</span>')


def _qa_reset_segment_edit(seg_id):
    """Drop every per-card edit widget for one segment + its override entry.

    Used as an `on_click` callback so the widget keys are removed BEFORE
    the next script run instantiates the corresponding `st.text_area`s. On
    the next render every editor for this segment falls back to its
    original `value=` (the segment's untouched target text).

    Per-card widget keys follow the pattern ``qa_edit_{seg_id}__{card_uid}``
    so a single segment may have multiple sibling keys when the same
    segment surfaces in several QA categories.
    """
    prefix = f"qa_edit_{seg_id}__"
    for k in [k for k in list(st.session_state.keys()) if k.startswith(prefix)]:
        st.session_state.pop(k, None)
    overrides = st.session_state.get("qa_target_overrides", {})
    overrides.pop(str(seg_id), None)
    st.session_state["qa_target_overrides"] = overrides


def _qa_reset_all_edits():
    """Clear every inline-edit widget value and the entire override map."""
    for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
        st.session_state.pop(k, None)
    st.session_state["qa_target_overrides"] = {}



def _render_qa_check_tab():
    st.markdown("### 🛡️ QA Check")
    st.markdown(
        "<small>Standalone QA check for TMX/MQXLIFF files. No anonymization performed.</small>",
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------------
    # Shortcut: load the in-memory anonymized output(s) directly into the QA
    # engine without re-uploading. Visible only after an anonymization run
    # produced results in session state. Avoids the download → re-upload
    # round-trip when QA-checking your own anonymized files. The original
    # uploader below still works for files anonymized elsewhere.
    # ------------------------------------------------------------------
    anon_results = st.session_state.get("results") or {}
    if anon_results:
        st.markdown("**🔁 From the current anonymization run**")
        anon_names = list(anon_results.keys())
        if len(anon_names) == 1:
            chosen_name = anon_names[0]
            st.caption(f"Anonymized file in memory: `{chosen_name}`")
        else:
            chosen_name = st.selectbox(
                f"Choose one of the {len(anon_names)} anonymized files",
                options=anon_names,
                key="qa_anon_pick",
            )
        if st.button(
            "🛡️ Check current anonymized output",
            key="qa_use_anon_btn",
            help="Run QA on the in-memory anonymized bytes — no need to download and re-upload.",
        ):
            anon_bytes = anon_results[chosen_name]
            new_sig = (chosen_name, hashlib.sha1(anon_bytes).hexdigest())
            for k in ("qa_results", "qa_target_overrides"):
                st.session_state.pop(k, None)
            for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
                st.session_state.pop(k, None)
            st.session_state["qa_filename"] = chosen_name
            st.session_state["qa_original_bytes"] = anon_bytes
            st.session_state["qa_file_signature"] = new_sig
            st.success(f"Loaded anonymized `{chosen_name}` into QA. Click 🚀 Run QA below.")
            st.rerun()
        st.markdown("<small>— or upload a different file —</small>", unsafe_allow_html=True)

    qa_file = st.file_uploader(
        "Upload one .tmx or .mqxliff file",
        type=["tmx", "mqxliff"],
        key="qa_file_uploader",
        help="Single file only. QA Check is independent from the Upload tab.",
    )

    if qa_file is not None:
        # Compare a content hash (sha1 of the bytes) plus the filename so any
        # change in the uploaded payload — including same-name/same-size with
        # different content — refreshes state correctly. Reading bytes once
        # per upload event is cheap relative to the QA pipeline that follows.
        new_bytes = qa_file.getvalue()
        new_sig = (qa_file.name, hashlib.sha1(new_bytes).hexdigest())
        prev_sig = st.session_state.get("qa_file_signature")
        if prev_sig != new_sig:
            for k in ("qa_results", "qa_target_overrides"):
                st.session_state.pop(k, None)
            # Drop any per-segment edit widget keys from the previous file
            # so the new file starts with empty edit fields.
            for k in [k for k in list(st.session_state.keys()) if k.startswith("qa_edit_")]:
                st.session_state.pop(k, None)
            st.session_state["qa_filename"] = qa_file.name
            st.session_state["qa_original_bytes"] = new_bytes
            st.session_state["qa_file_signature"] = new_sig

    with st.expander("⚙️ Configuration", expanded=False):
        # 33 checks available; 30 active by default. Three opt-IN checks
        # (`qa_checker.OPT_IN_CHECK_IDS`) ship OFF: `confusable_pairs`
        # (needs user-supplied word groups, high noise) and the two
        # custom-regex checks (Task #68 — the textareas only render once
        # one of those toggles is ticked so the panel stays uncluttered).
        cap_col, tick_col, untick_col = st.columns([6, 1, 1])
        with cap_col:
            st.markdown(
                "**33 checks available / 30 active by default.** "
                "Tick or untick as needed."
            )
        # Bulk toggles — mutate every per-check `qa_check_toggle_{cid}`
        # in session_state BEFORE the checkboxes render so the new
        # state is reflected on the same run. Streamlit forbids
        # writing to a widget-managed key after the widget is
        # instantiated, so this has to happen above the per-check
        # loop. `st.rerun()` ensures the user sees the change
        # immediately without needing a second interaction.
        with tick_col:
            if st.button("✓ Defaults", key="qa_tick_all_btn",
                         help=("Restore the 30 default-ON checks. Leaves the "
                               "three opt-in checks (Confusable, Custom "
                               "forbidden regex, Custom required regex) OFF "
                               "and resets the glossary sub-options "
                               "(case-sensitive, match inflected) to OFF.")):
                for _cid in qa_checker.CATEGORY_METADATA:
                    st.session_state[f"qa_check_toggle_{_cid}"] = (
                        _cid not in qa_checker.OPT_IN_CHECK_IDS
                    )
                # Glossary sub-options live outside CATEGORY_METADATA as
                # their own Streamlit checkboxes; reset them too so the
                # "Defaults" button truly restores the day-one state.
                # IMPORTANT: the checkbox widgets own the `_chk` keys —
                # if we only reset the mirror keys, on rerun the widget
                # repopulates the mirrors from its retained value and
                # the box stays ticked.
                st.session_state["qa_glossary_case_chk"] = False
                st.session_state["qa_glossary_case_sensitive"] = False
                st.session_state["qa_glossary_inflected_chk"] = False
                st.session_state["qa_glossary_inflected_forms"] = False
                st.rerun()
        with untick_col:
            if st.button("✗ None", key="qa_untick_all_btn",
                         help="Untick every check at once."):
                for _cid in qa_checker.CATEGORY_METADATA:
                    st.session_state[f"qa_check_toggle_{_cid}"] = False
                st.session_state["qa_glossary_case_chk"] = False
                st.session_state["qa_glossary_case_sensitive"] = False
                st.session_state["qa_glossary_inflected_chk"] = False
                st.session_state["qa_glossary_inflected_forms"] = False
                st.rerun()
        grouped = {}
        for cid, meta in qa_checker.CATEGORY_METADATA.items():
            grouped.setdefault(meta["group"], []).append((cid, meta))
        new_enabled = set()
        # Custom 3-column layout (Task #70). Content has the most
        # checks so it gets the left column alone; Tags has the fewest
        # so it gets the right column alone; the middle column stacks
        # Terminology → Numeric Elements → Spelling so the most
        # commonly tweaked block (Terminology) sits closest to the
        # Inconsistent-translation slider rendered below the columns.
        rendered_cols: Dict[int, List[str]] = {
            0: ["Content"],
            1: ["Terminology", "Numeric Elements", "Spelling"],
            2: ["Tags"],
        }
        # Any future group not listed above falls into the middle column
        # so it never disappears silently.
        _laid_out = {g for col in rendered_cols.values() for g in col}
        for group_name in grouped.keys():
            if group_name not in _laid_out:
                rendered_cols[1].append(group_name)
        cols = st.columns(3)
        for col_idx in (0, 1, 2):
            with cols[col_idx]:
                for group_name in rendered_cols[col_idx]:
                    items = grouped[group_name]
                    st.markdown(f"**{items[0][1]['icon']} {group_name}**")
                    for cid, meta in items:
                        key = f"qa_check_toggle_{cid}"
                        if key not in st.session_state:
                            # All checks default ON except the opt-IN trio
                            # (confusable pairs + the two custom-regex checks).
                            st.session_state[key] = (
                                cid not in qa_checker.OPT_IN_CHECK_IDS
                            )
                        checked = st.checkbox(
                            meta['label'],
                            key=key,
                        )
                        if checked:
                            new_enabled.add(cid)
        st.session_state["qa_enabled_checks"] = new_enabled

        # Task #51 — Inconsistent translation threshold slider. Only
        # rendered when the check is enabled, to keep the rest of the
        # config compact. 1.00 = exact source match only (default);
        # below 1.00 enables fuzzy matching via SequenceMatcher, mirroring
        # the Duplicates-tab behavior.
        if "inconsistent_translation" in new_enabled:
            # Constrain the slider to roughly the width of the left
            # column (1/3 of the expander) instead of stretching across
            # the full screen — easier to read and visually anchored to
            # the Content checks above.
            _thr_col, _ = st.columns([1, 2])
            with _thr_col:
                inconsist_thr = st.slider(
                    "Inconsistent translation: source-similarity threshold",
                    min_value=0.80, max_value=1.00,
                    value=float(st.session_state.get("qa_inconsistent_translation_threshold", 1.00)),
                    step=0.01,
                    key="qa_inconsistent_translation_thr",
                    help=(
                        "1.00 = only flag segments whose source is identical. "
                        "Lower values (e.g. 0.90) also flag near-duplicate sources, "
                        "matching the Duplicates tab's fuzzy mode."
                    ),
                )
                st.session_state["qa_inconsistent_translation_threshold"] = inconsist_thr

        # Spell-check sub-config: detected language + ignore-words list. Only
        # rendered when the user has actually enabled the check, to keep the
        # rest of the QA configuration uncluttered.
        if "spellcheck" in new_enabled:
            try:
                import spellcheck as _sc_mod
                detected = (
                    st.session_state.get("qa_results", {}) or {}
                ).get("target_lang") or ""
                norm = _sc_mod.normalize_lang_code(detected) if detected else None
                supported_count = len(_sc_mod.DICTIONARY_SOURCES)
                if detected and norm:
                    st.caption(
                        f"🔤 Detected language: **{norm}** — dictionary "
                        f"downloaded automatically on first use."
                    )
                elif detected and not norm:
                    st.caption(
                        f"⚠️ Detected language `{detected}` not supported "
                        f"({supported_count} languages available). "
                        f"Spell-check skipped."
                    )
                else:
                    st.caption(
                        f"🔤 Language auto-detected from the file; dictionary "
                        f"downloaded automatically on first use "
                        f"({supported_count} languages supported)."
                    )
            except Exception:
                pass

            ignore_text = st.text_area(
                "Spell-check: ignore words (one per line, proper nouns, brand names, acronyms…)",
                value=st.session_state.get("qa_spellcheck_ignore_text", ""),
                key="qa_spellcheck_ignore_text_area",
                height=90,
            )
            st.session_state["qa_spellcheck_ignore_text"] = ignore_text
            ignore_words = [w.strip() for w in (ignore_text or "").splitlines() if w.strip()]
            st.session_state["qa_spellcheck_ignore"] = ignore_words
            if ignore_words:
                st.caption(f"✅ {len(ignore_words)} word(s) will be ignored by spell-check")

        st.markdown("**📘 Forbidden terms** (one per line, case-insensitive)")
        forbidden_text = st.text_area(
            "Forbidden terms",
            value=st.session_state.get("qa_forbidden_text", ""),
            key="qa_forbidden_text_area",
            height=100,
            label_visibility="collapsed",
        )
        st.session_state["qa_forbidden_text"] = forbidden_text
        forbidden_terms = qa_checker.parse_forbidden_terms(forbidden_text)
        if forbidden_terms:
            st.caption(f"✅ {len(forbidden_terms)} forbidden term(s) loaded")

        st.markdown("**📒 Glossary** (TXT tab-separated, CSV or Excel `.xlsx`. 3 columns: source, target, *optional note*. Header row optional.)")
        glossary_file = st.file_uploader(
            "Glossary file",
            type=["txt", "csv", "xlsx"],
            key="qa_glossary_uploader",
            label_visibility="collapsed",
        )
        # Auto-enable "Match inflected forms" the first time a glossary is
        # uploaded (or when a different glossary file replaces the previous
        # one). Reviewers almost always want inflected matching active when
        # they bother to upload a glossary; the toggle stays user-overridable
        # afterwards because we only nudge the default on a new file event.
        if glossary_file is not None:
            _prev_name = st.session_state.get("qa_glossary_uploaded_name")
            if _prev_name != glossary_file.name:
                # Streamlit checkbox widgets own their state via their
                # `key=` once rendered. Setting only the mirror key
                # (`qa_glossary_inflected_forms`) wouldn't tick the box
                # on next render, so we seed the widget's own key
                # (`qa_glossary_inflected_chk`) BEFORE the widget is
                # instantiated below. Both are kept in sync so the rest
                # of the QA pipeline (which reads the mirror key) sees
                # the new value too.
                st.session_state["qa_glossary_inflected_chk"] = True
                st.session_state["qa_glossary_inflected_forms"] = True
                st.session_state["qa_glossary_uploaded_name"] = glossary_file.name
        # Widget keys are the source of truth — passing both `value=` and
        # mutating `st.session_state[key]` elsewhere triggers Streamlit's
        # "default value AND session state" warning. We seed the key once
        # (if missing) and let the widget own it from then on; the auto-tick
        # on a fresh glossary upload writes directly to this same key above.
        st.session_state.setdefault("qa_glossary_case_chk", False)
        st.session_state.setdefault("qa_glossary_inflected_chk", False)
        glossary_case_sensitive = st.checkbox(
            "Glossary case-sensitive matching",
            key="qa_glossary_case_chk",
        )
        st.session_state["qa_glossary_case_sensitive"] = glossary_case_sensitive
        glossary_inflected = st.checkbox(
            "Match inflected forms (Hunspell)",
            key="qa_glossary_inflected_chk",
            help=(
                "Match inflected forms of single-word entries "
                "(e.g. 'patient' matches 'patients'). "
                "Multi-word entries stay literal. Case-insensitive."
            ),
        )
        st.session_state["qa_glossary_inflected_forms"] = glossary_inflected
        glossary_entries = []
        if glossary_file is not None:
            glossary_entries = qa_checker.parse_glossary(
                glossary_file.getvalue(),
                glossary_file.name,
                case_sensitive=glossary_case_sensitive,
            )
            st.session_state["qa_glossary"] = glossary_entries
            st.caption(f"✅ {len(glossary_entries)} glossary entries loaded")
        else:
            glossary_entries = st.session_state.get("qa_glossary", [])
            if glossary_entries:
                st.caption(f"ℹ️ Using previously loaded glossary ({len(glossary_entries)} entries). Re-upload to replace.")

        # ----- Custom regex patterns (Task #61 / #68) -----
        # Both regex checks are opt-IN (default OFF). The intro line, the
        # case-sensitive toggle and the pattern textareas only render when
        # at least one toggle is ON, so the panel stays uncluttered for
        # users who never reach for custom regex. When a toggle is OFF its
        # textarea is skipped and the corresponding pattern list is empty,
        # so previously typed patterns simply don't fire until the user
        # re-enables the check.
        forbidden_on = "custom_forbidden_regex" in new_enabled
        required_on = "custom_required_regex" in new_enabled
        custom_forbidden_patterns: list = []
        custom_required_patterns: list = []
        if forbidden_on or required_on:
            st.markdown(
                "**🧩 Custom regex patterns:** your own Python regular expressions, "
                "evaluated alongside the built-in checks."
            )
            custom_regex_cs = st.checkbox(
                "Custom regex: case-sensitive matching",
                value=st.session_state.get("qa_custom_regex_case_sensitive", False),
                key="qa_custom_regex_cs_chk",
                help="Applies to both textareas below. Default = case-insensitive.",
            )
            st.session_state["qa_custom_regex_case_sensitive"] = custom_regex_cs
        else:
            custom_regex_cs = st.session_state.get(
                "qa_custom_regex_case_sensitive", False
            )

        if forbidden_on:
            custom_forbidden_text = st.text_area(
                "Forbidden patterns (regex, target). Must NOT appear in the target",
                value=st.session_state.get("qa_custom_forbidden_regex", ""),
                key="qa_custom_forbidden_regex_area",
                height=90,
            )
            st.caption(
                "Python regular expressions, one per line. Empty lines and lines "
                "starting with `#` are ignored. Example: `\\bplacebo\\b`."
            )
            st.session_state["qa_custom_forbidden_regex"] = custom_forbidden_text
            custom_forbidden_patterns, custom_forbidden_errors = (
                qa_checker.parse_custom_regex_patterns(
                    custom_forbidden_text, case_sensitive=custom_regex_cs
                )
            )
            for err in custom_forbidden_errors:
                st.error(
                    f"❌ Forbidden pattern line {err['line']} "
                    f"(`{err['raw']}`): {err['error']}"
                )
            if custom_forbidden_patterns:
                st.caption(
                    f"✅ {len(custom_forbidden_patterns)} forbidden regex pattern(s) loaded"
                )

        if required_on:
            custom_required_text = st.text_area(
                "Required patterns (regex, source → target). If matched in source, "
                "must match the same number of times in target",
                value=st.session_state.get("qa_custom_required_regex", ""),
                key="qa_custom_required_regex_area",
                height=90,
            )
            st.caption(
                "Python regular expressions, one per line. Empty lines and lines "
                "starting with `#` are ignored. Example: `\\bNCT\\d{8}\\b` keeps "
                "every clinical-trial ID present in the source."
            )
            st.session_state["qa_custom_required_regex"] = custom_required_text
            custom_required_patterns, custom_required_errors = (
                qa_checker.parse_custom_regex_patterns(
                    custom_required_text, case_sensitive=custom_regex_cs
                )
            )
            for err in custom_required_errors:
                st.error(
                    f"❌ Required pattern line {err['line']} "
                    f"(`{err['raw']}`): {err['error']}"
                )
            if custom_required_patterns:
                st.caption(
                    f"✅ {len(custom_required_patterns)} required regex pattern(s) loaded"
                )

        # ----- Confusable pairs (Task #42) -----
        if "confusable_pairs" in new_enabled:
            try:
                import confusable_pairs as _cp_mod
            except Exception:
                _cp_mod = None
            if _cp_mod is not None:
                st.markdown("**🔤 Confusable pairs** (real-word errors that spell-check can't catch)")
                st.caption(
                    "Add your own confusable word groups via the textarea or "
                    "an .xlsx upload. No built-in language packs are shipped."
                )

                custom_text = st.text_area(
                    "Custom confusable pairs (one group per line, members separated by `|` or `/`)",
                    value=st.session_state.get("qa_confusable_custom_text", ""),
                    key="qa_confusable_custom_area",
                    height=90,
                    help="Examples:\nefectivo|eficaz|eficiente\nprincipal/principle",
                )
                st.session_state["qa_confusable_custom_text"] = custom_text
                custom_pairs = _cp_mod.parse_custom_pairs(custom_text)

                custom_xlsx = st.file_uploader(
                    "…or upload an Excel `.xlsx` with custom pairs (one row per group, one column per member)",
                    type=["xlsx"],
                    key="qa_confusable_custom_uploader",
                )
                if custom_xlsx is not None:
                    xlsx_pairs = _cp_mod.parse_pairs_xlsx(custom_xlsx.getvalue(), custom_xlsx.name)
                    if xlsx_pairs:
                        # Merge avoiding duplicates with textarea entries.
                        seen_keys = {tuple(sorted(m.lower() for m in g)) for g in custom_pairs}
                        for g in xlsx_pairs:
                            k = tuple(sorted(m.lower() for m in g))
                            if k not in seen_keys:
                                seen_keys.add(k)
                                custom_pairs.append(g)
                st.session_state["qa_confusable_custom_pairs"] = custom_pairs
                if custom_pairs:
                    st.caption(f"✅ {len(custom_pairs)} custom confusable group(s) loaded")

    # Show the Run QA button ONLY when a file is actually loaded (either
    # uploaded in this tab or piped in from the Anonymizer). This keeps the
    # initial QA tab clean — no orphan disabled button before there's anything
    # to check.
    can_run = qa_file is not None or st.session_state.get("qa_original_bytes") is not None
    run_clicked = False
    if can_run:
        _qa_run_l, _qa_run_c, _qa_run_r = st.columns([2, 1, 2])
        with _qa_run_c:
            run_clicked = st.button(
                "🚀 Run QA",
                type="primary",
                key="qa_run_btn",
                use_container_width=True,
            )

    if run_clicked and can_run:
        config = {
            "enabled_checks": st.session_state.get("qa_enabled_checks", qa_checker.DEFAULT_PROFILE),
            "forbidden_terms": forbidden_terms,
            "glossary": st.session_state.get("qa_glossary", []),
            "glossary_case_sensitive": st.session_state.get("qa_glossary_case_sensitive", False),
            "glossary_inflected_forms": st.session_state.get("qa_glossary_inflected_forms", False),
            "inconsistent_translation_threshold": st.session_state.get("qa_inconsistent_translation_threshold", 1.00),
            "spellcheck_ignore": st.session_state.get("qa_spellcheck_ignore", []),
            "confusable_pairs_custom": st.session_state.get("qa_confusable_custom_pairs", []),
            "custom_forbidden_regex_patterns": custom_forbidden_patterns,
            "custom_required_regex_patterns": custom_required_patterns,
            "custom_regex_case_sensitive": custom_regex_cs,
        }
        try:
            with st.spinner("Running QA checks..."):
                results = qa_checker.run_qa_checks(
                    st.session_state["qa_original_bytes"],
                    st.session_state["qa_filename"],
                    config,
                )
            st.session_state["qa_results"] = results
            st.session_state["qa_config"] = config
        except ValueError as exc:
            st.error(f"❌ {exc}")
            return
        except Exception as exc:
            st.error(f"❌ Unexpected error during QA: {exc}")
            return

    results = st.session_state.get("qa_results")
    if not results:
        if not can_run:
            st.info("Upload a TMX or MQXLIFF file to begin.")
        return

    st.markdown("---")
    summary = results["summary"]
    target_overrides = st.session_state.setdefault("qa_target_overrides", {})

    # ---- Pre-pass: synchronise per-card editor values into the canonical
    # `qa_target_overrides` map BEFORE the summary line is rendered, so the
    # "N segments overridden manually" count reflects the just-typed edit
    # on the very next rerun (rather than lagging by one).
    #
    # Per-card widget keys follow the pattern `qa_edit_{seg_id}__{card_uid}`
    # (see comments around the inner render loop). For each segment that
    # has at least one such key in session_state, we:
    #   1. Look up its canonical value (override entry, or original target).
    #   2. Promote any sibling widget whose value differs from the
    #      canonical (i.e. the user just typed) to the new canonical.
    #   3. Pop ALL sibling widget keys so each per-card editor below
    #      re-initialises this run from `value=canonical_target`.
    seg_originals: dict[str, str] = {}
    for _cat in results["categories"].values():
        for _issue in _cat["issues"]:
            sid = str(_issue["segment_id"])
            if sid not in seg_originals:
                seg_originals[sid] = _issue.get("target") or ""

    sibling_groups: dict[str, list[str]] = {}
    for _k in list(st.session_state.keys()):
        if not _k.startswith("qa_edit_"):
            continue
        rest = _k[len("qa_edit_"):]
        if "__" not in rest:
            continue
        sid = rest.split("__", 1)[0]
        sibling_groups.setdefault(sid, []).append(_k)

    synced_segs_this_run: set = set()
    for sid, keys in sibling_groups.items():
        if sid not in seg_originals:
            # Orphaned widget (e.g. results no longer contain this segment).
            for _k in keys:
                st.session_state.pop(_k, None)
            continue
        original = seg_originals[sid]
        canonical = target_overrides.get(sid, original)
        for _k in keys:
            v = st.session_state.get(_k)
            if v != canonical:
                canonical = v
                break
        if canonical != original:
            target_overrides[sid] = canonical
        else:
            target_overrides.pop(sid, None)
        for _k in keys:
            st.session_state.pop(_k, None)
        synced_segs_this_run.add(sid)

    st.session_state["qa_target_overrides"] = target_overrides
    overrides_pending = len(target_overrides)

    info_line = (
        f"**File:** `{results['filename']}` · **Format:** {results['format'].upper()} · "
        f"**Source:** `{results['source_lang']}` · **Target:** `{results['target_lang']}` · "
        f"**Segments:** {results['segment_count']}"
    )
    st.markdown(info_line)

    # Top-of-results notices (e.g. "Spell-check skipped: dictionary unavailable…").
    for _notice in results.get("notices", []) or []:
        st.info(_notice)

    if summary["total"] == 0 and overrides_pending == 0:
        st.success("No issues found 🎉. The file passes all enabled checks.")
    else:
        # Hide the manual-override badge when count is 0 — it's noise the
        # rest of the time. Task #68 — single TOTAL pill replaces the
        # previous HIGH / LOW headline badges.
        extras_html = ""
        if overrides_pending > 0:
            extras_html = (
                f" &nbsp; <span style='color:#1a5488;font-weight:600;'>"
                f"{overrides_pending} segments overridden manually</span>"
            )
        st.markdown(
            f"**Summary:** {_qa_total_badge(summary['total'])}{extras_html}",
            unsafe_allow_html=True,
        )

    # Color legend (Task #36): one swatch per QA group, shared between the
    # in-app highlight and the HTML report so users can map a color back to
    # a group at a glance. Hidden when there are no issues to look at.
    if summary["total"] > 0:
        st.markdown(_qa_legend_html(), unsafe_allow_html=True)

    raw_search = st.text_input(
        "🔍 Filter issues by text or segment number",
        value="",
        key="qa_search_input",
        placeholder="Text, or segment numbers: 12 · 12,15,20 · 10-20 · 5, 8-11, 30",
    ).strip()
    # xbench/memoQ-style search options. Segment-number queries bypass these
    # (they're parsed before the text path even runs).
    opt_cols = st.columns(5)
    with opt_cols[0]:
        search_in_source = st.checkbox("Source", value=True, key="qa_search_in_source")
    with opt_cols[1]:
        search_in_target = st.checkbox("Target", value=True, key="qa_search_in_target")
    with opt_cols[2]:
        search_regex = st.checkbox("Regex", value=False, key="qa_search_regex")
    with opt_cols[3]:
        search_whole_words = st.checkbox("Whole words", value=False, key="qa_search_whole_words")
    with opt_cols[4]:
        search_case_sensitive = st.checkbox("Case-sensitive", value=False, key="qa_search_case_sensitive")

    seg_filter = _parse_segment_filter(raw_search)
    search_query = "" if seg_filter is not None else raw_search
    if seg_filter is not None and not seg_filter:
        st.info("No segment numbers parsed from your query.")

    # Compile the search pattern once. None = no active text filter; a
    # compiled pattern = use it; "invalid" sentinel = warn user and skip.
    compiled_pattern = None
    if search_query:
        if not (search_in_source or search_in_target):
            st.info("Tick **Source** and/or **Target** to search inside segments.")
            compiled_pattern = "invalid"
        else:
            flags = 0 if search_case_sensitive else re.IGNORECASE
            pattern_str = search_query if search_regex else re.escape(search_query)
            if search_whole_words:
                pattern_str = r"(?<!\w)" + pattern_str + r"(?!\w)"
            try:
                compiled_pattern = re.compile(pattern_str, flags)
            except re.error as exc:
                st.warning(f"Invalid regex: {exc}")
                compiled_pattern = "invalid"

    # The same segment may appear in several QA categories (e.g. forbidden term +
    # whitespace edges). Every card must show its own editable text_area
    # while sharing the same logical edit value across cards. Streamlit
    # forbids two widgets sharing the same `key` in the same script run, so
    # each card uses a unique per-card widget key
    # (``qa_edit_{seg_id}__{card_uid}``) and we sync them via a single
    # canonical entry in ``qa_target_overrides[seg_id]``.
    #
    # The sync itself was already performed by the pre-pass above (so the
    # summary line's override count is accurate this very rerun). All sibling
    # widget keys for synced segments have been popped, so each per-card
    # editor below initialises from ``value=canonical_target``.
    cat_items = list(results["categories"].items())
    shown_issue_count = 0
    for idx, (cid, cat) in enumerate(cat_items):
        issues = cat["issues"]
        if seg_filter is not None:
            def _seg_id(i):
                v = i.get("segment_id")
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return None
            issues = [i for i in issues if _seg_id(i) in seg_filter]
        elif compiled_pattern == "invalid":
            issues = []
        elif compiled_pattern is not None:
            def _matches(i):
                if search_in_source and compiled_pattern.search(i.get("source") or ""):
                    return True
                if search_in_target and compiled_pattern.search(i.get("target") or ""):
                    return True
                return False
            issues = [i for i in issues if _matches(i)]
        if not issues:
            continue
        shown_issue_count += len(issues)
        header = f"{cat['icon']} {cat['label']} ({len(issues)})"
        with st.expander(header, expanded=False):
            # Task #68 — per-category severity badge removed; the group
            # color in the legend / highlighter already conveys the tag.
            visible_issues = issues[:200]
            for issue_idx, issue in enumerate(visible_issues):
                seg_id = str(issue["segment_id"])
                original_target = issue.get("target") or ""
                canonical_target = target_overrides.get(seg_id, original_target)
                is_edited = (canonical_target != original_target)

                edited_badge = (
                    " <span style=\"background:#1a5488;color:#fff;padding:1px 6px;"
                    "border-radius:3px;font-size:0.7rem;font-weight:700;margin-left:6px;\">EDITED</span>"
                    if is_edited else ""
                )
                # Optional reviewer note (glossary column C). Rendered
                # inside the same header card so it visually belongs to
                # the issue, with a softer warm background so it reads
                # as auxiliary context rather than the main message.
                note_text = issue.get("note") or ""
                note_html = (
                    f'<div style="margin-top:6px;padding:5px 10px;background:#fff8e1;'
                    f'border-left:3px solid #f0b429;border-radius:3px;font-size:0.82rem;'
                    f'color:#5e5f6b;line-height:1.35;">'
                    f'<strong>Note:</strong> {_html_escape(note_text)}</div>'
                ) if note_text else ""
                st.markdown(
                    f'<div style="border-left:4px solid #1a5488;padding:8px 14px;margin:14px 0 8px 0;background:#d4dde4;border-radius:4px;">'
                    f'<strong>Segment {issue["segment_id"]}:</strong> {_html_escape(issue["message"])}{edited_badge}'
                    f'{note_html}</div>',
                    unsafe_allow_html=True,
                )
                col_s, col_t = st.columns(2)
                # Task #51 — inconsistent_translation paints BOTH cells
                # with the Content group background. Source = plain text
                # (the source is by definition the same / near-same as
                # the reference, so there's nothing meaningful to diff).
                # Target = word-level diff vs the reference target so
                # the reviewer immediately sees what diverged.
                is_inconsist = (
                    cid == "inconsistent_translation"
                    and issue.get("reference_segment_id") is not None
                )
                if is_inconsist:
                    # Task #66 — derive the card chrome from the Content
                    # palette via the single-source-of-truth lookup so the
                    # in-app card matches the exported HTML report (both
                    # render `inconsistent_translation` under Content).
                    _ic_bg, _ic_fg = qa_checker.get_highlight_color(
                        "inconsistent_translation")
                    cell_style = (
                        f"background:{_ic_bg};border:1px solid {_ic_fg};"
                        f"border-left:4px solid {_ic_fg};border-radius:4px;"
                        f"padding:6px 10px;font-size:0.9rem;"
                    )
                else:
                    cell_style = (
                        "background:#fff;border:1px solid #dee2e6;"
                        "border-radius:4px;padding:6px 10px;font-size:0.9rem;"
                    )

                # Task #55 / #66 — stacked "seg X" + "seg Y" block so the
                # reviewer sees both segments together. Top row is the
                # reference (lower seg_id by construction); diff is painted
                # on the cell whose axis diverges. Task #66 dropped the
                # "Ref ·" / "This ·" prefixes — the two-line stack already
                # communicates which row is the reference.
                if is_inconsist:
                    axis = issue.get("mismatch_axis") or "target"
                    ref_src_txt = issue.get("reference_source") or ""
                    ref_tgt_txt = issue.get("reference_target") or ""
                    this_src_txt = issue.get("source") or ""
                    this_tgt_txt = issue.get("target") or ""
                    ref_id = issue.get("reference_segment_id")
                    if axis == "source":
                        ref_src_html_i, this_src_html_i = highlight_diff(ref_src_txt, this_src_txt)
                        ref_tgt_html_i = _html_escape(ref_tgt_txt)
                        this_tgt_html_i = _html_escape(this_tgt_txt)
                    elif axis == "both":
                        ref_src_html_i, this_src_html_i = highlight_diff(ref_src_txt, this_src_txt)
                        ref_tgt_html_i, this_tgt_html_i = highlight_diff(ref_tgt_txt, this_tgt_txt)
                    else:  # "target"
                        ref_src_html_i = _html_escape(ref_src_txt)
                        this_src_html_i = _html_escape(this_src_txt)
                        ref_tgt_html_i, this_tgt_html_i = highlight_diff(ref_tgt_txt, this_tgt_txt)
                    # When the user has edited the segment, the diff vs
                    # the original target no longer maps onto the edited
                    # text — fall back to plain text on the "This" target.
                    if is_edited:
                        this_tgt_html_i = _html_escape(canonical_target)

                    def _stack_inconsist(ref_html: str, this_html: str) -> str:
                        return (
                            f'<div style="font-size:0.74rem;color:#5e5f6b;'
                            f'font-weight:600;margin-bottom:2px;">seg {ref_id}</div>'
                            f'<div style="margin-bottom:8px;">{ref_html}</div>'
                            f'<div style="font-size:0.74rem;color:#5e5f6b;'
                            f'font-weight:600;margin-bottom:2px;">seg {seg_id}</div>'
                            f'<div>{this_html}</div>'
                        )

                with col_s:
                    st.markdown("**Source:**")
                    if is_inconsist:
                        source_html = _stack_inconsist(ref_src_html_i, this_src_html_i)
                    else:
                        source_html = _qa_highlight(issue["source"], issue.get("span_source"), cid)
                    # Overlay yellow search highlight on top of the QA
                    # finding highlight, but only on the field(s) the
                    # user is actually searching in.
                    if search_in_source:
                        source_html = _qa_highlight_search(source_html, compiled_pattern)
                    st.markdown(
                        f'<div style="{cell_style}">{source_html}</div>',
                        unsafe_allow_html=True,
                    )
                with col_t:
                    st.markdown("**Target:**")
                    # When the user has edited the segment, the highlight
                    # span computed at QA time no longer maps onto the
                    # edited text — show the canonical edited text
                    # plain-escaped instead.
                    if is_inconsist:
                        target_html = _stack_inconsist(ref_tgt_html_i, this_tgt_html_i)
                    elif is_edited:
                        target_html = _html_escape(canonical_target)
                    else:
                        target_html = _qa_highlight(issue["target"], issue.get("span_target"), cid)
                    if search_in_target:
                        target_html = _qa_highlight_search(target_html, compiled_pattern)
                    st.markdown(
                        f'<div style="{cell_style}">{target_html}</div>',
                        unsafe_allow_html=True,
                    )

                # Visual breathing room between the read-only Source/Target
                # boxes and the editable target widget below.
                st.write("")

                # Per-card editor. Unique widget key per card so multiple
                # cards for the same segment can coexist in one Streamlit
                # script run; their values are kept in lock-step by the
                # sync block above on every rerun.
                card_uid = f"{cid}__{issue_idx}"
                widget_key = f"qa_edit_{seg_id}__{card_uid}"
                edit_cols = st.columns([5, 1])
                with edit_cols[0]:
                    st.text_area(
                        "Edit target (plain text)",
                        value=canonical_target,
                        key=widget_key,
                        height=80,
                        label_visibility="collapsed",
                    )
                with edit_cols[1]:
                    st.button(
                        "↺ Reset",
                        key=f"qa_reset_btn_{seg_id}__{card_uid}",
                        on_click=_qa_reset_segment_edit,
                        args=(seg_id,),
                        disabled=not is_edited,
                        help="Revert this segment to its original target text.",
                    )
                st.caption(
                    "⚠️ Saving an edit replaces the segment as plain text. "
                    "Any inline tags inside this segment will be lost in the downloaded file."
                )

            if len(issues) > 200:
                st.caption(f"… and {len(issues) - 200} more issues in this category (showing first 200)")

    # Empty-state message when an active filter (segment numbers or text
    # search) hides every issue. Without this, the QA results area silently
    # collapses and the user can't tell whether the filter matched nothing
    # or the file truly has no issues.
    if shown_issue_count == 0:
        if seg_filter is not None:
            st.info("🔍 No issues found for the segment number(s) you entered.")
        elif compiled_pattern == "invalid":
            pass  # warning/info already shown above by the search-bar logic
        elif compiled_pattern is not None:
            st.info("🔍 No issues match your search.")

    # `overrides_pending` is already accurate from the pre-pass at the top
    # of this function; per-card editors below cannot mutate it within the
    # same script run (they only enqueue edits for the next rerun's sync).
    if overrides_pending > 0:
        st.button(
            f"↺ Reset all edits ({overrides_pending})",
            key="qa_reset_all_btn",
            on_click=_qa_reset_all_edits,
            help="Revert every manually edited segment to its original target text.",
        )

    st.markdown("---")

    base_name = (st.session_state["qa_filename"] or "file").rsplit(".", 1)
    stem = base_name[0]
    ext = base_name[1] if len(base_name) > 1 else ("tmx" if results["format"] == "tmx" else "mqxliff")
    cleaned_name = f"Cleaned_{stem}.{ext}"
    mime = "application/xml"

    _spacer_l, col_dl, col_rep_html, _spacer_r = st.columns([2, 3, 3, 2])
    with col_dl:
        try:
            cleaned_bytes = qa_checker.prepare_qa_download(
                st.session_state["qa_original_bytes"],
                st.session_state["qa_filename"],
                target_overrides=target_overrides,
            )
            if target_overrides:
                help_text = (
                    f"Applies {len(target_overrides)} inline target edit(s). "
                    "Manual edits replace the target as plain text and discard "
                    "inline tags inside the segment. The original XML structure "
                    "is preserved for every other segment."
                )
            else:
                help_text = (
                    "Round trips the original file unchanged. Use the inline "
                    "editor on any issue card to override a target. Those "
                    "edits will be applied here."
                )
            st.download_button(
                "📥 Download cleaned file",
                data=cleaned_bytes,
                file_name=cleaned_name,
                mime=mime,
                key="qa_download_cleaned",
                disabled=not target_overrides,
                help=help_text,
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"Could not prepare cleaned file: {exc}")
    with col_rep_html:
        try:
            html_bytes = qa_checker.export_qa_report(
                results, "html", target_overrides=target_overrides,
            )
            st.download_button(
                "📰 Export HTML report",
                data=html_bytes,
                file_name=f"QA_Report_{stem}.html",
                mime="text/html",
                key="qa_download_html",
                disabled=summary["total"] == 0,
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"HTML export failed: {exc}")


# Informational threshold for the O(n²) fuzzy paths in the two
# duplicate-detection helpers below. Above this segment count the UI
# shows a one-line "this may take a while" notice so users don't
# assume the app has frozen. The comparison itself still runs to
# completion — no segments are dropped.
LARGE_FUZZY_NOTICE_THRESHOLD = 15000


def detect_conflicting_translations(previews: dict, dedup_threshold: int = 100,
                                     no_anon_segments: dict = None,
                                     filter_junk: bool = False, min_words_junk: int = 2,
                                     filter_short: bool = False, min_words: int = 5,
                                     dedup_used_keys: set = None) -> list:
    if not no_anon_segments:
        no_anon_segments = {}
    if not dedup_used_keys:
        dedup_used_keys = set()
    replacement_token = st.session_state.get('replacement_token', '███')
    token_esc = re.escape(replacement_token)
    consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'

    all_segments = []
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            if segment_key in dedup_used_keys:
                continue
            if filter_junk and is_junk_segment(preview, min_words_junk):
                if not st.session_state.get(f"skipjunk_{segment_key}", False):
                    continue
            if no_anon_segments.get(segment_key, False):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue

            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)

            if not source_text and not target_text:
                continue

            all_segments.append({
                "file": filename,
                "segment": preview['segment'],
                "source": source_text,
                "target": target_text,
                "key": segment_key
            })

    groups = []
    used = set()

    if dedup_threshold >= 100:
        source_map = {}
        for seg in all_segments:
            src = seg["source"]
            if src not in source_map:
                source_map[src] = []
            source_map[src].append(seg)
        for src, members in source_map.items():
            if len(members) < 2:
                continue
            targets = set(m["target"] for m in members)
            if len(targets) < 2:
                continue
            groups.append({
                "members": members,
                "similarity": 100.0
            })
            for m in members:
                used.add(m["key"])
    else:
        # Fuzzy mode is O(n²): every segment is compared to every other
        # via SequenceMatcher. On large TMs this is slow but not broken
        # — just let the user know up front so they don't think the app
        # has frozen. The threshold below is purely informational.
        if len(all_segments) > LARGE_FUZZY_NOTICE_THRESHOLD:
            st.info(
                f"ℹ️ Large file ({len(all_segments):,} segments) with fuzzy "
                "matching enabled — this analysis may take a while. The app "
                "isn't frozen; please wait."
            )
        threshold_ratio = dedup_threshold / 100.0
        for i, seg_a in enumerate(all_segments):
            if seg_a["key"] in used:
                continue
            group_members = [seg_a]
            best_similarity = 0.0
            for j in range(i + 1, len(all_segments)):
                seg_b = all_segments[j]
                if seg_b["key"] in used:
                    continue
                ratio = difflib.SequenceMatcher(None, seg_a["source"], seg_b["source"]).ratio()
                if ratio >= threshold_ratio:
                    group_members.append({
                        **seg_b,
                        "_similarity": ratio * 100
                    })
                    if ratio * 100 > best_similarity:
                        best_similarity = ratio * 100
            if len(group_members) > 1:
                targets = set(m["target"] for m in group_members)
                if len(targets) < 2:
                    continue
                group_members[0]["_similarity"] = best_similarity if best_similarity > 0 else 100.0
                groups.append({
                    "members": group_members,
                    "similarity": best_similarity if best_similarity > 0 else 100.0
                })
                for m in group_members:
                    used.add(m["key"])

    return groups


def detect_duplicate_groups(previews: dict, dedup_threshold: int = 100,
                            no_anon_segments: dict = None,
                            filter_junk: bool = False, min_words_junk: int = 2,
                            filter_short: bool = False, min_words: int = 5) -> list:
    if not no_anon_segments:
        no_anon_segments = {}
    replacement_token = st.session_state.get('replacement_token', '███')
    token_esc = re.escape(replacement_token)
    consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'

    all_segments = []
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            if filter_junk and is_junk_segment(preview, min_words_junk):
                if not st.session_state.get(f"skipjunk_{segment_key}", False):
                    continue
            if no_anon_segments.get(segment_key, False):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue

            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)

            if not source_text and not target_text:
                continue

            all_segments.append({
                "file": filename,
                "segment": preview['segment'],
                "source": source_text,
                "target": target_text,
                "key": segment_key
            })

    groups = []
    used = set()

    if dedup_threshold >= 100:
        text_map = {}
        for seg in all_segments:
            pair = (seg["source"], seg["target"])
            if pair not in text_map:
                text_map[pair] = []
            text_map[pair].append(seg)
        for pair, members in text_map.items():
            if len(members) > 1:
                group = {
                    "members": members,
                    "similarity": 100.0
                }
                groups.append(group)
                for m in members:
                    used.add(m["key"])
    else:
        # See LARGE_FUZZY_NOTICE_THRESHOLD above — purely informational
        # notice; the fuzzy comparison still runs to completion.
        if len(all_segments) > LARGE_FUZZY_NOTICE_THRESHOLD:
            st.info(
                f"ℹ️ Large file ({len(all_segments):,} segments) with fuzzy "
                "matching enabled — this analysis may take a while. The app "
                "isn't frozen; please wait."
            )
        threshold_ratio = dedup_threshold / 100.0
        for i, seg_a in enumerate(all_segments):
            if seg_a["key"] in used:
                continue
            combined_a = seg_a["source"] + " ||| " + seg_a["target"]
            group_members = [seg_a]
            best_similarity = 0.0
            for j in range(i + 1, len(all_segments)):
                seg_b = all_segments[j]
                if seg_b["key"] in used:
                    continue
                combined_b = seg_b["source"] + " ||| " + seg_b["target"]
                ratio = difflib.SequenceMatcher(None, combined_a, combined_b).ratio()
                if ratio >= threshold_ratio:
                    group_members.append({
                        **seg_b,
                        "_similarity": ratio * 100
                    })
                    if ratio * 100 > best_similarity:
                        best_similarity = ratio * 100
                    used.add(seg_b["key"])
            if len(group_members) > 1:
                group_members[0]["_similarity"] = best_similarity
                groups.append({
                    "members": group_members,
                    "similarity": best_similarity
                })
                used.add(seg_a["key"])

    return groups


def extract_existing_canonical_ids(originals: dict) -> dict:
    """Extract existing x-document canonical IDs from original TMX files.
    Returns dict: {(filename, segment_number) -> canonical_id}"""
    from lxml import etree
    canonical_map = {}
    for filename, content in originals.items():
        if not filename.lower().endswith(".tmx"):
            continue
        try:
            if isinstance(content, str):
                content = content.encode('utf-8')
            tree = etree.fromstring(content)
            for i, tu in enumerate(tree.xpath("//tu")):
                doc_prop = tu.find("prop[@type='x-document']")
                if doc_prop is not None and doc_prop.text:
                    canonical_map[(filename, i + 1)] = doc_prop.text.strip()
        except Exception:
            pass
    return canonical_map


def generate_clean_tmx(previews: dict, results: dict, originals: dict,
                       filter_junk: bool, min_words_junk: int,
                       filter_short: bool, min_words: int,
                       exclude_modified: bool, exclusion_threshold: float,
                       excluded_segments: dict, no_anon_segments: dict,
                       dedup_tmx: bool = True, dedup_threshold: int = 100,
                       dedup_keep_choices: dict = None,
                       tmx_filename: str = None,
                       canonical_id: str = None,
                       existing_canonical_map: dict = None) -> tuple:
    """Generate a clean TMX containing only valid anonymized segments."""
    from lxml import etree
    
    src_lang = "en"
    tgt_lang = "es"
    
    for filename in originals:
        try:
            tree = etree.fromstring(originals[filename])
            is_tmx = filename.lower().endswith(".tmx")
            if is_tmx:
                header = tree.find(".//header")
                if header is not None:
                    src_lang = header.get("srclang", "en")
                    tus = tree.xpath("//tu/tuv")
                    langs = set()
                    for tuv in tus:
                        lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                        if lang:
                            langs.add(lang)
                    for lang in langs:
                        if lang.lower() != src_lang.lower():
                            tgt_lang = lang
                            break
            else:
                nsmap = tree.nsmap
                default_ns = nsmap.get(None, '')
                if default_ns:
                    ns = {'x': default_ns}
                    file_els = tree.xpath('//x:file', namespaces=ns)
                else:
                    file_els = tree.xpath('//file')
                if file_els:
                    src_lang = file_els[0].get("source-language", "en")
                    tgt_lang = file_els[0].get("target-language", "es")
            break
        except Exception:
            pass
    
    src_lang = _normalize_lang_code(src_lang)
    tgt_lang = _normalize_lang_code(tgt_lang)
    
    tmx_root = etree.Element("tmx", version="1.4")
    header = etree.SubElement(tmx_root, "header",
                              creationtool="Anonymizer",
                              datatype="PlainText",
                              segtype="sentence",
                              srclang=src_lang)
    has_existing = existing_canonical_map and len(existing_canonical_map) > 0
    if has_existing:
        header_filename = tmx_filename
    else:
        header_filename = canonical_id if canonical_id else tmx_filename
    if header_filename:
        prop = etree.SubElement(header, "prop", type="x-filename")
        prop.text = header_filename
    body = etree.SubElement(tmx_root, "body")
    
    valid_count = 0
    no_anon_skipped = 0
    dedup_count = 0
    dedup_details = []
    excluded_ids = []
    replacement_token = st.session_state.get('replacement_token', '███')
    seen_exact = set()
    seen_segments = []
    
    dedup_discard_keys = set()
    if dedup_keep_choices and dedup_tmx:
        dedup_group_keys = st.session_state.get('dedup_group_keys', {})
        for stable_id, kept_key in dedup_keep_choices.items():
            if kept_key == "__keep_all__":
                continue
            group_keys = dedup_group_keys.get(stable_id, [])
            for k in group_keys:
                if k != kept_key:
                    dedup_discard_keys.add(k)

    conflict_keep_choices = st.session_state.get('conflict_keep', {})
    if conflict_keep_choices and dedup_tmx:
        conflict_group_keys = st.session_state.get('conflict_group_keys', {})
        for stable_id, kept_key in conflict_keep_choices.items():
            if kept_key == "__keep_all__":
                continue
            group_keys = conflict_group_keys.get(stable_id, [])
            for k in group_keys:
                if k != kept_key:
                    dedup_discard_keys.add(k)
    
    short_excluded = []
    short_anon_excluded = []
    empty_excluded = []
    heavy_excluded = []
    
    for filename, file_previews in previews.items():
        for preview in file_previews:
            segment_key = f"{filename}_{preview['segment']}"
            
            if filter_junk and is_junk_segment(preview, min_words_junk):
                sk = f"skipjunk_{segment_key}"
                if not st.session_state.get(sk, False):
                    excluded_ids.append(preview['segment'])
                    short_excluded.append({
                        "file": filename, "segment": preview['segment'],
                        "source": preview.get('source_before', ''),
                        "target": preview.get('target_before', '')
                    })
                    continue
            
            if no_anon_segments.get(segment_key, False):
                no_anon_skipped += 1
                continue
            
            if filter_short and segment_word_count(preview) < min_words:
                sk_short = f"skipshort_{segment_key}"
                if not st.session_state.get(sk_short, False):
                    excluded_ids.append(preview['segment'])
                    short_anon_excluded.append({
                        "file": filename, "segment": preview['segment'],
                        "source_before": preview.get('source_before', ''),
                        "target_before": preview.get('target_before', ''),
                        "source_after": preview.get('source_after', ''),
                        "target_after": preview.get('target_after', '')
                    })
                    continue
            
            changed = preview.get('changed', False)
            source_text = preview.get('source_after' if changed else 'source_before', '').strip()
            target_text = preview.get('target_after' if changed else 'target_before', '').strip()
            
            source_text = strip_inline_tags(source_text)
            target_text = strip_inline_tags(target_text)
            
            token_esc = re.escape(replacement_token)
            consol_pat = rf'{token_esc}(?:\s*[,;.\-–—/|]\s*{token_esc}|\s+{token_esc}|\s*\(\s*{token_esc}\s*\))+'
            source_text = re.sub(consol_pat, replacement_token, source_text)
            target_text = re.sub(consol_pat, replacement_token, target_text)
            
            if not source_text and not target_text:
                excluded_ids.append(preview['segment'])
                empty_excluded.append({
                    "file": filename, "segment": preview['segment']
                })
                continue
            
            if exclude_modified:
                should_exclude = excluded_segments.get(segment_key, None)
                if should_exclude is not False:
                    has_token_src = replacement_token in (preview.get('source_after', ''))
                    has_token_tgt = replacement_token in (preview.get('target_after', ''))
                    if has_token_src or has_token_tgt:
                        src_before = preview.get('source_before', '')
                        tgt_before = preview.get('target_before', '')
                        src_after = preview.get('source_after', '')
                        tgt_after = preview.get('target_after', '')
                        src_pct = (1 - len(src_after.replace(replacement_token, '')) / max(len(src_before), 1)) * 100 if src_before else 0
                        tgt_pct = (1 - len(tgt_after.replace(replacement_token, '')) / max(len(tgt_before), 1)) * 100 if tgt_before else 0
                        max_pct = max(src_pct, tgt_pct)
                        if max_pct >= exclusion_threshold:
                            excluded_ids.append(preview['segment'])
                            heavy_excluded.append({
                                "file": filename, "segment": preview['segment'],
                                "pct": max_pct,
                                "source_after": src_after,
                                "target_after": tgt_after
                            })
                            continue
            
            if dedup_tmx:
                if segment_key in dedup_discard_keys:
                    dedup_count += 1
                    dedup_details.append({
                        "file": filename, "segment": preview['segment'],
                        "source": source_text, "target": target_text,
                        "similarity": 100.0
                    })
                    continue
                
                seg_pair = (source_text, target_text)
                if dedup_threshold >= 100:
                    if seg_pair in seen_exact:
                        dedup_count += 1
                        dedup_details.append({
                            "file": filename, "segment": preview['segment'],
                            "source": source_text, "target": target_text,
                            "similarity": 100.0
                        })
                        continue
                    seen_exact.add(seg_pair)
                else:
                    combined = source_text + " ||| " + target_text
                    is_dup = False
                    match_ratio = 0.0
                    threshold_ratio = dedup_threshold / 100.0
                    for seen_combined in seen_segments:
                        ratio = difflib.SequenceMatcher(None, combined, seen_combined).ratio()
                        if ratio >= threshold_ratio:
                            is_dup = True
                            match_ratio = ratio * 100
                            break
                    if is_dup:
                        dedup_count += 1
                        dedup_details.append({
                            "file": filename, "segment": preview['segment'],
                            "source": source_text, "target": target_text,
                            "similarity": match_ratio
                        })
                        continue
                    seen_segments.append(combined)
            
            tu = etree.SubElement(body, "tu")
            
            existing_cid = None
            if existing_canonical_map:
                existing_cid = existing_canonical_map.get((filename, preview['segment']))
            
            if existing_cid:
                doc_prop = etree.SubElement(tu, "prop", type="x-document")
                doc_prop.text = existing_cid
            elif canonical_id:
                doc_prop = etree.SubElement(tu, "prop", type="x-document")
                doc_prop.text = canonical_id
            
            tuv_src = etree.SubElement(tu, "tuv")
            tuv_src.set("{http://www.w3.org/XML/1998/namespace}lang", src_lang)
            seg_src = etree.SubElement(tuv_src, "seg")
            seg_src.text = source_text
            
            tuv_tgt = etree.SubElement(tu, "tuv")
            tuv_tgt.set("{http://www.w3.org/XML/1998/namespace}lang", tgt_lang)
            seg_tgt = etree.SubElement(tuv_tgt, "seg")
            seg_tgt.text = target_text
            
            valid_count += 1
    
    exclusion_breakdown = {
        "short": short_excluded,
        "short_anon": short_anon_excluded,
        "empty": empty_excluded,
        "heavy": heavy_excluded,
        "dedup": dedup_details
    }
    
    result = b'<?xml version="1.0" encoding="UTF-8"?>\n'
    result += etree.tostring(tmx_root, encoding="unicode", pretty_print=True).encode("utf-8")
    
    return result, valid_count, no_anon_skipped, excluded_ids, dedup_count, dedup_details, exclusion_breakdown


def generate_changes_excel(dedup_details: list = None, exclusion_breakdown: dict = None,
                           file_canonical_map: dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anonymization Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a5488", end_color="1a5488", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1a5488")
    border = Border(
        left=Side(style='thin', color='bcbdbe'),
        right=Side(style='thin', color='bcbdbe'),
        top=Side(style='thin', color='bcbdbe'),
        bottom=Side(style='thin', color='bcbdbe')
    )
    alt_fill = PatternFill(start_color="e8f4fc", end_color="e8f4fc", fill_type="solid")
    total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    total_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="d6e4f0", end_color="d6e4f0", fill_type="solid")
    section_font = Font(bold=True, size=12, color="1a5488")
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "ANONYMIZATION REPORT - MQXLIFF/TMX Anonymizer v6.2"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    if file_canonical_map:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "SOURCE FILES & CANONICAL IDs"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'B{row}'].fill = section_fill
        ws[f'C{row}'].fill = section_fill
        row += 1
        for fname, cids in file_canonical_map.items():
            ws.cell(row=row, column=1, value=fname).border = border
            if cids == "None":
                ws.merge_cells(f'B{row}:C{row}')
            ws.cell(row=row, column=2, value=cids).border = border
            ws.cell(row=row, column=3).border = border
            row += 1
        row += 1
    
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "STATISTICS SUMMARY"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    
    row += 1
    stats_headers = ["File", "Safe Regex", "Regex CT IDs", "Presidio", "ScispaCy", "Proper Names", "Dictionary", "Total"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    grand_total = 0
    for filename, stats in st.session_state.all_stats.items():
        file_total = stats.get("safe_regex", 0) + stats.get("regex_ct", 0) + stats.get("presidio_pii", 0) + stats.get("biomedical", 0) + stats.get("proper_names", 0) + stats.get("dictionary", 0)
        grand_total += file_total
        data = [filename, stats.get("safe_regex", 0), stats["regex_ct"], stats["presidio_pii"], stats["biomedical"], stats.get("proper_names", 0), stats["dictionary"], file_total]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
        row += 1
    
    total_data = ["GRAND TOTAL", "", "", "", "", "", "", grand_total]
    for col, value in enumerate(total_data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='center')
    
    row += 3
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "CHANGES DETAIL"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    
    row += 1
    detail_headers = ["File", "Segment", "Type", "Original Text", "Anonymized Text"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    alt_row = False
    filter_junk = st.session_state.get('filter_junk', False)
    min_words_junk = st.session_state.get('min_words_junk', 2)
    filter_short = st.session_state.get('filter_short_segments', False)
    min_words = st.session_state.get('min_words', 5)
    
    for filename, file_previews in st.session_state.previews.items():
        for preview in file_previews:
            if filter_junk and is_junk_segment(preview, min_words_junk):
                continue
            if not preview.get('changed', True):
                continue
            if filter_short and segment_word_count(preview) < min_words:
                continue
            
            if preview['source_before'] != preview['source_after']:
                data = [filename, preview['segment'], "Source", preview['source_before'], preview['source_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
            if preview['target_before'] != preview['target_after']:
                data = [filename, preview['segment'], "Target", preview['target_before'], preview['target_after']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if alt_row:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
                alt_row = not alt_row
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    ws2 = wb.create_sheet(title="Filtered Segments")
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "FILTERED SEGMENTS - Anonymizer v6.2"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ex_row = 3
    
    if exclusion_breakdown is None:
        exclusion_breakdown = {"short": [], "short_anon": [], "empty": [], "heavy": [], "dedup": []}
    
    bd_short = exclusion_breakdown.get("short", [])
    bd_short_anon = exclusion_breakdown.get("short_anon", [])
    bd_empty = exclusion_breakdown.get("empty", [])
    bd_heavy = exclusion_breakdown.get("heavy", [])
    bd_dedup = exclusion_breakdown.get("dedup", [])
    
    junk_fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
    short_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    empty_fill = PatternFill(start_color="d6d8db", end_color="d6d8db", fill_type="solid")
    tm_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    dedup_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    
    threshold = st.session_state.get('exclusion_threshold', 20)
    dedup_threshold_val = st.session_state.get('dedup_threshold', 100)
    has_any = len(bd_short) + len(bd_short_anon) + len(bd_empty) + len(bd_heavy) + len(bd_dedup) > 0
    
    if bd_short:
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT SEGMENTS - Excluded from TM (<{min_words_junk} words or only numbers/symbols)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="495057")
        ws2[f'A{ex_row}'].fill = junk_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Source (original)", "Target (original)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_short:
            for col, value in enumerate([item["file"], item["segment"], item["source"], item["target"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:D{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_short)} short segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = junk_fill
        ex_row += 1
        ex_row += 2
    
    if bd_short_anon:
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"SHORT ANONYMIZED SEGMENTS - Excluded from TM (less than {min_words} words)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="856404")
        ws2[f'A{ex_row}'].fill = short_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Source (original)", "Target (original)", "Source (anonymized)", "Target (anonymized)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_short_anon:
            for col, value in enumerate([item["file"], item["segment"], item["source_before"], item["target_before"], item["source_after"], item["target_after"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="fffbe6", end_color="fffbe6", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:F{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_short_anon)} short anonymized segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = short_fill
        ex_row += 1
        ex_row += 2
    
    if bd_empty:
        ws2.merge_cells(f'A{ex_row}:B{ex_row}')
        ws2[f'A{ex_row}'] = "EMPTY SEGMENTS - Excluded from TM (no text content)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="495057")
        ws2[f'A{ex_row}'].fill = empty_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_empty:
            ws2.cell(row=ex_row, column=1, value=item["file"]).border = border
            ws2.cell(row=ex_row, column=2, value=item["segment"]).border = border
            ws2.cell(row=ex_row, column=2).alignment = Alignment(horizontal='center')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:B{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_empty)} empty segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = empty_fill
        ex_row += 1
        ex_row += 2
    
    if bd_heavy:
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"HEAVILY ANONYMIZED SEGMENTS (redaction >= {threshold}%)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="721c24")
        ws2[f'A{ex_row}'].fill = tm_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Redaction %", "Source (anonymized)", "Target (anonymized)"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for item in bd_heavy:
            for col, value in enumerate([item["file"], item["segment"], f"{item['pct']:.1f}%", item["source_after"], item["target_after"]], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_heavy)} heavily anonymized segments excluded from TM"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = tm_fill
        ex_row += 1
        ex_row += 2
    
    if bd_dedup:
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"DUPLICATE SEGMENTS - Excluded from TMX (similarity >= {dedup_threshold_val}%)"
        ws2[f'A{ex_row}'].font = Font(bold=True, size=12, color="0c5460")
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
        for col, header in enumerate(["File", "Segment", "Similarity", "Source", "Target"], 1):
            cell = ws2.cell(row=ex_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        ex_row += 1
        for dd in bd_dedup:
            for col, value in enumerate([dd['file'], dd['segment'], f"{dd['similarity']:.1f}%", dd['source'], dd['target']], 1):
                cell = ws2.cell(row=ex_row, column=col, value=value)
                cell.border = border
                cell.fill = PatternFill(start_color="e8f6f8", end_color="e8f6f8", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ex_row += 1
        ws2.merge_cells(f'A{ex_row}:E{ex_row}')
        ws2[f'A{ex_row}'] = f"Total: {len(bd_dedup)} duplicate segments excluded from TMX"
        ws2[f'A{ex_row}'].font = total_font
        ws2[f'A{ex_row}'].fill = dedup_fill
        ex_row += 1
    
    if not has_any:
        ws2.cell(row=ex_row, column=1, value="No segments were excluded.").font = Font(italic=True, color="666666")
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 45
    ws2.column_dimensions['F'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0e7bc0 0%, #1a5488 100%); 
                padding: 2rem; 
                border-radius: 12px; 
                margin-bottom: 2rem;
                box-shadow: 0 4px 15px rgba(26, 84, 136, 0.3);">
        <h1 style="color: white !important; margin: 0; font-size: 2.5rem; font-weight: 700; letter-spacing: 8px;">
            Anonymizer <span style="background: rgba(255,255,255,0.2); color: rgba(255,255,255,0.95); font-size: 0.9rem; font-weight: 500; padding: 3px 10px; border-radius: 20px; vertical-align: middle; letter-spacing: 1px; border: 1px solid rgba(255,255,255,0.3);">v6.2</span>
        </h1>
        <p style="color: rgba(255,255,255,0.95) !important; margin: 0.5rem 0 0 0; font-size: 1.1rem; font-weight: 400; letter-spacing: 1px;">
            Anonymize & clean bilingual memoQ documents
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        replacement_token = st.text_input("Replacement token", value="███")
        st.session_state['replacement_token'] = replacement_token
        process_source = True
        process_target = True
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Multilingual layers")
        use_safe_regex = st.checkbox("Safe Regex", value=True, help="Emails, phones, URLs, IDs, addresses, titled names, etc.")
        use_proper_names = st.checkbox("Proper Names", value=False, help="Structural person name detection (labels, initials, scoring)")
        use_dictionary = st.checkbox("Custom dictionary", value=True, help="Blacklist: manually loaded terms to anonymize")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### EN > ES layers")
        use_regex = st.checkbox("Clinical ID Regex", value=False, help="NCT IDs, EudraCT, Protocol IDs, Subject IDs, etc.")
        use_presidio = st.checkbox("Presidio", value=False, help="Emails, phone numbers, person names, addresses")
        use_biomedical = st.checkbox("ScispaCy", value=False, help="Drugs, pharmaceutical organizations")
        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
        st.markdown("#### Filters")
        dedup_tmx = st.checkbox("Deduplicate TMX segments", value=True, help="Remove duplicate or similar segments from the clean TMX")
        dedup_threshold = st.slider("Similarity threshold (≥%)", 50, 100, 100, 5, help="100% = exact match (recommended for large files). Lower values use fuzzy matching.") if dedup_tmx else 100
        st.session_state['dedup_tmx'] = dedup_tmx
        st.session_state['dedup_threshold'] = dedup_threshold
        filter_junk = st.checkbox("Exclude short segments", value=True, help="Removes short segments: less than min. words or only numbers/symbols")
        min_words_junk = st.slider("Minimum words (short segments)", 2, 10, 2, 1) if filter_junk else 2
        st.session_state['filter_junk'] = filter_junk
        st.session_state['min_words_junk'] = min_words_junk
        filter_short_segments = st.checkbox("Exclude short anon. segments", value=True, help="Excludes segments that are too short after anonymization")
        min_words = st.slider("Minimum words (anon. segments)", 2, 10, 5, 1) if filter_short_segments else 5
        st.session_state['filter_short_segments'] = filter_short_segments
        st.session_state['min_words'] = min_words
        exclude_modified_targets = st.checkbox("Exclude heavily anonymized", value=True)
        if exclude_modified_targets:
            exclusion_threshold = st.slider("Threshold (≥%)", 10, 90, 50, 5, help="Segments above this % anonymized are excluded")
            exclude_source_too = True
        else:
            exclusion_threshold = 20
            exclude_source_too = False
        st.session_state['exclusion_threshold'] = exclusion_threshold
        st.session_state['exclude_source_too'] = exclude_source_too
        st.session_state['exclude_modified_targets'] = exclude_modified_targets
    
    tab1, tab2, tab_dedup, tab_qa, tab3 = st.tabs(["📤 Upload", "📝 Preview", "🔄 Duplicates", "🛡️ QA Check", "📥 Download"])
    
    with tab1:
        st.markdown("### Upload MQXLIFF / TMX files")
        
        mqxliff_files = st.file_uploader(
            "Select one or more .mqxliff or .tmx files",
            type=["mqxliff", "tmx"],
            accept_multiple_files=True,
            help="You can upload multiple files for batch processing"
        )
        
        if mqxliff_files:
            st.success(f"✅ {len(mqxliff_files)} file(s) loaded")
            for f in mqxliff_files:
                ext = f.name.rsplit(".", 1)[-1].upper() if "." in f.name else "?"
                st.write(f"- {f.name} ({f.size / 1024:.1f} KB) — {ext}")
        
        st.markdown("---")
        st.markdown("### Custom dictionary (blacklist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will be anonymized:</strong><br>
            These terms will be forcefully anonymized even if not detected automatically.<br>
            One term per line or separated by commas:<br>
            <code>Example</code><br>
            <code>Example project, info@company.com</code><br>
            <code>Brand®, www.example.com</code>
        </div>
        """, unsafe_allow_html=True)
        
        dictionary_file = st.file_uploader(
            "Upload TXT file with sensitive terms (optional)",
            type=["txt"],
            help="Custom terms to anonymize, applied to all active layers"
        )
        
        dictionary_terms = set()
        if dictionary_file:
            dict_raw = dictionary_file.read()
            if dict_raw[:3] == b'\xef\xbb\xbf':
                dict_raw = dict_raw[3:]
            elif dict_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                dict_raw = dict_raw.decode('utf-16').encode('utf-8')
            try:
                content = dict_raw.decode("utf-8")
            except UnicodeDecodeError:
                content = dict_raw.decode("latin-1")
            dictionary_terms = load_dictionary_terms(content)
            st.success(f"✅ {len(dictionary_terms)} unique terms loaded")
            
            with st.expander("View loaded terms"):
                for term in sorted(dictionary_terms):
                    st.write(f"- {term}")
        
        st.markdown("### Protected terms (whitelist)")
        
        st.markdown("""
        <div class="info-box">
            <strong>ℹ️ Terms that will NOT be anonymized:</strong><br>
            These terms will be preserved even if detected.<br>
            One term per line or separated by commas.
        </div>
        """, unsafe_allow_html=True)
        
        whitelist_file = st.file_uploader(
            "Upload TXT file with protected terms (optional)",
            type=["txt"],
            help="Terms that should never be anonymized",
            key="whitelist_uploader"
        )
        
        whitelist_terms = set()
        if whitelist_file:
            wl_raw = whitelist_file.read()
            if wl_raw[:3] == b'\xef\xbb\xbf':
                wl_raw = wl_raw[3:]
            elif wl_raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
                wl_raw = wl_raw.decode('utf-16').encode('utf-8')
            try:
                wl_content = wl_raw.decode("utf-8")
            except UnicodeDecodeError:
                wl_content = wl_raw.decode("latin-1")
            whitelist_terms = load_dictionary_terms(wl_content)
            st.success(f"🛡️ {len(whitelist_terms)} protected terms loaded")
            
            with st.expander("View protected terms"):
                for term in sorted(whitelist_terms):
                    st.write(f"- {term}")
        
        if mqxliff_files:
            st.markdown("---")
            if st.button("🚀 Process files", type="primary", use_container_width=True):
                process_files(
                    mqxliff_files, replacement_token, process_source, process_target,
                    use_safe_regex, use_regex, use_presidio, use_biomedical, use_proper_names,
                    use_dictionary, dictionary_terms, whitelist_terms
                )
    
    with tab2:
        if "previews" in st.session_state and st.session_state.previews:
            st.markdown("### Changes preview")
            
            col_search, col_show_junk = st.columns([3, 1])
            with col_search:
                search_term = st.text_input(
                    "🔍 Search in preview",
                    placeholder="Type to filter changes...",
                    key="preview_search"
                )
            with col_show_junk:
                show_junk_in_preview = st.checkbox(
                    "Show short",
                    value=False,
                    key="show_junk_preview",
                    help="Show/hide short and short anon. segments in the preview (does not affect downloads)"
                )
            
            replacement_token = st.session_state.get('replacement_token', '███')
            threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_enabled = st.session_state.get('exclude_modified_targets', False)
            exclude_source = st.session_state.get('exclude_source_too', False)
            
            if 'excluded_segments' not in st.session_state:
                st.session_state['excluded_segments'] = {}
            if 'no_anon_segments' not in st.session_state:
                st.session_state['no_anon_segments'] = {}
            if 'skip_junk_segments' not in st.session_state:
                st.session_state['skip_junk_segments'] = {}
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            
            total_matches = 0
            candidates_count = 0
            candidate_keys = set()
            junk_segments_count = 0
            short_segments_count = 0
            file_index = 0
            
            for filename, file_previews in st.session_state.previews.items():
                filtered_previews = []
                for preview in file_previews:
                    is_junk = is_junk_segment(preview, min_words_junk)
                    segment_key_junk = f"{filename}_{preview['segment']}"
                    
                    if is_junk:
                        is_skipped = st.session_state.get(f"skipjunk_{segment_key_junk}", False)
                        if not is_skipped:
                            junk_segments_count += 1
                        if show_junk_in_preview:
                            preview['_is_junk'] = True
                            preview['_junk_skipped'] = is_skipped
                            filtered_previews.append(preview)
                        continue
                    
                    preview['_is_junk'] = False
                    is_short_anon = filter_short and segment_word_count(preview) < min_words
                    if is_short_anon:
                        is_short_skipped = st.session_state.get(f"skipshort_{filename}_{preview['segment']}", False)
                        if not is_short_skipped:
                            short_segments_count += 1
                        preview['_is_short_anon'] = True
                        preview['_short_anon_skipped'] = is_short_skipped
                        if show_junk_in_preview:
                            filtered_previews.append(preview)
                        continue
                    
                    if not preview.get('changed', True):
                        continue
                    
                    has_token = (replacement_token in preview.get('source_after', '') or 
                                replacement_token in preview.get('target_after', ''))
                    if not has_token:
                        continue
                    
                    if search_term:
                        search_lower = search_term.lower()
                        if (search_lower in preview['source_before'].lower() or
                            search_lower in preview['source_after'].lower() or
                            search_lower in preview['target_before'].lower() or
                            search_lower in preview['target_after'].lower()):
                            filtered_previews.append(preview)
                    else:
                        filtered_previews.append(preview)
                
                non_junk_count = sum(1 for p in filtered_previews if not p.get('_is_junk', False) and not p.get('_is_short_anon', False))
                junk_in_file = sum(1 for p in filtered_previews if p.get('_is_junk', False))
                short_anon_in_file = sum(1 for p in filtered_previews if p.get('_is_short_anon', False))
                
                if filtered_previews:
                    if file_index > 0:
                        st.markdown('<hr style="border:none;border-top:2px solid #1a5488;margin:1.5rem 0;">', unsafe_allow_html=True)
                    file_index += 1
                    label_parts = []
                    if non_junk_count > 0:
                        label_parts.append(f"{non_junk_count} affected")
                    if junk_in_file > 0:
                        label_parts.append(f"{junk_in_file} short")
                    if short_anon_in_file > 0:
                        label_parts.append(f"{short_anon_in_file} short anon.")
                    total_matches += non_junk_count
                    with st.expander(f"📄 {filename} ({', '.join(label_parts)} segments)", expanded=True):
                        for preview in filtered_previews:
                            preview_is_junk = preview.get('_is_junk', False)
                            segment_key = f"{filename}_{preview['segment']}"
                            
                            if preview_is_junk:
                                junk_skipped = preview.get('_junk_skipped', False)
                                col_header, col_skip_junk = st.columns([4, 1])
                                with col_header:
                                    if junk_skipped:
                                        st.markdown(
                                            f'<div style="background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#155724;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                            f'<br><small>Source: {preview["source_before"][:80] or "(empty)"} | Target: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        st.markdown(
                                            f'<div style="background:#e2e3e5;border:1px solid #6c757d;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#495057;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#6c757d;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Short</span>'
                                            f'<br><small>Source: {preview["source_before"][:80] or "(empty)"} | Target: {preview["target_before"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                with col_skip_junk:
                                    cb_key = f"skipjunk_{segment_key}"
                                    if cb_key not in st.session_state:
                                        st.session_state[cb_key] = False
                                    skip_junk = st.checkbox(
                                        "Skip",
                                        key=cb_key,
                                        help="Keep this segment (don't exclude as short)"
                                    )
                                    st.session_state['skip_junk_segments'][segment_key] = skip_junk
                                continue
                            
                            preview_is_short_anon = preview.get('_is_short_anon', False)
                            if preview_is_short_anon:
                                short_anon_skipped = preview.get('_short_anon_skipped', False)
                                col_header, col_skip_short = st.columns([4, 1])
                                with col_header:
                                    if short_anon_skipped:
                                        st.markdown(
                                            f'<div style="background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#155724;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                            f'<br><small>Source: {preview["source_after"][:80] or "(empty)"} | Target: {preview["target_after"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                    else:
                                        st.markdown(
                                            f'<div style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;color:#856404;">'
                                            f'<strong>Segment {preview["segment"]}</strong> '
                                            f'<span style="background:#e0a800;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✂️ Short anon.</span>'
                                            f'<br><small>Source: {preview["source_after"][:80] or "(empty)"} | Target: {preview["target_after"][:80] or "(empty)"}</small>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                                with col_skip_short:
                                    cb_key = f"skipshort_{segment_key}"
                                    if cb_key not in st.session_state:
                                        st.session_state[cb_key] = False
                                    skip_short = st.checkbox(
                                        "Skip",
                                        key=cb_key,
                                        help="Keep this segment (don't exclude as short anon.)"
                                    )
                                    if 'skip_short_segments' not in st.session_state:
                                        st.session_state['skip_short_segments'] = {}
                                    st.session_state['skip_short_segments'][segment_key] = skip_short
                                continue
                            
                            src_before = preview.get('source_before', '')
                            tgt_before = preview.get('target_before', '')
                            src_after = preview.get('source_after', '')
                            tgt_after = preview.get('target_after', '')
                            if src_before:
                                source_pct = (1 - len(src_after.replace(replacement_token, '')) / max(len(src_before), 1)) * 100
                            else:
                                source_pct = 0
                            if tgt_before:
                                target_pct = (1 - len(tgt_after.replace(replacement_token, '')) / max(len(tgt_before), 1)) * 100
                            else:
                                target_pct = 0
                            max_pct = max(source_pct, target_pct)
                            is_candidate = max_pct >= threshold
                            
                            if is_candidate:
                                candidates_count += 1
                                candidate_keys.add(segment_key)
                            
                            has_changes = (preview['source_before'] != preview['source_after'] or 
                                         preview['target_before'] != preview['target_after'])
                            
                            col_header, col_no_anon, col_exclude = st.columns([3, 1, 1])
                            with col_header:
                                header_placeholder = st.empty()
                            
                            with col_no_anon:
                                if has_changes:
                                    no_anon_val = st.session_state['no_anon_segments'].get(segment_key, False)
                                    no_anon = st.checkbox(
                                        "Skip",
                                        value=no_anon_val,
                                        key=f"noanon_{segment_key}",
                                        help="Keep original text without anonymizing"
                                    )
                                    st.session_state['no_anon_segments'][segment_key] = no_anon
                            
                            with col_exclude:
                                if is_candidate and exclude_enabled:
                                    default_val = st.session_state['excluded_segments'].get(segment_key, True)
                                    exclude_this = st.checkbox(
                                        "Exclude TM",
                                        value=default_val,
                                        key=f"excl_{segment_key}"
                                    )
                                    st.session_state['excluded_segments'][segment_key] = exclude_this
                            
                            is_no_anon = st.session_state['no_anon_segments'].get(segment_key, False)
                            
                            segment_header = f"**Segment {preview['segment']}**"
                            if is_no_anon:
                                segment_header += ' <span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Not anonymized</span>'
                            elif is_candidate and exclude_enabled:
                                segment_header += f' <span class="exclude-badge">⚠️ {max_pct:.0f}% → exclusion</span>'
                            elif max_pct > 0:
                                segment_header += f' <span style="background:#c8ccd0;color:#333;padding:2px 6px;border-radius:4px;font-size:0.75rem;">{max_pct:.0f}% anonymized</span>'
                            header_placeholder.markdown(segment_header, unsafe_allow_html=True)
                            
                            if preview['source_before'] != preview['source_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Source - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["source_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Source - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["source_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label_src = "**Source - After:**"
                                        show_excluded = is_candidate and exclude_enabled and exclude_source and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label_src = "**Source - After** _(will be excluded)_:"
                                        st.markdown(label_src)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["source_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            if preview['target_before'] != preview['target_after']:
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Target - Before:**")
                                    st.markdown(f'<div class="before-text">{preview["target_before"]}</div>', 
                                              unsafe_allow_html=True)
                                with col2:
                                    if is_no_anon:
                                        st.markdown("**Target - After** _(no changes)_:")
                                        st.markdown(f'<div class="before-text" style="border-color:#28a745;">{preview["target_before"]}</div>', 
                                                  unsafe_allow_html=True)
                                    else:
                                        label = "**Target - After:**"
                                        show_excluded = is_candidate and exclude_enabled and st.session_state['excluded_segments'].get(segment_key, True)
                                        if show_excluded:
                                            label = "**Target - After** _(will be excluded)_:"
                                        st.markdown(label)
                                        css_class = "after-text excluded-segment" if show_excluded else "after-text"
                                        st.markdown(f'<div class="{css_class}">{preview["target_after"]}</div>', 
                                                  unsafe_allow_html=True)
                            
                            st.markdown("---")
            
            if filter_junk and junk_segments_count > 0:
                st.info(f"ℹ️ {junk_segments_count} short segments excluded from TM (<{min_words_junk} words or only numbers/symbols)")
            
            if filter_short and short_segments_count > 0:
                st.info(f"ℹ️ {short_segments_count} short anonymized segments excluded from TM (less than {min_words} words)")
            
            if exclude_enabled and candidates_count > 0:
                excluded_count = sum(1 for k, v in st.session_state['excluded_segments'].items() if v and k in candidate_keys)
                st.info(f"ℹ️ {excluded_count} heavily anonymized segments excluded from TM (threshold: {threshold}%)")
            
            if total_matches == 0:
                if search_term:
                    st.warning(f"No changes found containing '{search_term}'")
                else:
                    st.info("No anonymized segments to preview. Check that the appropriate layers are enabled and reprocess the files.")
        else:
            st.info("Upload and process files to see the changes preview")
    
    with tab_dedup:
        if "results" in st.session_state and st.session_state.results:
            previews = st.session_state.get('previews', {})
            dedup_tmx = st.session_state.get('dedup_tmx', True)
            dedup_threshold = st.session_state.get('dedup_threshold', 100)
            no_anon_segments = st.session_state.get('no_anon_segments', {})
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            replacement_token = st.session_state.get('replacement_token', '███')

            if not dedup_tmx:
                st.info("Deduplication is disabled. Enable it in the sidebar to detect duplicates.")
                st.session_state['dedup_keep'] = {}
                st.session_state['dedup_group_keys'] = {}
                st.session_state['conflict_keep'] = {}
                st.session_state['conflict_group_keys'] = {}
            else:
                dup_groups = detect_duplicate_groups(
                    previews, dedup_threshold, no_anon_segments,
                    filter_junk, min_words_junk, filter_short, min_words
                )

                dedup_used_keys = set()
                for group in dup_groups:
                    for m in group["members"]:
                        dedup_used_keys.add(m["key"])

                if not dup_groups:
                    st.success("No duplicate segments detected.")
                    st.session_state['dedup_keep'] = {}
                    st.session_state['dedup_group_keys'] = {}
                else:
                    st.markdown(f"### 🔄 {len(dup_groups)} duplicate matches found")
                    st.markdown(f"<small>Select which segment to keep in each match. By default, the first occurrence by upload order is kept and later duplicates are discarded from the clean TMX.</small>", unsafe_allow_html=True)

                    old_keep = st.session_state.get('dedup_keep', {})
                    new_keep = {}
                    new_group_keys = {}

                    for g_idx, group in enumerate(dup_groups):
                        members = group["members"]
                        similarity = group["similarity"]

                        group_keys = sorted([m["key"] for m in members])
                        stable_id = "|".join(group_keys)
                        new_group_keys[stable_id] = group_keys

                        if stable_id in old_keep and (old_keep[stable_id] == "__keep_all__" or old_keep[stable_id] in group_keys):
                            new_keep[stable_id] = old_keep[stable_id]
                        else:
                            new_keep[stable_id] = members[0]["key"]

                        with st.expander(
                            f"Match {g_idx + 1}: {len(members)} segments — {similarity:.0f}% similarity",
                            expanded=(g_idx == 0)
                        ):
                            options = []
                            option_keys = []
                            for m_idx, member in enumerate(members):
                                label = f"📄 {member['file']} — Segment {member['segment']}"
                                options.append(label)
                                option_keys.append(member["key"])
                            options.append("Keep all")
                            option_keys.append("__keep_all__")

                            current_kept = new_keep.get(stable_id, option_keys[0])
                            if current_kept not in option_keys:
                                current_kept = option_keys[0]
                            current_index = option_keys.index(current_kept)

                            selected = st.radio(
                                "Keep:",
                                options=options,
                                index=current_index,
                                key=f"dedup_radio_{stable_id}",
                                horizontal=False
                            )
                            selected_idx = options.index(selected)
                            new_keep[stable_id] = option_keys[selected_idx]

                            keep_all = new_keep[stable_id] == "__keep_all__"

                            ref_source = members[0]["source"][:500]
                            ref_target = members[0]["target"][:500]

                            for m_idx, member in enumerate(members):
                                is_kept = keep_all or (option_keys[m_idx] == new_keep[stable_id])
                                if is_kept:
                                    badge = '<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                else:
                                    badge = '<span style="background:#dc3545;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Discarded</span>'

                                member_sim = member.get("_similarity", similarity)
                                st.markdown(
                                    f'<div style="background:{"#d4edda" if is_kept else "#f8d7da"};border:1px solid {"#28a745" if is_kept else "#dc3545"};border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;">'
                                    f'<strong>📄 {member["file"]} — Segment {member["segment"]}</strong> {badge} '
                                    f'<small style="color:#666;">({member_sim:.0f}% similar)</small>'
                                    f'</div>',
                                    unsafe_allow_html=True
                                )

                                src_text = member["source"][:500]
                                tgt_text = member["target"][:500]
                                if m_idx == 0:
                                    src_html = _html_escape(src_text)
                                    tgt_html = _html_escape(tgt_text)
                                else:
                                    src_html, _ = highlight_diff(ref_source, src_text)
                                    _, tgt_html = highlight_diff(ref_target, tgt_text)

                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Source:**")
                                    st.markdown(
                                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{src_html}</div>',
                                        unsafe_allow_html=True
                                    )
                                with col2:
                                    st.markdown("**Target:**")
                                    st.markdown(
                                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{tgt_html}</div>',
                                        unsafe_allow_html=True
                                    )

                            st.markdown("---")

                    st.session_state['dedup_keep'] = new_keep
                    st.session_state['dedup_group_keys'] = new_group_keys

                st.markdown("---")

                conflict_groups = detect_conflicting_translations(
                    previews, dedup_threshold, no_anon_segments,
                    filter_junk, min_words_junk, filter_short, min_words,
                    dedup_used_keys=dedup_used_keys
                )

                if not conflict_groups:
                    st.success("No conflicting translations detected.")
                    st.session_state['conflict_keep'] = {}
                    st.session_state['conflict_group_keys'] = {}
                else:
                    st.markdown(f"### ⚠️ {len(conflict_groups)} conflicting translation matches found")
                    st.markdown(f"<small>Select which translation to keep in each match. By default, the first occurrence by upload order is kept and later duplicates are discarded from the clean TMX. Segments with similar source (≥{dedup_threshold}% threshold applied to source only) but different translations.</small>", unsafe_allow_html=True)

                    old_conflict_keep = st.session_state.get('conflict_keep', {})
                    new_conflict_keep = {}
                    new_conflict_group_keys = {}

                    for g_idx, group in enumerate(conflict_groups):
                        members = group["members"]
                        similarity = group["similarity"]

                        group_keys = sorted([m["key"] for m in members])
                        stable_id = "|".join(group_keys)
                        new_conflict_group_keys[stable_id] = group_keys

                        if stable_id in old_conflict_keep and (old_conflict_keep[stable_id] == "__keep_all__" or old_conflict_keep[stable_id] in group_keys):
                            new_conflict_keep[stable_id] = old_conflict_keep[stable_id]
                        else:
                            new_conflict_keep[stable_id] = members[0]["key"]

                        with st.expander(
                            f"Match {g_idx + 1}: {len(members)} translations — {similarity:.0f}% source similarity",
                            expanded=(g_idx == 0)
                        ):
                            options = []
                            option_keys = []
                            for m_idx, member in enumerate(members):
                                label = f"📄 {member['file']} — Segment {member['segment']}"
                                options.append(label)
                                option_keys.append(member["key"])
                            options.append("Keep all")
                            option_keys.append("__keep_all__")

                            current_kept = new_conflict_keep.get(stable_id, option_keys[0])
                            if current_kept not in option_keys:
                                current_kept = option_keys[0]
                            current_index = option_keys.index(current_kept)

                            selected = st.radio(
                                "Keep:",
                                options=options,
                                index=current_index,
                                key=f"conflict_radio_{stable_id}",
                                horizontal=False
                            )
                            selected_idx = options.index(selected)
                            new_conflict_keep[stable_id] = option_keys[selected_idx]

                            keep_all = new_conflict_keep[stable_id] == "__keep_all__"

                            ref_source = members[0]["source"][:500]
                            ref_target = members[0]["target"][:500]

                            for m_idx, member in enumerate(members):
                                is_kept = keep_all or (option_keys[m_idx] == new_conflict_keep[stable_id])
                                if is_kept:
                                    badge = '<span style="background:#28a745;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">✓ Kept</span>'
                                else:
                                    badge = '<span style="background:#dc3545;color:white;padding:2px 6px;border-radius:4px;font-size:0.75rem;">🗑️ Discarded</span>'

                                member_sim = member.get("_similarity", similarity)
                                st.markdown(
                                    f'<div style="background:{"#d4edda" if is_kept else "#f8d7da"};border:1px solid {"#28a745" if is_kept else "#dc3545"};border-radius:6px;padding:0.5rem 0.8rem;margin:0.3rem 0;">'
                                    f'<strong>📄 {member["file"]} — Segment {member["segment"]}</strong> {badge} '
                                    f'<small style="color:#666;">({member_sim:.0f}% similar)</small>'
                                    f'</div>',
                                    unsafe_allow_html=True
                                )

                                src_text = member["source"][:500]
                                tgt_text = member["target"][:500]
                                if m_idx == 0:
                                    src_html = _html_escape(src_text)
                                    tgt_html = _html_escape(tgt_text)
                                else:
                                    src_html, _ = highlight_diff(ref_source, src_text)
                                    _, tgt_html = highlight_diff(ref_target, tgt_text)

                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**Source:**")
                                    st.markdown(
                                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{src_html}</div>',
                                        unsafe_allow_html=True
                                    )
                                with col2:
                                    st.markdown("**Target:**")
                                    st.markdown(
                                        f'<div style="background:#f8f9fa;border:1px solid #dee2e6;border-radius:4px;padding:0.5rem;font-size:0.9rem;">{tgt_html}</div>',
                                        unsafe_allow_html=True
                                    )

                            st.markdown("---")

                    st.session_state['conflict_keep'] = new_conflict_keep
                    st.session_state['conflict_group_keys'] = new_conflict_group_keys
        else:
            st.info("Process the files to see duplicate analysis")

    with tab_qa:
        _render_qa_check_tab()

    with tab3:
        if "results" in st.session_state and st.session_state.results:
            st.markdown("### Anonymization statistics")
            
            total_stats = {"safe_regex": 0, "regex_ct": 0, "presidio_pii": 0, "biomedical": 0, "proper_names": 0, "dictionary": 0}
            for stats in st.session_state.all_stats.values():
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)
            
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            render_stat_card("Safe Regex", total_stats["safe_regex"], col1, "stat-card-safe-regex")
            render_stat_card("Regex CT IDs", total_stats["regex_ct"], col2, "stat-card-regex-ct")
            render_stat_card("Presidio", total_stats["presidio_pii"], col3, "stat-card-presidio")
            render_stat_card("Biomedical", total_stats["biomedical"], col4, "stat-card-biomedical")
            render_stat_card("Proper Names", total_stats["proper_names"], col5, "stat-card-proper-names")
            render_stat_card("Dictionary", total_stats["dictionary"], col6, "stat-card-dictionary")
            
            st.markdown("---")
            st.markdown("### Download anonymized files")
            
            exclude_modified_targets = st.session_state.get('exclude_modified_targets', False)
            exclusion_threshold = st.session_state.get('exclusion_threshold', 20)
            exclude_source_too = st.session_state.get('exclude_source_too', False)
            excluded_segments = st.session_state.get('excluded_segments', {})
            
            def get_junk_segment_keys(filename: str, file_previews: list, min_words_junk: int = 2) -> set:
                """Gets the keys of junk segments that should be cleared."""
                junk_keys = set()
                for preview in file_previews:
                    if is_junk_segment(preview, min_words_junk):
                        seg_key = f"{filename}_{preview['segment']}"
                        if not st.session_state.get(f"skipjunk_{seg_key}", False):
                            junk_keys.add(seg_key)
                return junk_keys
            
            def get_short_segment_keys(filename: str, file_previews: list, min_words: int) -> set:
                """Gets the keys of short segments that should be restored."""
                short_keys = set()
                for preview in file_previews:
                    if segment_word_count(preview) < min_words:
                        short_keys.add(f"{filename}_{preview['segment']}")
                return short_keys
            
            def apply_no_anon_segments(anon_content: bytes, orig_content: bytes, filename: str, no_anon_segs: dict) -> bytes:
                if not no_anon_segs:
                    return anon_content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    anon_tree = etree.fromstring(anon_content)
                    orig_tree = etree.fromstring(orig_content)
                    
                    if is_tmx:
                        anon_units = anon_tree.xpath('//tu')
                        orig_units = orig_tree.xpath('//tu')
                    else:
                        nsmap = anon_tree.nsmap
                        default_ns = nsmap.get(None, '')
                        
                        if default_ns:
                            ns = {'x': default_ns}
                            anon_units = anon_tree.xpath('//x:trans-unit', namespaces=ns)
                            orig_units = orig_tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            anon_units = anon_tree.xpath('//trans-unit')
                            orig_units = orig_tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, (anon_tu, orig_tu) in enumerate(zip(anon_units, orig_units)):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        should_restore_only = no_anon_segs.get(segment_key, False)
                        
                        if should_restore_only:
                            if is_tmx:
                                anon_tuvs = anon_tu.xpath('tuv')
                                orig_tuvs = orig_tu.xpath('tuv')
                                for anon_tuv, orig_tuv in zip(anon_tuvs, orig_tuvs):
                                    anon_segs = anon_tuv.xpath('seg')
                                    orig_segs = orig_tuv.xpath('seg')
                                    for anon_seg, orig_seg in zip(anon_segs, orig_segs):
                                        anon_seg.getparent().replace(anon_seg, orig_seg)
                            else:
                                if default_ns:
                                    anon_sources = anon_tu.xpath('.//x:source', namespaces=ns)
                                    orig_sources = orig_tu.xpath('.//x:source', namespaces=ns)
                                    anon_targets = anon_tu.xpath('.//x:target', namespaces=ns)
                                    orig_targets = orig_tu.xpath('.//x:target', namespaces=ns)
                                else:
                                    anon_sources = anon_tu.xpath('.//source')
                                    orig_sources = orig_tu.xpath('.//source')
                                    anon_targets = anon_tu.xpath('.//target')
                                    orig_targets = orig_tu.xpath('.//target')
                                
                                for anon_src, orig_src in zip(anon_sources, orig_sources):
                                    anon_src.getparent().replace(anon_src, orig_src)
                                
                                for anon_tgt, orig_tgt in zip(anon_targets, orig_targets):
                                    anon_tgt.getparent().replace(anon_tgt, orig_tgt)
                    
                    return etree.tostring(anon_tree, encoding='utf-8', xml_declaration=True)
                except Exception:
                    # Malformed/encrypted XML — fall back to the anonymized
                    # content as-is so the user still gets a usable file.
                    return anon_content
            
            def _clear_element(elem):
                for child in list(elem):
                    elem.remove(child)
                elem.text = None
                elem.tail = None

            def prepare_download_content(content: bytes, filename: str, exclude_targets: bool, threshold: float, exclude_source: bool, excluded_segs: dict, short_segs: set = None, junk_segs: set = None) -> bytes:
                if short_segs is None:
                    short_segs = set()
                if junk_segs is None:
                    junk_segs = set()
                
                if not exclude_targets and not short_segs and not junk_segs:
                    return content
                
                from lxml import etree
                is_tmx = filename.lower().endswith(".tmx")
                
                try:
                    tree = etree.fromstring(content)
                    
                    if is_tmx:
                        trans_units = tree.xpath('//tu')
                    else:
                        nsmap = tree.nsmap
                        default_ns = nsmap.get(None, '')
                        if default_ns:
                            ns = {'x': default_ns}
                            trans_units = tree.xpath('//x:trans-unit', namespaces=ns)
                        else:
                            trans_units = tree.xpath('//trans-unit')
                    
                    replacement_token = st.session_state.get('replacement_token', '███')
                    
                    for idx, tu in enumerate(trans_units):
                        segment_num = idx + 1
                        segment_key = f"{filename}_{segment_num}"
                        
                        is_short = segment_key in short_segs
                        is_junk = segment_key in junk_segs
                        
                        if is_short or is_junk:
                            if is_tmx:
                                for tuv in tu.xpath('tuv'):
                                    for seg in tuv.xpath('seg'):
                                        _clear_element(seg)
                            else:
                                if default_ns:
                                    targets = tu.xpath('.//x:target', namespaces=ns)
                                    sources = tu.xpath('.//x:source', namespaces=ns)
                                else:
                                    targets = tu.xpath('.//target')
                                    sources = tu.xpath('.//source')
                                for elem in sources + targets:
                                    _clear_element(elem)
                            continue
                        
                        if not exclude_targets:
                            continue
                        
                        should_exclude_this = excluded_segs.get(segment_key, None)
                        if should_exclude_this is False:
                            continue
                        
                        if is_tmx:
                            target_tuvs = []
                            source_tuvs = []
                            for tuv in tu.xpath('tuv'):
                                tuv_lang = tuv.get("{http://www.w3.org/XML/1998/namespace}lang", tuv.get("lang", ""))
                                if tuv_lang.lower().startswith("es"):
                                    target_tuvs.extend(tuv.xpath('seg'))
                                elif tuv_lang.lower().startswith("en") and exclude_source:
                                    source_tuvs.extend(tuv.xpath('seg'))
                            
                            for seg in target_tuvs:
                                text_content = ''.join(seg.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    if redacted_pct >= threshold:
                                        _clear_element(seg)
                                        if exclude_source:
                                            for src_seg in source_tuvs:
                                                _clear_element(src_seg)
                        else:
                            if default_ns:
                                targets = tu.xpath('.//x:target', namespaces=ns)
                                sources = tu.xpath('.//x:source', namespaces=ns) if exclude_source else []
                            else:
                                targets = tu.xpath('.//target')
                                sources = tu.xpath('.//source') if exclude_source else []
                            
                            for target in targets:
                                text_content = ''.join(target.itertext())
                                words = text_content.split()
                                if replacement_token in text_content and words:
                                    redacted_pct = sum(1 for w in words if replacement_token in w) / len(words) * 100
                                    
                                    if redacted_pct >= threshold:
                                        _clear_element(target)
                                        
                                        if exclude_source:
                                            for source in sources:
                                                _clear_element(source)
                    
                    return etree.tostring(tree, encoding='utf-8', xml_declaration=True)
                except Exception:
                    # Malformed/encrypted XML — return the original bytes
                    # unmodified rather than crashing the download.
                    return content
            
            no_anon_segments = st.session_state.get('no_anon_segments', {})
            originals = st.session_state.get('originals', {})
            
            no_anon_count = sum(1 for v in no_anon_segments.values() if v)
            if no_anon_count > 0:
                st.markdown(f"""
                <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>✓ {no_anon_count} segments will keep their original text</strong><br>
                    <small>Marked as "Skip" in the Preview tab</small>
                </div>
                """, unsafe_allow_html=True)
            
            filter_junk = st.session_state.get('filter_junk', False)
            min_words_junk = st.session_state.get('min_words_junk', 2)
            filter_short = st.session_state.get('filter_short_segments', False)
            min_words = st.session_state.get('min_words', 5)
            previews = st.session_state.get('previews', {})
            
            dedup_tmx = st.session_state.get('dedup_tmx', True)
            dedup_threshold = st.session_state.get('dedup_threshold', 100)
            
            existing_canonical_map = extract_existing_canonical_ids(originals)
            existing_canonical_count = len(existing_canonical_map)
            total_preview_segments = sum(len(fp) for fp in previews.values())
            all_have_canonical = existing_canonical_count >= total_preview_segments and existing_canonical_count > 0
            has_some_canonical = existing_canonical_count > 0
            segments_without_canonical = total_preview_segments - existing_canonical_count
            
            st.markdown("---")
            
            canonical_id_value = None
            if all_have_canonical:
                existing_ids = sorted(set(existing_canonical_map.values()))
                ids_display = ", ".join(existing_ids)
                st.markdown(f"""
                <div style="background-color: #d4edda; border: 1px solid #28a745; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>🔒 All {existing_canonical_count} segments already have canonical IDs</strong><br>
                    <small>Existing IDs: {ids_display}</small>
                </div>
                """, unsafe_allow_html=True)
            else:
                if has_some_canonical:
                    existing_ids = sorted(set(existing_canonical_map.values()))
                    ids_display = ", ".join(existing_ids)
                    st.markdown(f"""
                    <div style="background-color: #fff3cd; border: 1px solid #ffc107; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <strong>⚠️ {existing_canonical_count} segments already have canonical IDs ({ids_display})</strong><br>
                        <small>{segments_without_canonical} segments without canonical ID</small>
                    </div>
                    """, unsafe_allow_html=True)
                
                if 'use_canonical' not in st.session_state:
                    st.session_state['use_canonical'] = True
                use_canonical = st.checkbox("Assign canonical ID to TMX segments", key='use_canonical',
                                            help="Each segment will carry a permanent document identifier for traceability")
                
                if use_canonical:
                    if 'canonical_counter' not in st.session_state:
                        st.session_state['canonical_counter'] = 1
                    suggested_id = f"TMX-{date.today().isoformat()}-{st.session_state['canonical_counter']:03d}"
                    canonical_id_value = st.text_input("Canonical ID", value=suggested_id, key='canonical_id_input',
                                                        help="All segments in this TMX will be tagged with this ID")
                    if canonical_id_value:
                        canonical_id_value = canonical_id_value.strip()
                    if not canonical_id_value:
                        st.warning("Please enter a canonical ID or uncheck the option.")
                        canonical_id_value = None
                    elif has_some_canonical:
                        st.info(f"Only {segments_without_canonical} new segments will receive this ID. Existing canonical IDs are preserved.")
            
            tmx_output_name = f"Redacted_TMX_{date.today().isoformat()}.tmx"
            
            clean_tmx_data, valid_segments, no_anon_in_tmx, excluded_ids, dedup_count, dedup_details, exclusion_breakdown = generate_clean_tmx(
                previews=previews,
                results=st.session_state.results,
                originals=originals,
                filter_junk=filter_junk,
                min_words_junk=min_words_junk,
                filter_short=filter_short,
                min_words=min_words,
                exclude_modified=exclude_modified_targets,
                exclusion_threshold=exclusion_threshold,
                excluded_segments=excluded_segments,
                no_anon_segments=no_anon_segments,
                dedup_tmx=dedup_tmx,
                dedup_threshold=dedup_threshold,
                dedup_keep_choices=st.session_state.get('dedup_keep', {}),
                tmx_filename=tmx_output_name,
                canonical_id=canonical_id_value,
                existing_canonical_map=existing_canonical_map
            )
            
            st.session_state['dedup_details'] = dedup_details
            st.session_state['exclusion_breakdown'] = exclusion_breakdown
            
            total_segs = sum(len(fp) for fp in previews.values())
            empty_total = total_segs - valid_segments - no_anon_in_tmx
            if empty_total > 0:
                breakdown_parts = []
                if len(exclusion_breakdown.get("short", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['short'])} short")
                if len(exclusion_breakdown.get("short_anon", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['short_anon'])} short anon.")
                if len(exclusion_breakdown.get("empty", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['empty'])} empty")
                if len(exclusion_breakdown.get("heavy", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['heavy'])} heavily anon.")
                if len(exclusion_breakdown.get("dedup", [])) > 0:
                    breakdown_parts.append(f"{len(exclusion_breakdown['dedup'])} duplicates")
                breakdown_text = " | ".join(breakdown_parts) if breakdown_parts else "Excluded by filters"
                st.markdown(f"""
                <div style="background-color: #e2e3e5; border: 1px solid #6c757d; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                    <strong>🗑️ {empty_total} segments excluded from clean TMX</strong> <small>(of {total_segs} total)</small><br>
                    <small>{breakdown_text}</small>
                </div>
                """, unsafe_allow_html=True)
            
            col_mqxliff, col_tmx_clean, col_excel = st.columns(3)
            
            with col_mqxliff:
                if len(st.session_state.results) == 1:
                    filename, content = list(st.session_state.results.items())[0]
                    orig_content = originals.get(filename, content)
                    junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                    short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                    content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                    download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                    file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                    redacted_filename = f"Redacted_{date.today().isoformat()}.{file_ext}"
                    st.download_button(
                        label=f"📥 Download {file_ext.upper()}",
                        data=download_content,
                        file_name=redacted_filename,
                        mime="application/xml",
                        use_container_width=True
                    )
                else:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                        for idx, (filename, content) in enumerate(st.session_state.results.items()):
                            orig_content = originals.get(filename, content)
                            junk_segs = get_junk_segment_keys(filename, previews.get(filename, []), min_words_junk) if filter_junk else set()
                            short_segs = get_short_segment_keys(filename, previews.get(filename, []), min_words) if filter_short else set()
                            content = apply_no_anon_segments(content, orig_content, filename, no_anon_segments)
                            download_content = prepare_download_content(content, filename, exclude_modified_targets, exclusion_threshold, exclude_source_too, excluded_segments, short_segs, junk_segs)
                            file_ext = filename.rsplit(".", 1)[-1] if "." in filename else "xml"
                            zip_entry_name = f"Redacted_{idx + 1}_{date.today().isoformat()}.{file_ext}"
                            zf.writestr(zip_entry_name, download_content)
                    
                    st.download_button(
                        label=f"📥 Download ZIP ({len(st.session_state.results)} files)",
                        data=zip_buffer.getvalue(),
                        file_name=f"Redacted_{date.today().isoformat()}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            
            with col_tmx_clean:
                tmx_label = "📥 Download clean TMX"
                tmx_help = "TMX without empty, excluded or filtered segments"
                if st.download_button(
                    label=tmx_label,
                    data=clean_tmx_data,
                    file_name=tmx_output_name,
                    mime="application/xml",
                    use_container_width=True,
                    help=tmx_help
                ):
                    if canonical_id_value and 'canonical_counter' in st.session_state:
                        st.session_state['canonical_counter'] += 1
            
            with col_excel:
                file_canonical_map = {}
                for fname in st.session_state.get('originals', {}).keys():
                    cids = set()
                    file_existing_segments = set()
                    if existing_canonical_map:
                        for (f, seg), cid in existing_canonical_map.items():
                            if f == fname:
                                cids.add(cid)
                                file_existing_segments.add(seg)
                    if canonical_id_value:
                        file_previews = previews.get(fname, [])
                        has_new_segments = any(
                            p['segment'] not in file_existing_segments
                            for p in file_previews
                        )
                        if has_new_segments:
                            cids.add(canonical_id_value)
                    file_canonical_map[fname] = ", ".join(sorted(cids)) if cids else "None"
                
                excel_data = generate_changes_excel(
                    dedup_details=st.session_state.get('dedup_details', []),
                    exclusion_breakdown=st.session_state.get('exclusion_breakdown', None),
                    file_canonical_map=file_canonical_map if file_canonical_map else None
                )
                st.download_button(
                    label="📊 Download changes report",
                    data=excel_data,
                    file_name=f"Report_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("Process the files to download them")


def process_files(files, replacement_token, process_source, process_target,
                  use_safe_regex, use_regex, use_presidio, use_biomedical, use_proper_names,
                  use_dictionary, dictionary_terms, whitelist_terms=None):
    
    anonymizer = MQXLIFFAnonymizer(replacement_token=replacement_token)
    
    results = {}
    originals = {}
    all_stats = {}
    previews = {}
    
    progress_bar = st.progress(0, text="Processing files...")
    
    for i, file in enumerate(files):
        file_label = file.name if len(file.name) <= 40 else file.name[:37] + "..."
        
        def make_progress_cb(file_idx, total_files, fname):
            last_pct = [-1]
            def cb(current, total):
                pct = int((current / total) * 100) if total > 0 else 0
                if pct == last_pct[0] and pct < 100:
                    return
                last_pct[0] = pct
                file_base = file_idx / total_files
                file_share = 1.0 / total_files
                segment_pct = current / total if total > 0 else 0
                overall = file_base + file_share * segment_pct
                progress_bar.progress(min(overall, 1.0), text=f"Processing {fname}... {pct}% ({current}/{total} segments)")
                time.sleep(0.01)
            return cb
        
        progress_cb = make_progress_cb(i, len(files), file_label)
        progress_cb(0, 1)
        
        try:
            content = file.read()
            originals[file.name] = content
            
            is_tmx = file.name.lower().endswith(".tmx")
            
            if is_tmx:
                result_xml, stats, file_previews = anonymizer.anonymize_tmx(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_presidio=use_presidio,
                    use_biomedical=use_biomedical,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            else:
                result_xml, stats, file_previews = anonymizer.anonymize_mqxliff(
                    xml_content=content,
                    process_source=process_source,
                    process_target=process_target,
                    use_safe_regex=use_safe_regex,
                    use_regex=use_regex,
                    use_presidio=use_presidio,
                    use_biomedical=use_biomedical,
                    use_proper_names=use_proper_names,
                    use_dictionary=use_dictionary,
                    dictionary_terms=dictionary_terms,
                    whitelist_terms=whitelist_terms,
                    progress_callback=progress_cb
                )
            
            results[file.name] = result_xml
            all_stats[file.name] = stats
            previews[file.name] = file_previews
            
        except Exception as e:
            st.error(f"Error processing {file.name}: {str(e)}")
    
    progress_bar.empty()
    
    st.session_state.results = results
    st.session_state.originals = originals
    st.session_state.all_stats = all_stats
    st.session_state.previews = previews
    st.session_state['use_canonical'] = True
    st.session_state['dedup_keep'] = {}
    st.session_state['dedup_group_keys'] = {}
    st.session_state['conflict_keep'] = {}
    st.session_state['conflict_group_keys'] = {}
    
    st.success(f"✅ Successfully processed {len(results)} file(s)")
    st.info("Go to the **Preview** and **Download** tabs to see the results")


if __name__ == "__main__":
    main()
